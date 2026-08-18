import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction
from django.db import models
from django.db.models import Case, F, Sum, Value, When
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

_HEADER_FONT = Font(bold=True, color='FFFFFF')
_HEADER_FILL = PatternFill('solid', fgColor='0284C7')

from ..api.serializers import (
    InventoryBatchSerializer,
    InventoryItemSerializer,
    ItemSyncSerializer,
    WarehouseSerializer,
)
from ..models import AppUser, AuditLog, InventoryBatch, InventoryItem, PencacahanRecord, StockOutLog, Warehouse


def _actor(request):
    return request.user if isinstance(request.user, AppUser) else None


# ── Items ──────────────────────────────────────────────────────────────────

class InventoryItemListCreateView(APIView):
    def get(self, request):
        qs = InventoryItem.objects.annotate(
            total_stock=Coalesce(Sum('batches__quantity_remaining'), Value(Decimal('0')), output_field=models.DecimalField())
        )
        # Inventory admin pages pass stock_only=1 to hide service mirror items.
        # The POS omits this flag so package mirror items are scannable by code.
        if request.GET.get('stock_only') == '1':
            qs = qs.filter(is_service=False)
        if request.GET.get('active_only') == '1':
            qs = qs.filter(is_active=True)
        if search := request.GET.get('search', '').strip():
            qs = qs.filter(
                models.Q(code__icontains=search) | models.Q(name__icontains=search)
            )
        return Response(InventoryItemSerializer(qs, many=True).data)

    def post(self, request):
        serializer = InventoryItemSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        instance = serializer.save(created_by=_actor(request))
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='CREATE',
            entity_type='InventoryItem',
            entity_id=str(instance.id),
            description=f'Inventory item created: [{instance.code}] {instance.name}',
        )
        return Response(InventoryItemSerializer(instance).data, status=status.HTTP_201_CREATED)


class InventoryItemDetailView(APIView):
    def _get(self, pk):
        try:
            return InventoryItem.objects.get(pk=pk)
        except InventoryItem.DoesNotExist:
            return None

    def get(self, request, pk):
        instance = self._get(pk)
        if instance is None:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(InventoryItemSerializer(instance).data)

    def put(self, request, pk):
        instance = self._get(pk)
        if instance is None:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if instance.is_service:
            return Response({'error': 'Service items are managed through Treatments.'}, status=status.HTTP_400_BAD_REQUEST)
        serializer = InventoryItemSerializer(instance, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='InventoryItem',
            entity_id=str(pk),
            description=f'Inventory item updated: [{instance.code}] {instance.name}',
        )
        return Response(serializer.data)

    def delete(self, request, pk):
        instance = self._get(pk)
        if instance is None:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if instance.is_service:
            return Response({'error': 'Service items are managed through Treatments.'}, status=status.HTTP_400_BAD_REQUEST)
        instance.is_active = False
        instance.save(update_fields=['is_active'])
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='InventoryItem',
            entity_id=str(pk),
            description=f'Inventory item deactivated: [{instance.code}] {instance.name}',
        )
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Warehouses ─────────────────────────────────────────────────────────────

class WarehouseListCreateView(APIView):
    def get(self, request):
        return Response(WarehouseSerializer(Warehouse.objects.all(), many=True).data)

    def post(self, request):
        serializer = WarehouseSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='CREATE',
            entity_type='Warehouse',
            entity_id=str(instance.id),
            description=f'Warehouse created: [{instance.code}] {instance.name}',
        )
        return Response(WarehouseSerializer(instance).data, status=status.HTTP_201_CREATED)


class WarehouseDetailView(APIView):
    def _get(self, pk):
        try:
            return Warehouse.objects.get(pk=pk)
        except Warehouse.DoesNotExist:
            return None

    def put(self, request, pk):
        instance = self._get(pk)
        if instance is None:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = WarehouseSerializer(instance, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='Warehouse',
            entity_id=str(pk),
            description=f'Warehouse updated: {instance.name}',
        )
        return Response(serializer.data)

    def delete(self, request, pk):
        instance = self._get(pk)
        if instance is None:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if InventoryBatch.objects.filter(warehouse_id=pk).exists():
            return Response(
                {'error': 'Cannot delete warehouse with existing inventory records.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='DELETE',
            entity_type='Warehouse',
            entity_id=str(pk),
            description=f'Warehouse deleted: {instance.name}',
        )
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Stock levels ───────────────────────────────────────────────────────────

def _batch_value_expr():
    """The remaining value of a batch, apportioned from its total purchase cost.

    ``(value / quantity_initial) * quantity_remaining`` is the same
    original-cost-per-unit basis ``production_page.unit_cost_small`` uses when
    it walks batches in Python — this is the DB-side form of the identical
    maths, so both a per-item weighted-average lookup and a bulk aggregate can
    share one definition of "what is this stock worth". Guarded against
    ``quantity_initial=0`` so a malformed batch cannot raise DivisionByZero
    inside an aggregate query.
    """
    return Case(
        When(quantity_initial__gt=0,
             then=F('value') / F('quantity_initial') * F('quantity_remaining')),
        default=Value(Decimal('0')),
        output_field=models.DecimalField(max_digits=20, decimal_places=6),
    )


def stock_valuation_by_item(item_ids=None, warehouse_id=None):
    """{item_id: {'qty_on_hand': Decimal, 'stock_value': Decimal}} across
    remaining batches, optionally scoped to a warehouse or a set of items.

    The one place bulk stock valuation happens. ``StockLevelView`` below and
    the stock-movement report (``views/stock_movement_report.py``) both call
    this instead of each summing ``InventoryBatch`` rows their own way — two
    different stock valuations in one system is a bug waiting to be argued
    about (see docs/stock-movement-patient-activity-design.md §2).
    """
    qs = InventoryBatch.objects.filter(quantity_remaining__gt=0)
    if item_ids is not None:
        qs = qs.filter(item_id__in=item_ids)
    if warehouse_id:
        qs = qs.filter(warehouse_id=warehouse_id)
    rows = qs.values('item_id').annotate(
        qty_on_hand=Sum('quantity_remaining'),
        stock_value=Sum(_batch_value_expr()),
    )
    return {
        r['item_id']: {
            'qty_on_hand': r['qty_on_hand'] or Decimal('0'),
            'stock_value': r['stock_value'] or Decimal('0'),
        }
        for r in rows
    }


class StockLevelView(APIView):
    def get(self, request):
        qs = (
            InventoryBatch.objects
            .values(
                'item_id', 'warehouse_id',
                'item__code', 'item__name', 'item__unit_small',
                'item__unit_medium', 'item__unit_medium_qty',
                'item__unit_large', 'item__unit_large_qty',
                'item__min_stock', 'item__selling_price', 'item__is_active',
                'warehouse__code', 'warehouse__name',
            )
            .annotate(
                total_quantity=Sum('quantity_remaining'),
                stock_value=Sum(_batch_value_expr()),
            )
            .order_by('item__name', 'warehouse__name')
        )
        result = [
            {
                'item_id': r['item_id'],
                'item_code': r['item__code'],
                'item_name': r['item__name'],
                'unit_small': r['item__unit_small'],
                'unit_medium': r['item__unit_medium'] or '',
                'unit_medium_qty': r['item__unit_medium_qty'],
                'unit_large': r['item__unit_large'] or '',
                'unit_large_qty': r['item__unit_large_qty'],
                'min_stock': r['item__min_stock'],
                'selling_price': str(r['item__selling_price']),
                'is_active': r['item__is_active'],
                'warehouse_id': r['warehouse_id'],
                'warehouse_code': r['warehouse__code'],
                'warehouse_name': r['warehouse__name'],
                'total_quantity': r['total_quantity'] or 0,
                'stock_value': str(r['stock_value'] or Decimal('0')),
            }
            for r in qs
        ]
        return Response(result)


# ── Batches (stock-in history) ─────────────────────────────────────────────

class InventoryBatchListView(APIView):
    def get(self, request):
        qs = InventoryBatch.objects.select_related('item', 'warehouse', 'created_by')
        if item_id := request.GET.get('item_id'):
            qs = qs.filter(item_id=item_id)
        if wh_id := request.GET.get('warehouse_id'):
            qs = qs.filter(warehouse_id=wh_id)
        return Response(InventoryBatchSerializer(qs.order_by('-input_date', '-created_at'), many=True).data)


# ── Stock In ───────────────────────────────────────────────────────────────

class StockInView(APIView):
    def post(self, request):
        data = request.data

        # Bulk import path: { entries: [...] }
        if 'entries' in data:
            return self._handle_bulk(request, data['entries'])

        # Single-entry path
        return self._handle_single(request, data)

    def _handle_single(self, request, data):
        try:
            item = InventoryItem.objects.get(pk=data['item_id'])
            warehouse = Warehouse.objects.get(pk=data['warehouse_id'])
        except (InventoryItem.DoesNotExist, Warehouse.DoesNotExist, KeyError):
            return Response({'error': 'Invalid item or warehouse.'}, status=status.HTTP_400_BAD_REQUEST)
        if item.is_service:
            return Response({'error': 'Service items have no physical stock.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            qty_raw = Decimal(str(data['quantity']))
            unit = data.get('unit', 'small')
            qty_small = _to_small(item, qty_raw, unit)
            value = Decimal(str(data['value']))
            input_date = data['input_date']
        except (KeyError, ValueError) as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        if qty_small <= 0:
            return Response({'error': 'Quantity must be positive.'}, status=status.HTTP_400_BAD_REQUEST)

        batch = InventoryBatch.objects.create(
            item=item,
            warehouse=warehouse,
            input_date=input_date,
            quantity_initial=qty_small,
            quantity_remaining=qty_small,
            value=value,
            created_by=_actor(request),
        )
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='CREATE',
            entity_type='InventoryBatch',
            entity_id=str(batch.id),
            description=(
                f'Stock in: {qty_small} {item.unit_small} of '
                f'[{item.code}] {item.name} @ {warehouse.name}'
            ),
        )
        return Response(InventoryBatchSerializer(batch).data, status=status.HTTP_201_CREATED)

    def _handle_bulk(self, request, entries):
        actor = _actor(request)
        created_batches = []
        skipped = []

        for i, entry in enumerate(entries):
            try:
                item = InventoryItem.objects.get(pk=entry['item_id'])
                warehouse = Warehouse.objects.get(pk=entry['warehouse_id'])
            except (InventoryItem.DoesNotExist, Warehouse.DoesNotExist, KeyError):
                skipped.append({'index': i, 'reason': 'Invalid item or warehouse'})
                continue

            if item.is_service:
                skipped.append({'index': i, 'reason': f'Item "{item.code}" is a service item'})
                continue

            try:
                qty_raw = Decimal(str(entry['quantity']))
                unit = entry.get('unit', 'small')
                qty_small = _to_small(item, qty_raw, unit)
                value = Decimal(str(entry['value']))
                input_date = entry['input_date']
            except (KeyError, ValueError) as exc:
                skipped.append({'index': i, 'reason': str(exc)})
                continue

            if qty_small <= 0:
                skipped.append({'index': i, 'reason': 'Quantity must be positive'})
                continue

            batch = InventoryBatch.objects.create(
                item=item,
                warehouse=warehouse,
                input_date=input_date,
                quantity_initial=qty_small,
                quantity_remaining=qty_small,
                value=value,
                created_by=actor,
            )
            AuditLog.objects.create(
                performed_by=actor,
                action='CREATE',
                entity_type='InventoryBatch',
                entity_id=str(batch.id),
                description=(
                    f'Stock in (bulk import): {qty_small} {item.unit_small} of '
                    f'[{item.code}] {item.name} @ {warehouse.name}'
                ),
            )
            created_batches.append(batch)

        if not created_batches and skipped:
            return Response(
                {'error': 'All entries were invalid. Nothing was imported.', 'skipped': skipped},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response({
            'imported': len(created_batches),
            'skipped': len(skipped),
            'skipped_details': skipped,
            'batches': InventoryBatchSerializer(created_batches, many=True).data,
        }, status=status.HTTP_201_CREATED)


# ── Stock Out (FIFO) ───────────────────────────────────────────────────────

VALID_STOCK_OUT_REASONS = {code for code, _label in StockOutLog.REASON_CHOICES}


def _resolve_stock_out_reason(raw):
    """Validate the caller's reason code. Returns (reason, error_message).

    A stock issue with no reason cannot be journalled — the reason is the only
    thing that decides which account absorbs the cost — so an unknown or missing
    value is rejected rather than silently defaulting. Callers written before
    Phase 5 send nothing at all; they get ``REASON_OTHER``, which charges
    5000900 and is the safest place for an unexplained write-off to land.
    """
    if raw in (None, ''):
        return StockOutLog.REASON_OTHER, None
    reason = str(raw).strip()
    if reason not in VALID_STOCK_OUT_REASONS:
        allowed = ', '.join(sorted(VALID_STOCK_OUT_REASONS))
        return None, f'Unknown reason "{reason}". Expected one of: {allowed}.'
    return reason, None


def _stock_out_posting_status(reason, value):
    """A row with no journal to write is born 'posted'.

    A warehouse transfer has no P&L effect and a zero-cost draw has nothing to
    charge, so neither will ever produce legs. Marking them posted up front
    keeps them out of ``_gather_events`` instead of having every future journal
    preview pick them up and stage nothing.
    """
    journalable = bool(StockOutLog.REASON_ACCOUNTS.get(reason)) and value > 0
    return 'unposted' if journalable else 'posted'


class StockOutView(APIView):
    def post(self, request):
        data = request.data

        # Bulk import path: { entries: [...] }
        if 'entries' in data:
            return self._handle_bulk(request, data['entries'])

        return self._handle_single(request, data)

    def _handle_single(self, request, data):
        try:
            item = InventoryItem.objects.get(pk=data['item_id'])
            warehouse = Warehouse.objects.get(pk=data['warehouse_id'])
            qty_raw = Decimal(str(data['quantity']))
            unit = data.get('unit', 'small')
            qty_small = _to_small(item, qty_raw, unit)
        except (InventoryItem.DoesNotExist, Warehouse.DoesNotExist, KeyError, ValueError) as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        reason, reason_error = _resolve_stock_out_reason(data.get('reason'))
        if reason_error:
            return Response({'error': reason_error}, status=status.HTTP_400_BAD_REQUEST)

        if item.is_service:
            return Response({'error': 'Service items have no physical stock.'}, status=status.HTTP_400_BAD_REQUEST)

        if qty_small <= 0:
            return Response({'error': 'Quantity must be positive.'}, status=status.HTTP_400_BAD_REQUEST)

        available = (
            InventoryBatch.objects
            .filter(item=item, warehouse=warehouse, quantity_remaining__gt=0)
            .aggregate(total=Sum('quantity_remaining'))['total'] or 0
        )
        if available < qty_small:
            return Response(
                {
                    'error': (
                        f'Insufficient stock. Available: {available} {item.unit_small}, '
                        f'requested: {qty_small} {item.unit_small}.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # The FIFO cost is captured here and nowhere else. Once these batches are
        # drawn down the number cannot be recomputed, and it is the entire
        # amount of the journal entry the sweep will write.
        shortfall, cogs = _fifo_deduct(item.id, warehouse.id, qty_small)
        if shortfall > 0:
            return Response(
                {'error': 'Stock deduction failed due to a concurrent modification.'},
                status=status.HTTP_409_CONFLICT,
            )

        import datetime
        notes = data.get('notes', '')
        out_date = data.get('out_date') or str(datetime.date.today())
        actor = _actor(request)
        AuditLog.objects.create(
            performed_by=actor,
            action='CREATE',
            entity_type='StockOut',
            entity_id=str(item.id),
            description=(
                f'Stock out: {qty_small} {item.unit_small} of '
                f'[{item.code}] {item.name} @ {warehouse.name}'
                + (f'. Notes: {notes}' if notes else '')
            ),
        )
        log = StockOutLog.objects.create(
            item=item,
            warehouse=warehouse,
            out_date=out_date,
            quantity=qty_small,
            reason=reason,
            value=cogs,
            notes=notes,
            created_by=actor,
            posting_status=_stock_out_posting_status(reason, cogs),
        )
        return Response({
            'deducted': qty_small,
            'unit': item.unit_small,
            'reason': reason,
            'value': str(cogs),
            'posting_status': log.posting_status,
        })

    def _handle_bulk(self, request, entries):
        actor = _actor(request)
        deducted_count = 0
        skipped = []

        for i, entry in enumerate(entries):
            try:
                item = InventoryItem.objects.get(pk=entry['item_id'])
                warehouse = Warehouse.objects.get(pk=entry['warehouse_id'])
                qty_raw = Decimal(str(entry['quantity']))
                unit = entry.get('unit', 'small')
                qty_small = _to_small(item, qty_raw, unit)
            except (InventoryItem.DoesNotExist, Warehouse.DoesNotExist, KeyError, ValueError) as exc:
                skipped.append({'index': i, 'reason': str(exc)})
                continue

            out_reason, reason_error = _resolve_stock_out_reason(entry.get('reason'))
            if reason_error:
                skipped.append({'index': i, 'reason': reason_error})
                continue

            if item.is_service:
                skipped.append({'index': i, 'reason': f'Item "{item.code}" is a service item'})
                continue

            if qty_small <= 0:
                skipped.append({'index': i, 'reason': 'Quantity must be positive'})
                continue

            available = (
                InventoryBatch.objects
                .filter(item=item, warehouse=warehouse, quantity_remaining__gt=0)
                .aggregate(total=Sum('quantity_remaining'))['total'] or 0
            )
            if available < qty_small:
                skipped.append({
                    'index': i,
                    'reason': (
                        f'Insufficient stock for [{item.code}]. '
                        f'Available: {available} {item.unit_small}, requested: {qty_small}.'
                    ),
                })
                continue

            shortfall, cogs = _fifo_deduct(item.id, warehouse.id, qty_small)
            if shortfall > 0:
                skipped.append({'index': i, 'reason': f'Concurrent modification for [{item.code}]'})
                continue

            import datetime
            notes = entry.get('notes', '')
            out_date = entry.get('out_date') or str(datetime.date.today())
            AuditLog.objects.create(
                performed_by=actor,
                action='CREATE',
                entity_type='StockOut',
                entity_id=str(item.id),
                description=(
                    f'Stock out (bulk import): {qty_small} {item.unit_small} of '
                    f'[{item.code}] {item.name} @ {warehouse.name}'
                    + (f'. Notes: {notes}' if notes else '')
                ),
            )
            StockOutLog.objects.create(
                item=item,
                warehouse=warehouse,
                out_date=out_date,
                quantity=qty_small,
                reason=out_reason,
                value=cogs,
                notes=notes,
                created_by=actor,
                posting_status=_stock_out_posting_status(out_reason, cogs),
            )
            deducted_count += 1

        if deducted_count == 0 and skipped:
            return Response(
                {'error': 'All entries were invalid. Nothing was deducted.', 'skipped': skipped},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response({
            'imported': deducted_count,
            'skipped': len(skipped),
            'skipped_details': skipped,
        }, status=status.HTTP_201_CREATED)


# ── Stock-out history ──────────────────────────────────────────────────────

class StockOutBatchListView(APIView):
    def get(self, request):
        qs = StockOutLog.objects.select_related('item', 'warehouse', 'created_by').order_by('-created_at')
        if item_id := request.GET.get('item_id'):
            qs = qs.filter(item_id=item_id)
        if warehouse_id := request.GET.get('warehouse_id'):
            qs = qs.filter(warehouse_id=warehouse_id)
        data = [
            {
                'id': log.id,
                'item_id': log.item_id,
                'item_code': log.item.code,
                'item_name': log.item.name,
                'item_unit_small': log.item.unit_small,
                'warehouse_id': log.warehouse_id,
                'warehouse_name': log.warehouse.name,
                'out_date': str(log.out_date),
                'quantity': log.quantity,
                'reason': log.reason,
                'reason_label': log.get_reason_display(),
                'value': str(log.value),
                'posting_status': log.posting_status,
                'notes': log.notes,
                'created_by_name': log.created_by.display_name if log.created_by else None,
                'created_at': log.created_at.isoformat(),
            }
            for log in qs
        ]
        return Response(data)


# ── Cashier sync ───────────────────────────────────────────────────────────

class ItemSyncView(APIView):
    """
    GET /api/inventory/sync/items/?since=<ISO-8601>&page=<n>&page_size=<n>

    Returns inventory items updated since the given timestamp, paginated.
    Omit `since` for a full sync (first-time setup).
    Default page_size=200, max=1000.
    """
    def get(self, request):
        from django.utils.dateparse import parse_datetime
        synced_at = timezone.now()
        since_raw = request.GET.get('since')

        try:
            page = max(1, int(request.GET.get('page', 1)))
            page_size = min(max(1, int(request.GET.get('page_size', 200))), 1000)
        except (ValueError, TypeError):
            return Response(
                {'error': 'Invalid `page` or `page_size` value.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        qs = InventoryItem.objects.all().order_by('updated_at', 'id')

        if since_raw:
            try:
                since_dt = parse_datetime(since_raw)
                if since_dt is None:
                    return Response(
                        {'error': 'Invalid `since` value. Use ISO 8601 format, e.g. 2025-01-01T00:00:00Z.'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
                if timezone.is_naive(since_dt):
                    since_dt = timezone.make_aware(since_dt)
                qs = qs.filter(updated_at__gte=since_dt)
            except (ValueError, TypeError):
                return Response(
                    {'error': 'Invalid `since` value. Use ISO 8601 format, e.g. 2025-01-01T00:00:00Z.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        total_count = qs.count()
        start = (page - 1) * page_size
        end = start + page_size
        items = ItemSyncSerializer(qs[start:end], many=True).data
        has_more = end < total_count

        return Response({
            'synced_at': synced_at,
            'count': len(items),
            'total_count': total_count,
            'has_more': has_more,
            'next_page': page + 1 if has_more else None,
            'items': items,
        })


# ── Helpers ────────────────────────────────────────────────────────────────

def _to_small(item: InventoryItem, qty: Decimal, unit: str) -> Decimal:
    if unit == 'large':
        if not item.unit_large or not item.unit_large_qty:
            raise ValueError('This item has no large unit defined.')
        if not item.unit_medium or not item.unit_medium_qty:
            raise ValueError('Large unit requires a medium unit to be defined.')
        return qty * Decimal(item.unit_large_qty) * Decimal(item.unit_medium_qty)
    if unit == 'medium':
        if not item.unit_medium or not item.unit_medium_qty:
            raise ValueError('This item has no medium unit defined.')
        return qty * Decimal(item.unit_medium_qty)
    return qty


class FifoSimulation:
    """Virtual consumption ledger for dry-run FIFO.

    A journal *preview* has to know what COGS an invoice would produce without
    consuming any stock. Reading the live batch rows once per line is not enough:
    a preview covers many invoices, and two of them may draw on the same batch.
    Without a memory of what earlier previewed lines "used", the second invoice
    would see stock the first one already claimed and both would report a full
    COGS figure.

    So dry runs share one of these for the whole preview. It records how much of
    each batch has been virtually consumed and subtracts that from the batch's
    real ``quantity_remaining``. Nothing here is written to the database.
    """

    __slots__ = ('_consumed',)

    def __init__(self):
        self._consumed = {}

    def available(self, batch) -> Decimal:
        used = self._consumed.get(batch.pk, Decimal('0'))
        left = Decimal(batch.quantity_remaining) - used
        return left if left > 0 else Decimal('0')

    def consume(self, batch, quantity: Decimal) -> None:
        self._consumed[batch.pk] = self._consumed.get(batch.pk, Decimal('0')) + quantity


def _fifo_batches(item_id: int, warehouse_id=None, *, lock: bool):
    """The FIFO queue for an item, oldest batch first.

    ``lock`` is False only for dry runs. The real deduction locks the rows for
    the life of the enclosing transaction, which is correct there; a preview
    that took the same locks would hold them across hundreds of documents and
    stall the POS. A dry run therefore reads unlocked and accepts that its
    answer is a snapshot — which is why previewed COGS is recomputed at commit.
    """
    criteria = {'item_id': item_id, 'quantity_remaining__gt': 0}
    if warehouse_id is not None:
        criteria['warehouse_id'] = warehouse_id
    # Kept as a single .filter() on the manager so the chain is byte-for-byte
    # what it was before the dry-run split — test_fifo.py mocks this exact call
    # sequence, and a queryset that behaves the same but is *built* differently
    # would break it for no reason.
    manager = InventoryBatch.objects.select_for_update() if lock else InventoryBatch.objects
    return manager.filter(**criteria).order_by('input_date', 'created_at')


def _fifo_walk(batches, quantity: Decimal, *, commit: bool, sim: 'FifoSimulation | None' = None) -> tuple:
    """Walk the FIFO queue consuming ``quantity``. Returns (shortfall, cogs).

    The single implementation behind both the real deduction and its dry run —
    if these two ever diverged, the preview would confidently show a COGS figure
    the commit would never post.
    """
    remaining = Decimal(quantity)
    cogs = Decimal('0')
    for batch in batches:
        if remaining <= 0:
            break
        on_hand = sim.available(batch) if sim is not None else Decimal(batch.quantity_remaining)
        if on_hand <= 0:
            continue
        deduct = min(on_hand, remaining)
        if batch.quantity_initial:
            cogs += (batch.value / batch.quantity_initial) * deduct
        if commit:
            batch.quantity_remaining -= deduct
            batch.save(update_fields=['quantity_remaining'])
        elif sim is not None:
            sim.consume(batch, deduct)
        remaining -= deduct
    return remaining, cogs


def _fifo_deduct(item_id: int, warehouse_id: int, quantity: Decimal,
                 *, commit: bool = True, sim: 'FifoSimulation | None' = None) -> tuple:
    """Deduct stock FIFO. Returns (shortfall, cogs_amount). shortfall=0 means fully deducted.

    ``commit=False`` computes the same answer without touching a single row —
    used by the journal preview. Pass the same ``sim`` for every dry-run call in
    one preview so repeated draws on a batch stay honest.
    """
    if not commit:
        return _fifo_walk(_fifo_batches(item_id, warehouse_id, lock=False),
                          quantity, commit=False, sim=sim)
    with transaction.atomic():
        return _fifo_walk(_fifo_batches(item_id, warehouse_id, lock=True),
                          quantity, commit=True)


def _fifo_deduct_global(item_id: int, quantity_small: Decimal,
                        *, commit: bool = True, sim: 'FifoSimulation | None' = None) -> tuple:
    """
    Deduct a decimal quantity (in unit_small) from any warehouse, FIFO by date.
    Returns (shortfall, cogs_amount). See ``_fifo_deduct`` for ``commit``/``sim``.
    """
    qty = Decimal(quantity_small)
    if qty <= 0:
        return Decimal('0'), Decimal('0')
    return _fifo_walk(_fifo_batches(item_id, None, lock=commit),
                      qty, commit=commit, sim=None if commit else sim)


@transaction.atomic
def _fifo_restock(item_id: int, warehouse_id: int, quantity: Decimal) -> Decimal:
    """Return stock to batches (newest-first). Returns the COGS amount reversed."""
    batches = (
        InventoryBatch.objects
        .select_for_update()
        .filter(item_id=item_id, warehouse_id=warehouse_id)
        .order_by('-input_date', '-created_at')
    )
    return _restock_batches(batches, quantity)


@transaction.atomic
def _fifo_restock_global(item_id: int, quantity: Decimal) -> Decimal:
    """
    Return a decimal quantity (in unit_small) to any warehouse, newest-first.
    Mirror of _fifo_deduct_global. Returns the COGS amount reversed.
    """
    batches = (
        InventoryBatch.objects
        .select_for_update()
        .filter(item_id=item_id)
        .order_by('-input_date', '-created_at')
    )
    return _restock_batches(batches, quantity)


def _restock_batches(batches, quantity: Decimal) -> Decimal:
    """
    Refill batches newest-first, capped at each batch's original quantity.

    Note: deduction is oldest-first while restock is newest-first, so a
    reverse-then-repost of the same line can shift COGS between batches when
    an item has batches at differing unit costs. Exact reversal would require
    recording which batches each line consumed.
    """
    remaining = Decimal(quantity)
    cogs = Decimal('0')
    if remaining <= 0:
        return cogs
    for batch in batches:
        if remaining <= 0:
            break
        capacity = (batch.quantity_initial or Decimal('0')) - batch.quantity_remaining
        if capacity <= 0:
            continue
        restock = min(capacity, remaining)
        if batch.quantity_initial:
            cogs += (batch.value / batch.quantity_initial) * restock
        batch.quantity_remaining += restock
        batch.save(update_fields=['quantity_remaining'])
        remaining -= restock
    return cogs


# ── Excel template + import ───────────────────────────────────────────────────

class InventoryItemTemplateView(APIView):
    """GET — download a blank .xlsx template for bulk item import."""

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Items'

        headers = [
            'code', 'name', 'selling_price',
            'unit_small', 'unit_medium', 'unit_medium_qty',
            'unit_large', 'unit_large_qty',
            'min_stock', 'is_active',
            'category', 'legal_code',
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = Alignment(horizontal='center')

        # Notes row explaining each column
        notes = [
            'Unique item code (required)',
            'Item name (required)',
            'Selling price (required, e.g. 85000)',
            'Smallest unit, e.g. ml or pcs (required)',
            'Medium unit, e.g. bottle (optional)',
            'Small units per 1 medium unit (optional)',
            'Large unit, e.g. box (optional)',
            'Medium units per 1 large unit (optional)',
            'Minimum stock level (optional, default 0)',
            'yes / no (optional, default yes)',
            'Item category (optional)',
            'Legal/regulatory code (optional)',
        ]
        ws.append(notes)
        note_font = Font(italic=True, color='595959')
        for cell in ws[2]:
            cell.font      = note_font
            cell.alignment = Alignment(wrap_text=True)

        # Example data rows
        examples = [
            ('ITEM-001', 'Vitamin C Serum',  85000,  'ml', 'bottle', 30,  '',    '',  10, 'yes', 'Skincare', ''),
            ('ITEM-002', 'Facial Toner',     45000,  'ml', 'bottle', 200, 'box', 6,   5,  'yes', 'Skincare', 'REG-001'),
            ('ITEM-003', 'Sunscreen SPF50',  120000, 'g',  '',       '',  '',    '',  0,  'yes', 'Suncare',  ''),
        ]
        for row in examples:
            ws.append(list(row))

        # Column widths
        col_widths = [14, 28, 16, 12, 14, 18, 12, 16, 12, 12, 18, 18]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        ws.row_dimensions[2].height = 36

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="inventory_items_template.xlsx"'
        return response


def _parse_import_rows(rows):
    """
    Parse spreadsheet rows (skipping header row 1) into validated item dicts.
    Returns (valid_rows, errors).
    valid_rows: list of dicts with keys matching InventoryItem fields plus 'row' and 'action'.
    errors: list of {row, message}.
    """
    def _int_or_none(v):
        if v is None or str(v).strip() == '':
            return None
        try:
            return int(v)
        except (ValueError, TypeError):
            return None

    existing_codes = set(InventoryItem.objects.values_list('code', flat=True))
    valid_rows = []
    errors = []

    for i, row in enumerate(rows[1:], start=2):
        padded = list(row) + [None] * 12
        code           = str(padded[0]).strip() if padded[0] is not None else ''
        name           = str(padded[1]).strip() if padded[1] is not None else ''
        price_raw      = padded[2]
        unit_small     = str(padded[3]).strip() if padded[3] is not None else ''
        unit_medium    = str(padded[4]).strip() if padded[4] is not None else ''
        unit_medium_qty_raw = padded[5]
        unit_large     = str(padded[6]).strip() if padded[6] is not None else ''
        unit_large_qty_raw  = padded[7]
        min_stock_raw  = padded[8]
        active_raw     = padded[9]
        category       = str(padded[10]).strip() if padded[10] is not None else ''
        legal_code     = str(padded[11]).strip() if padded[11] is not None else None
        if legal_code == '':
            legal_code = None

        if not code:
            errors.append({'row': i, 'message': 'Code is required.'})
            continue
        if not name:
            errors.append({'row': i, 'message': f'Row {i}: name is required.'})
            continue
        if not unit_small:
            errors.append({'row': i, 'message': f'Row {i}: unit_small is required.'})
            continue
        try:
            selling_price = Decimal(str(price_raw)) if price_raw not in (None, '') else None
            if selling_price is None:
                raise ValueError
        except (ValueError, TypeError):
            errors.append({'row': i, 'message': f'Row {i}: invalid selling_price "{price_raw}".'})
            continue

        unit_medium_qty = _int_or_none(unit_medium_qty_raw)
        unit_large_qty  = _int_or_none(unit_large_qty_raw)
        min_stock       = _int_or_none(min_stock_raw) or 0

        if active_raw is None or str(active_raw).strip() == '':
            is_active = True
        else:
            is_active = str(active_raw).strip().lower() in ('yes', 'true', '1')

        valid_rows.append({
            'row': i,
            'action': 'update' if code in existing_codes else 'create',
            'code': code,
            'name': name,
            'selling_price': selling_price,
            'unit_small': unit_small,
            'unit_medium': unit_medium,
            'unit_medium_qty': unit_medium_qty,
            'unit_large': unit_large,
            'unit_large_qty': unit_large_qty,
            'min_stock': min_stock,
            'is_active': is_active,
            'category': category,
            'legal_code': legal_code,
        })

    return valid_rows, errors


def _load_workbook_from_request(request):
    uploaded = request.FILES.get('file')
    if not uploaded:
        return None, None, Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), read_only=True, data_only=True)
    except Exception:
        return None, None, Response({'error': 'Could not parse file. Upload a valid .xlsx file.'}, status=status.HTTP_400_BAD_REQUEST)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, None, Response({'error': 'File is empty.'}, status=status.HTTP_400_BAD_REQUEST)
    return wb, rows, None


class InventoryItemImportPreviewView(APIView):
    """
    POST multipart/form-data with file=<xlsx>
    Parses and validates the file without writing to the database.
    Returns { rows: [{row, action, code, name, ...}], errors: [{row, message}] }
    where action is 'create' or 'update'.
    """

    def post(self, request):
        _wb, rows, err = _load_workbook_from_request(request)
        if err:
            return err
        valid_rows, errors = _parse_import_rows(rows)
        preview_rows = [
            {
                'row': r['row'],
                'action': r['action'],
                'code': r['code'],
                'name': r['name'],
                'selling_price': str(r['selling_price']),
                'unit_small': r['unit_small'],
                'unit_medium': r['unit_medium'],
                'category': r['category'],
                'legal_code': r['legal_code'],
                'is_active': r['is_active'],
            }
            for r in valid_rows
        ]
        return Response({'rows': preview_rows, 'errors': errors}, status=status.HTTP_200_OK)


class InventoryItemImportView(APIView):
    """
    POST multipart/form-data with file=<xlsx>
    Columns (row 1 = header, skipped):
      code | name | selling_price | unit_small | unit_medium | unit_medium_qty | unit_large | unit_large_qty | min_stock | is_active | category | legal_code
    Parses, validates, then commits all valid rows.
    Returns { created, updated, errors: [{row, message}] }
    """

    def post(self, request):
        _wb, rows, err = _load_workbook_from_request(request)
        if err:
            return err
        valid_rows, errors = _parse_import_rows(rows)
        created = updated = 0

        for r in valid_rows:
            try:
                with transaction.atomic():
                    obj, was_created = InventoryItem.objects.update_or_create(
                        code=r['code'],
                        defaults={
                            'name': r['name'],
                            'selling_price': r['selling_price'],
                            'unit_small': r['unit_small'],
                            'unit_medium': r['unit_medium'],
                            'unit_medium_qty': r['unit_medium_qty'],
                            'unit_large': r['unit_large'],
                            'unit_large_qty': r['unit_large_qty'],
                            'min_stock': r['min_stock'],
                            'is_active': r['is_active'],
                            'category': r['category'],
                            'legal_code': r['legal_code'],
                        },
                    )
                    AuditLog.objects.create(
                        performed_by=_actor(request),
                        action='CREATE' if was_created else 'UPDATE',
                        entity_type='InventoryItem',
                        entity_id=str(obj.id),
                        description=f'Item {"imported" if was_created else "updated via import"}: {obj.name} ({obj.code})',
                    )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except Exception as exc:
                errors.append({'row': r['row'], 'message': f'Row {r["row"]}: {exc}'})

        return Response({'created': created, 'updated': updated, 'errors': errors}, status=status.HTTP_200_OK)


# ── Pencacahan (Unit Conversion) ───────────────────────────────────────────

class PencacahanListCreateView(APIView):
    """
    GET  /api/inventory/pencacahan/   — list all pencacahan records, newest first
    POST /api/inventory/pencacahan/   — execute a unit conversion

    POST body:
    {
        "date": "YYYY-MM-DD",
        "source_item_id": <int>,
        "source_warehouse_id": <int>,
        "source_quantity": <number>,      // in source item's unit_small
        "source_unit": "small"|"medium"|"large",
        "target_item_id": <int>,
        "target_warehouse_id": <int>,
        "target_quantity": <number>,      // in target item's unit_small
        "target_unit": "small"|"medium"|"large",
        "notes": ""
    }

    The value (cost) transferred equals the FIFO cost of the source stock
    consumed. The same total value is added as a new InventoryBatch for the
    target item, so inventory value is unchanged.
    """

    def get(self, request):
        qs = PencacahanRecord.objects.select_related(
            'source_item', 'source_warehouse',
            'target_item', 'target_warehouse',
            'created_by',
        )
        if item_id := request.GET.get('item_id'):
            qs = qs.filter(
                models.Q(source_item_id=item_id) | models.Q(target_item_id=item_id)
            )
        records = [
            {
                'id': r.id,
                'pencacahan_no': r.pencacahan_no,
                'date': str(r.date),
                'source_item_id': r.source_item_id,
                'source_item_code': r.source_item.code,
                'source_item_name': r.source_item.name,
                'source_item_unit': r.source_item.unit_small,
                'source_warehouse_id': r.source_warehouse_id,
                'source_warehouse_name': r.source_warehouse.name,
                'source_quantity': str(r.source_quantity),
                'target_item_id': r.target_item_id,
                'target_item_code': r.target_item.code,
                'target_item_name': r.target_item.name,
                'target_item_unit': r.target_item.unit_small,
                'target_warehouse_id': r.target_warehouse_id,
                'target_warehouse_name': r.target_warehouse.name,
                'target_quantity': str(r.target_quantity),
                'value_transferred': str(r.value_transferred),
                'notes': r.notes,
                'created_by_name': r.created_by.display_name if r.created_by else None,
                'created_at': r.created_at.isoformat(),
            }
            for r in qs
        ]
        return Response(records)

    @transaction.atomic
    def post(self, request):
        data = request.data
        import datetime

        # ── Resolve items and warehouses ──────────────────────────────────
        try:
            source_item = InventoryItem.objects.get(pk=data['source_item_id'])
            source_wh   = Warehouse.objects.get(pk=data['source_warehouse_id'])
            target_item = InventoryItem.objects.get(pk=data['target_item_id'])
            target_wh   = Warehouse.objects.get(pk=data['target_warehouse_id'])
        except (InventoryItem.DoesNotExist, Warehouse.DoesNotExist, KeyError) as exc:
            return Response({'error': f'Invalid item or warehouse: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

        if source_item.is_service or target_item.is_service:
            return Response({'error': 'Service items cannot be used in pencacahan.'}, status=status.HTTP_400_BAD_REQUEST)

        if source_item.id == target_item.id and source_wh.id == target_wh.id:
            return Response({'error': 'Source and target must differ (item or warehouse).'}, status=status.HTTP_400_BAD_REQUEST)

        # ── Parse quantities ──────────────────────────────────────────────
        try:
            src_qty_raw  = Decimal(str(data['source_quantity']))
            src_unit     = data.get('source_unit', 'small')
            src_qty      = _to_small(source_item, src_qty_raw, src_unit)

            tgt_qty_raw  = Decimal(str(data['target_quantity']))
            tgt_unit     = data.get('target_unit', 'small')
            tgt_qty      = _to_small(target_item, tgt_qty_raw, tgt_unit)
        except (KeyError, ValueError) as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        if src_qty <= 0 or tgt_qty <= 0:
            return Response({'error': 'Quantities must be positive.'}, status=status.HTTP_400_BAD_REQUEST)

        # ── Parse date ────────────────────────────────────────────────────
        try:
            conv_date = datetime.date.fromisoformat(data.get('date') or str(datetime.date.today()))
        except ValueError:
            return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)

        notes = data.get('notes', '')

        # ── Check source availability ─────────────────────────────────────
        available = (
            InventoryBatch.objects
            .filter(item=source_item, warehouse=source_wh, quantity_remaining__gt=0)
            .aggregate(total=Sum('quantity_remaining'))['total'] or Decimal('0')
        )
        if available < src_qty:
            return Response(
                {
                    'error': (
                        f'Insufficient stock for [{source_item.code}]. '
                        f'Available: {available} {source_item.unit_small}, '
                        f'requested: {src_qty} {source_item.unit_small}.'
                    )
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── FIFO deduct source ────────────────────────────────────────────
        shortfall, value_transferred = _fifo_deduct(source_item.id, source_wh.id, src_qty)
        if shortfall > 0:
            return Response(
                {'error': 'Stock deduction failed due to a concurrent modification.'},
                status=status.HTTP_409_CONFLICT,
            )

        # ── Add target batch with same value ──────────────────────────────
        target_batch = InventoryBatch.objects.create(
            item=target_item,
            warehouse=target_wh,
            input_date=conv_date,
            quantity_initial=tgt_qty,
            quantity_remaining=tgt_qty,
            value=value_transferred,
            created_by=_actor(request),
        )

        # ── Record pencacahan ─────────────────────────────────────────────
        actor = _actor(request)
        pencacahan_no = PencacahanRecord.next_number(conv_date)
        record = PencacahanRecord.objects.create(
            pencacahan_no=pencacahan_no,
            date=conv_date,
            source_item=source_item,
            source_warehouse=source_wh,
            source_quantity=src_qty,
            target_item=target_item,
            target_warehouse=target_wh,
            target_quantity=tgt_qty,
            value_transferred=value_transferred,
            notes=notes,
            created_by=actor,
        )

        AuditLog.objects.create(
            performed_by=actor,
            action='CREATE',
            entity_type='PencacahanRecord',
            entity_id=str(record.id),
            description=(
                f'Pencacahan {pencacahan_no}: '
                f'{src_qty} {source_item.unit_small} [{source_item.code}] → '
                f'{tgt_qty} {target_item.unit_small} [{target_item.code}] '
                f'| nilai Rp {value_transferred:,.2f}'
            ),
        )

        return Response(
            {
                'pencacahan_no': pencacahan_no,
                'value_transferred': str(value_transferred),
                'target_batch_id': target_batch.id,
                'source_quantity_consumed': str(src_qty),
                'target_quantity_added': str(tgt_qty),
            },
            status=status.HTTP_201_CREATED,
        )


class PencacahanDetailView(APIView):
    def get(self, request, pk):
        try:
            r = PencacahanRecord.objects.select_related(
                'source_item', 'source_warehouse',
                'target_item', 'target_warehouse',
                'created_by',
            ).get(pk=pk)
        except PencacahanRecord.DoesNotExist:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response({
            'id': r.id,
            'pencacahan_no': r.pencacahan_no,
            'date': str(r.date),
            'source_item_id': r.source_item_id,
            'source_item_code': r.source_item.code,
            'source_item_name': r.source_item.name,
            'source_item_unit': r.source_item.unit_small,
            'source_warehouse_id': r.source_warehouse_id,
            'source_warehouse_name': r.source_warehouse.name,
            'source_quantity': str(r.source_quantity),
            'target_item_id': r.target_item_id,
            'target_item_code': r.target_item.code,
            'target_item_name': r.target_item.name,
            'target_item_unit': r.target_item.unit_small,
            'target_warehouse_id': r.target_warehouse_id,
            'target_warehouse_name': r.target_warehouse.name,
            'target_quantity': str(r.target_quantity),
            'value_transferred': str(r.value_transferred),
            'notes': r.notes,
            'created_by_name': r.created_by.display_name if r.created_by else None,
            'created_at': r.created_at.isoformat(),
        })
