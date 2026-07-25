import io
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from managementsys.models import (
    InventoryItem, InventoryBatch, ProductionRecipe, ProductionRecipeIngredient,
    ProductionRun, ProductionRunIngredient, Warehouse, AuditLog,
)
from managementsys.api.serializers import (
    ProductionRecipeSerializer, ProductionRunSerializer,
)
from managementsys.views.inventory_page import (  # reuse, do not copy
    _to_small, _actor, _fifo_deduct, _fifo_deduct_global,
)

_HEADER_FONT = Font(bold=True, color='FFFFFF')
_HEADER_FILL = PatternFill('solid', fgColor='0284C7')


def unit_cost_small(item_id, warehouse_id=None) -> Decimal:
    """Weighted-average cost per smallest unit across remaining batches."""
    qs = InventoryBatch.objects.filter(item_id=item_id, quantity_remaining__gt=0)
    if warehouse_id:
        qs = qs.filter(warehouse_id=warehouse_id)
    total_qty = Decimal('0')
    total_val = Decimal('0')
    for b in qs:
        if b.quantity_initial:
            total_val += (b.value / b.quantity_initial) * b.quantity_remaining
        total_qty += b.quantity_remaining
    if total_qty <= 0:
        return Decimal('0')
    return total_val / total_qty


def available_small(item_id, warehouse_id=None) -> Decimal:
    qs = InventoryBatch.objects.filter(item_id=item_id, quantity_remaining__gt=0)
    if warehouse_id:
        qs = qs.filter(warehouse_id=warehouse_id)
    return qs.aggregate(t=Sum('quantity_remaining'))['t'] or Decimal('0')


def build_cost_breakdown(lines, output_quantity, warehouse_id=None):
    """
    lines: iterable of dicts {item_id, quantity, unit}
    output_quantity: Decimal (smallest unit yield)
    Returns (breakdown_list, total_cost, cost_per_output_unit).
    Each breakdown row: {item_id, item_code, item_name, unit, quantity,
                         quantity_small, unit_cost_small, line_cost,
                         available_small, insufficient}
    Raises ValueError on bad unit/qty (propagate as 400 upstream).
    """
    breakdown = []
    total = Decimal('0')
    for ln in lines:
        item = InventoryItem.objects.get(pk=ln['item_id'])
        qty = Decimal(str(ln['quantity']))
        unit = ln.get('unit', 'small')
        qty_small = _to_small(item, qty, unit)
        ucost = unit_cost_small(item.id, warehouse_id)
        line_cost = (ucost * qty_small).quantize(Decimal('0.01'))
        avail = available_small(item.id, warehouse_id)
        breakdown.append({
            'item_id': item.id,
            'item_code': item.code,
            'item_name': item.name,
            'unit': unit,
            'quantity': str(qty),
            'quantity_small': str(qty_small),
            'unit_cost_small': str(ucost.quantize(Decimal('0.0001'))),
            'line_cost': str(line_cost),
            'available_small': str(avail),
            'insufficient': avail < qty_small,
        })
        total += line_cost
    out_q = Decimal(str(output_quantity or '0'))
    per_unit = (total / out_q).quantize(Decimal('0.0001')) if out_q > 0 else Decimal('0')
    return breakdown, total.quantize(Decimal('0.01')), per_unit


# ── Recipe CRUD ──────────────────────────────────────────────────────────────

class RecipeListCreateView(APIView):
    def get(self, request):
        qs = ProductionRecipe.objects.prefetch_related('ingredients').all()
        if request.GET.get('active') == '1':
            qs = qs.filter(is_active=True)
        return Response(ProductionRecipeSerializer(qs, many=True).data)

    def post(self, request):
        ser = ProductionRecipeSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        recipe = ser.save(created_by=_actor(request))
        return Response(ProductionRecipeSerializer(recipe).data,
                        status=status.HTTP_201_CREATED)


class RecipeDetailView(APIView):
    def _get(self, pk):
        return get_object_or_404(
            ProductionRecipe.objects.prefetch_related('ingredients'), pk=pk)

    def get(self, request, pk):
        return Response(ProductionRecipeSerializer(self._get(pk)).data)

    def put(self, request, pk):
        ser = ProductionRecipeSerializer(self._get(pk), data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)

    def delete(self, request, pk):
        self._get(pk).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class RecipeCostView(APIView):
    """GET live cost breakdown for a saved recipe. ?warehouse_id= optional."""
    def get(self, request, pk):
        recipe = get_object_or_404(
            ProductionRecipe.objects.prefetch_related('ingredients'), pk=pk)
        wh = request.GET.get('warehouse_id') or None
        lines = [{'item_id': i.item_id, 'quantity': i.quantity, 'unit': i.unit}
                 for i in recipe.ingredients.all()]
        try:
            breakdown, total, per_unit = build_cost_breakdown(
                lines, recipe.output_quantity, wh)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({
            'recipe_id': recipe.id,
            'output_quantity': str(recipe.output_quantity),
            'lines': breakdown,
            'total_cost': str(total),
            'cost_per_output_unit': str(per_unit),
        })


# ── Recipe Excel import/export ────────────────────────────────────────────────
# Flat layout: one row per ingredient. Recipe-level columns (name/output/notes/
# active) only need to be filled on the first row of each recipe's block —
# blank cells on later rows inherit the most recent non-blank value.

_RECIPE_HEADERS = [
    'recipe_name', 'output_item_code', 'output_quantity', 'notes', 'is_active',
    'ingredient_item_code', 'ingredient_quantity', 'ingredient_unit',
]
_RECIPE_COL_WIDTHS = [22, 18, 16, 30, 10, 20, 20, 16]


def _style_recipe_header(ws):
    ws.append(_RECIPE_HEADERS)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
    for i, width in enumerate(_RECIPE_COL_WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width


def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class RecipeExportView(APIView):
    """GET — download all recipes (with ingredients) as an .xlsx file."""

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Recipes'
        _style_recipe_header(ws)

        recipes = ProductionRecipe.objects.prefetch_related('ingredients__item').all()
        for recipe in recipes:
            ingredients = list(recipe.ingredients.all())
            if not ingredients:
                ws.append([recipe.name, recipe.output_item.code, str(recipe.output_quantity),
                          recipe.notes, 'yes' if recipe.is_active else 'no', '', '', ''])
                continue
            for idx, ing in enumerate(ingredients):
                ws.append([
                    recipe.name if idx == 0 else '',
                    recipe.output_item.code if idx == 0 else '',
                    str(recipe.output_quantity) if idx == 0 else '',
                    recipe.notes if idx == 0 else '',
                    ('yes' if recipe.is_active else 'no') if idx == 0 else '',
                    ing.item.code, str(ing.quantity), ing.unit,
                ])

        return _xlsx_response(wb, 'production_recipes.xlsx')


class RecipeTemplateView(APIView):
    """GET — download a blank sample .xlsx showing the expected import format."""

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Recipes'
        _style_recipe_header(ws)

        notes = [
            'Recipe name (required, first row of each block)',
            'Output item code (required, first row of each block)',
            'Output quantity in smallest unit (required, first row of each block)',
            'Notes (optional)',
            'yes / no (optional, default yes)',
            'Ingredient item code (required, one per row)',
            'Ingredient quantity (required)',
            'small / medium / large (optional, default small)',
        ]
        ws.append(notes)
        note_font = Font(italic=True, color='595959')
        for cell in ws[2]:
            cell.font = note_font
            cell.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 48

        samples = [
            ('Basic Facial Serum', 'ITEM-001', 100, 'Standard batch', 'yes', 'ITEM-010', 60, 'small'),
            ('', '', '', '', '', 'ITEM-011', 40, 'small'),
            ('Whitening Cream 50g', 'ITEM-002', 50, '', 'yes', 'ITEM-012', 1, 'medium'),
        ]
        for row in samples:
            ws.append(list(row))

        return _xlsx_response(wb, 'production_recipes_sample.xlsx')


def _parse_bool(value, default=True):
    if value in (None, ''):
        return default
    s = str(value).strip().lower()
    return s in ('yes', 'true', '1', 'y')


class RecipeImportView(APIView):
    """
    POST multipart/form-data with file=<xlsx>
    Columns (row 1 = header, row 2 may be a notes row, skipped as data only if
    it fails to parse as a valid ingredient line):
      recipe_name | output_item_code | output_quantity | notes | is_active |
      ingredient_item_code | ingredient_quantity | ingredient_unit
    One row per ingredient. Recipe-level columns only need to be present on the
    first row of each recipe's block. Recipes matched by name are updated
    (ingredients replaced wholesale); unmatched names create a new recipe.
    Returns { created, updated, errors: [{row, message}] }
    """

    def post(self, request):
        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), read_only=True, data_only=True)
        except Exception:
            return Response({'error': 'Could not parse file. Upload a valid .xlsx file.'},
                            status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'error': 'File is empty.'}, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        groups = {}
        order = []
        last_name = None

        for i, row in enumerate(rows[1:], start=2):
            padded = (list(row) + [None] * 8)[:8]
            if all(v in (None, '') for v in padded):
                continue  # fully blank row — skip silently

            name = str(padded[0]).strip() if padded[0] not in (None, '') else ''
            if not name:
                name = last_name
            else:
                last_name = name

            ing_code = str(padded[5]).strip() if padded[5] not in (None, '') else ''
            ing_qty_raw = padded[6]
            ing_unit = str(padded[7]).strip().lower() if padded[7] not in (None, '') else 'small'

            if not name:
                errors.append({'row': i, 'message': 'recipe_name is required.'})
                continue

            grp = groups.get(name)
            if grp is None:
                grp = {'output_code': None, 'output_qty': None, 'notes': '',
                      'is_active': True, 'lines': [], 'first_row': i}
                groups[name] = grp
                order.append(name)

            out_code = str(padded[1]).strip() if padded[1] not in (None, '') else ''
            if out_code:
                grp['output_code'] = out_code
            if padded[2] not in (None, ''):
                grp['output_qty'] = padded[2]
            if padded[3] not in (None, ''):
                grp['notes'] = str(padded[3]).strip()
            if padded[4] not in (None, ''):
                grp['is_active'] = _parse_bool(padded[4])

            if not ing_code:
                errors.append({'row': i, 'message': 'ingredient_item_code is required.'})
                continue
            if ing_unit not in ('small', 'medium', 'large'):
                errors.append({'row': i, 'message': f'Invalid ingredient_unit "{ing_unit}".'})
                continue
            try:
                qty_dec = Decimal(str(ing_qty_raw))
                if qty_dec <= 0:
                    raise InvalidOperation
            except (InvalidOperation, ValueError, TypeError):
                errors.append({'row': i, 'message': 'ingredient_quantity must be a positive number.'})
                continue

            grp['lines'].append({'row': i, 'code': ing_code, 'qty': qty_dec, 'unit': ing_unit})

        created = updated = 0

        for name in order:
            grp = groups[name]
            row0 = grp['first_row']

            if not grp['output_code']:
                errors.append({'row': row0, 'message': f'"{name}": output_item_code is required.'})
                continue
            try:
                output_item = InventoryItem.objects.get(code=grp['output_code'])
            except InventoryItem.DoesNotExist:
                errors.append({'row': row0,
                              'message': f'"{name}": output item code "{grp["output_code"]}" not found.'})
                continue

            try:
                out_qty_dec = Decimal(str(grp['output_qty'])) if grp['output_qty'] not in (None, '') else None
                if not out_qty_dec or out_qty_dec <= 0:
                    raise InvalidOperation
            except (InvalidOperation, ValueError, TypeError):
                errors.append({'row': row0, 'message': f'"{name}": output_quantity must be a positive number.'})
                continue

            if not grp['lines']:
                errors.append({'row': row0, 'message': f'"{name}": no valid ingredient rows.'})
                continue

            resolved_lines = []
            seen_items = set()
            group_ok = True
            for ln in grp['lines']:
                try:
                    item = InventoryItem.objects.get(code=ln['code'])
                except InventoryItem.DoesNotExist:
                    errors.append({'row': ln['row'], 'message': f'Ingredient item code "{ln["code"]}" not found.'})
                    group_ok = False
                    continue
                if item.id in seen_items:
                    errors.append({'row': ln['row'],
                                  'message': f'Duplicate ingredient "{ln["code"]}" in recipe "{name}".'})
                    group_ok = False
                    continue
                seen_items.add(item.id)
                resolved_lines.append({'item': item, 'quantity': ln['qty'], 'unit': ln['unit']})

            if not group_ok or not resolved_lines:
                continue

            with transaction.atomic():
                recipe = ProductionRecipe.objects.filter(name=name).first()
                was_created = recipe is None
                if recipe is None:
                    recipe = ProductionRecipe.objects.create(
                        name=name, output_item=output_item, output_quantity=out_qty_dec,
                        notes=grp['notes'], is_active=grp['is_active'], created_by=_actor(request))
                else:
                    recipe.output_item = output_item
                    recipe.output_quantity = out_qty_dec
                    recipe.notes = grp['notes']
                    recipe.is_active = grp['is_active']
                    recipe.save()
                    recipe.ingredients.all().delete()
                for idx, ln in enumerate(resolved_lines):
                    ProductionRecipeIngredient.objects.create(
                        recipe=recipe, item=ln['item'], quantity=ln['quantity'],
                        unit=ln['unit'], ordering=idx)

            AuditLog.objects.create(
                performed_by=_actor(request), action='CREATE' if was_created else 'UPDATE',
                entity_type='ProductionRecipe', entity_id=str(recipe.id),
                description=f'Recipe {"imported" if was_created else "updated via import"}: {recipe.name}')
            if was_created:
                created += 1
            else:
                updated += 1

        return Response({'created': created, 'updated': updated, 'errors': errors})


# ── Ad-hoc cost preview ──────────────────────────────────────────────────────

class ProductionPreviewView(APIView):
    """POST { lines:[{item_id, quantity, unit}], output_quantity, warehouse_id? }"""
    def post(self, request):
        d = request.data
        wh = d.get('warehouse_id') or None
        try:
            breakdown, total, per_unit = build_cost_breakdown(
                d.get('lines', []), d.get('output_quantity', 0), wh)
        except (ValueError, KeyError, InventoryItem.DoesNotExist) as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'lines': breakdown, 'total_cost': str(total),
                         'cost_per_output_unit': str(per_unit)})


# ── Production run execution ──────────────────────────────────────────────────

class ProductionRunListCreateView(APIView):
    def get(self, request):
        qs = ProductionRun.objects.prefetch_related('ingredients').all()[:200]
        return Response(ProductionRunSerializer(qs, many=True).data)

    def post(self, request):
        d = request.data
        # --- validate output ---
        try:
            output_item = InventoryItem.objects.get(pk=d['output_item'])
            warehouse = Warehouse.objects.get(pk=d['output_warehouse'])
            output_qty = Decimal(str(d['output_quantity']))
            lines = d['lines']  # [{item_id, quantity, unit, source_warehouse_id?}]
        except (KeyError, InventoryItem.DoesNotExist, Warehouse.DoesNotExist):
            return Response({'error': 'Invalid output item, warehouse, or payload.'},
                            status=status.HTTP_400_BAD_REQUEST)
        if output_item.is_service:
            return Response({'error': 'Cannot produce a service item.'},
                            status=status.HTTP_400_BAD_REQUEST)
        if output_qty <= 0 or not lines:
            return Response({'error': 'Output quantity and ingredients are required.'},
                            status=status.HTTP_400_BAD_REQUEST)

        # --- resolve qty_small + pre-check stock (deduct nothing yet) ---
        resolved = []
        shortages = []
        for ln in lines:
            try:
                item = InventoryItem.objects.get(pk=ln['item_id'])
            except (KeyError, InventoryItem.DoesNotExist):
                return Response({'error': 'Invalid ingredient item.'},
                                status=status.HTTP_400_BAD_REQUEST)
            if item.is_service:
                return Response({'error': f'{item.code} is a service item.'},
                                status=status.HTTP_400_BAD_REQUEST)
            src_wh = ln.get('source_warehouse_id') or None
            try:
                qty_small = _to_small(item, Decimal(str(ln['quantity'])),
                                      ln.get('unit', 'small'))
            except (ValueError, KeyError) as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
            avail = available_small(item.id, src_wh)
            if avail < qty_small:
                shortages.append({'item_code': item.code,
                                  'needed': str(qty_small), 'available': str(avail)})
            resolved.append((item, qty_small, src_wh))
        if shortages:
            return Response({'error': 'Insufficient stock.', 'shortages': shortages},
                            status=status.HTTP_400_BAD_REQUEST)

        # --- commit atomically ---
        try:
            with transaction.atomic():
                total_cost = Decimal('0')
                consumed = []
                for item, qty_small, src_wh in resolved:
                    if src_wh:
                        shortfall, cogs = _fifo_deduct(item.id, int(src_wh), qty_small)
                    else:
                        shortfall, cogs = _fifo_deduct_global(item.id, qty_small)
                    if shortfall > 0:                      # race-condition guard
                        raise ValueError(f'Stock changed for {item.code}; aborted.')
                    total_cost += cogs
                    consumed.append((item, qty_small, cogs, src_wh))

                total_cost = total_cost.quantize(Decimal('0.01'))
                batch = InventoryBatch.objects.create(
                    item=output_item, warehouse=warehouse, input_date=date.today(),
                    quantity_initial=output_qty, quantity_remaining=output_qty,
                    value=total_cost, created_by=_actor(request))

                run = ProductionRun.objects.create(
                    recipe_id=d.get('recipe') or None, output_item=output_item,
                    output_warehouse=warehouse, output_quantity=output_qty,
                    total_cost=total_cost, produced_batch=batch,
                    notes=d.get('notes', ''), produced_by=_actor(request))
                for item, qty_small, cogs, src_wh in consumed:
                    ProductionRunIngredient.objects.create(
                        run=run, item=item, quantity_small=qty_small,
                        cost=cogs.quantize(Decimal('0.01')),
                        source_warehouse_id=int(src_wh) if src_wh else None)

                AuditLog.objects.create(
                    performed_by=_actor(request), action='CREATE',
                    entity_type='ProductionRun', entity_id=str(run.id),
                    description=(f'Produced {output_qty} {output_item.unit_small} of '
                                 f'[{output_item.code}] {output_item.name} '
                                 f'(cost {total_cost}) from {len(consumed)} ingredients'))
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_409_CONFLICT)

        return Response(ProductionRunSerializer(run).data,
                        status=status.HTTP_201_CREATED)


class ProductionRunDetailView(APIView):
    def get(self, request, pk):
        try:
            run = ProductionRun.objects.prefetch_related('ingredients').get(pk=pk)
        except ProductionRun.DoesNotExist:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(ProductionRunSerializer(run).data)
