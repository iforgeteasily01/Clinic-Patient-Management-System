import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction
from django.db.models import Count, Prefetch, Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.response import Response
from rest_framework.views import APIView

from ..models import (
    AppUser, AuditLog,
    StockOpnameSession, StockOpnameItem, Warehouse,
    StockOpnameTemplate, StockOpnameTemplateItem,
    InventoryItem, InventoryBatch,
)
from .inventory_page import _fifo_deduct


def _actor(request):
    return request.user if isinstance(request.user, AppUser) else None


def _serialize_session_detail(session):
    """Serialise one session including all item lines. Expects items pre-fetched."""
    items = list(session.items.all())
    return {
        'id': session.id,
        'date': str(session.date),
        'conducted_by': session.conducted_by,
        'notes': session.notes,
        'status': session.status,
        'started_at': session.started_at.isoformat() if session.started_at else None,
        'completed_at': session.completed_at.isoformat() if session.completed_at else None,
        'created_at': session.created_at.isoformat(),
        'created_by_name': session.created_by.display_name if session.created_by else None,
        'warehouse_id': session.warehouse_id,
        'warehouse_name': session.warehouse.name if session.warehouse else None,
        'items': [
            {
                'item_id': it.item_id,
                'item_code': it.item.code,
                'item_name': it.item.name,
                'item_unit_small': it.item.unit_small,
                'warehouse_id': it.warehouse_id,
                'warehouse_name': it.warehouse.name,
                'shelf1_qty': it.shelf1_qty,
                'shelf2_qty': it.shelf2_qty,
                'system_qty': it.system_qty,
                'is_loss': it.is_loss,
            }
            for it in items
        ],
    }


def _detail_qs(pk):
    """Single-session queryset with all related data pre-fetched, items sorted by code."""
    return (
        StockOpnameSession.objects
        .select_related('created_by', 'warehouse')
        .prefetch_related(
            Prefetch(
                'items',
                queryset=StockOpnameItem.objects
                    .select_related('item', 'warehouse')
                    .order_by('item__code'),
            )
        )
        .get(pk=pk)
    )


def _save_items(session, warehouse_id, items_data, merge=False):
    """
    Upsert opname items for a session.

    When merge=True (collaborative saves), a zero incoming value does not
    overwrite an existing non-zero value — only filled-in (> 0) values win.
    This lets multiple users fill in different portions without clobbering
    each other's work.
    """
    if not items_data:
        return
    wh_id = int(warehouse_id)
    if merge:
        existing_map = {
            it.item_id: it
            for it in StockOpnameItem.objects.filter(session=session, warehouse_id=wh_id)
        }
    for row in items_data:
        item_id = int(row['item_id'])
        inc_s1 = int(row.get('shelf1_qty') or 0)
        inc_s2 = int(row.get('shelf2_qty') or 0)
        if merge:
            ex = existing_map.get(item_id)
            s1 = inc_s1 if inc_s1 > 0 else (ex.shelf1_qty if ex else 0)
            s2 = inc_s2 if inc_s2 > 0 else (ex.shelf2_qty if ex else 0)
        else:
            s1, s2 = inc_s1, inc_s2
        StockOpnameItem.objects.update_or_create(
            session=session,
            item_id=item_id,
            warehouse_id=wh_id,
            defaults={
                'shelf1_qty': s1,
                'shelf2_qty': s2,
                'system_qty': int(row.get('system_qty') or 0),
                'is_loss': bool(row.get('is_loss', False)),
            },
        )


class StockOpnameSessionListCreateView(APIView):
    def get(self, request):
        sessions = (
            StockOpnameSession.objects
            .select_related('created_by', 'warehouse')
            .annotate(item_count_ann=Count('items', distinct=True))
            .order_by('-date', '-created_at')
        )
        data = []
        for s in sessions:
            data.append({
                'id': s.id,
                'date': str(s.date),
                'conducted_by': s.conducted_by,
                'notes': s.notes,
                'status': s.status,
                'started_at': s.started_at.isoformat() if s.started_at else None,
                'completed_at': s.completed_at.isoformat() if s.completed_at else None,
                'created_at': s.created_at.isoformat(),
                'created_by_name': s.created_by.display_name if s.created_by else None,
                'warehouse_id': s.warehouse_id,
                'warehouse_name': s.warehouse.name if s.warehouse else None,
                'item_count': s.item_count_ann,
            })
        return Response(data)

    def post(self, request):
        data = request.data
        date = data.get('date')
        conducted_by = (data.get('conducted_by') or '').strip()
        notes = (data.get('notes') or '').strip()
        warehouse_id = data.get('warehouse_id')
        items_data = data.get('items', [])

        if not date or not conducted_by or not warehouse_id:
            return Response(
                {'error': 'Tanggal, nama petugas, dan gudang wajib diisi.'},
                status=400,
            )

        warehouse = Warehouse.objects.filter(pk=warehouse_id).first()
        with transaction.atomic():
            session = StockOpnameSession.objects.create(
                date=date,
                conducted_by=conducted_by,
                notes=notes,
                status='draft',
                warehouse=warehouse,
                started_at=timezone.now(),
                created_by=_actor(request),
            )
            _save_items(session, warehouse_id, items_data)

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='CREATE',
            entity_type='StockOpnameSession',
            entity_id=str(session.id),
            description=(
                f'Sesi stok opname #{session.id} dimulai oleh {conducted_by} '
                f'untuk tanggal {date}.'
            ),
        )
        return Response(_serialize_session_detail(_detail_qs(session.id)), status=201)


class StockOpnameSessionDetailView(APIView):
    def get(self, request, pk):
        try:
            session = _detail_qs(pk)
        except StockOpnameSession.DoesNotExist:
            return Response({'error': 'Tidak ditemukan.'}, status=404)
        return Response(_serialize_session_detail(session))

    def put(self, request, pk):
        try:
            session = StockOpnameSession.objects.get(pk=pk)
        except StockOpnameSession.DoesNotExist:
            return Response({'error': 'Tidak ditemukan.'}, status=404)
        if session.status == 'completed':
            return Response({'error': 'Sesi yang sudah selesai tidak dapat diubah.'}, status=400)

        data = request.data
        warehouse_id = data.get('warehouse_id') or session.warehouse_id
        items_data = data.get('items', [])

        if 'conducted_by' in data and data['conducted_by']:
            session.conducted_by = data['conducted_by'].strip()
        if 'notes' in data:
            session.notes = (data['notes'] or '').strip()
        if 'date' in data and data['date']:
            session.date = data['date']
        session.save()

        with transaction.atomic():
            _save_items(session, warehouse_id, items_data, merge=True)

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='StockOpnameSession',
            entity_id=str(session.id),
            description=f'Draft stok opname #{session.id} diperbarui.',
        )
        return Response(_serialize_session_detail(_detail_qs(pk)))


class StockOpnameCompleteView(APIView):
    def post(self, request, pk):
        try:
            session = StockOpnameSession.objects.get(pk=pk)
        except StockOpnameSession.DoesNotExist:
            return Response({'error': 'Tidak ditemukan.'}, status=404)

        if session.status == 'completed':
            return Response({'error': 'Sesi sudah diselesaikan.'}, status=400)

        data = request.data
        warehouse_id = data.get('warehouse_id') or session.warehouse_id
        items_data = data.get('items', [])
        loss_item_ids = {int(row['item_id']) for row in items_data if row.get('is_loss')}

        with transaction.atomic():
            # Merge counts — non-zero values from any contributor win.
            _save_items(session, warehouse_id, items_data, merge=True)

            # Apply is_loss decisions from this final submission.
            session.items.filter(item_id__in=loss_item_ids).update(is_loss=True)
            session.items.exclude(item_id__in=loss_item_ids).update(is_loss=False)

            total_loss_qty = 0
            total_loss_cogs = Decimal('0')

            for opname_item in session.items.select_related('item').filter(is_loss=True):
                physical = opname_item.shelf1_qty + opname_item.shelf2_qty
                shortage = opname_item.system_qty - physical
                if shortage <= 0:
                    continue

                shortfall, cogs = _fifo_deduct(opname_item.item_id, int(warehouse_id), Decimal(shortage))
                total_loss_qty += shortage - shortfall
                total_loss_cogs += cogs

                AuditLog.objects.create(
                    performed_by=_actor(request),
                    action='UPDATE',
                    entity_type='StockOpnameLoss',
                    entity_id=str(session.id),
                    description=(
                        f'Stok opname #{session.id}: item #{opname_item.item_id} '
                        f'dicatat susut {shortage} {opname_item.item.unit_small} '
                        f'(HPP Rp {cogs:,.0f}).'
                    ),
                )

            session.status = 'completed'
            session.completed_at = timezone.now()
            session.save()

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='CREATE',
            entity_type='StockOpnameSession',
            entity_id=str(session.id),
            description=(
                f'Stok opname #{session.id} oleh {session.conducted_by} '
                f'pada {session.date} diselesaikan. '
                f'Total susut: {total_loss_qty} unit (HPP Rp {total_loss_cogs:,.0f}).'
            ),
        )
        return Response(_serialize_session_detail(_detail_qs(pk)))


# ── Stock Opname Template ──────────────────────────────────────────────────

class StockOpnameTemplateView(APIView):
    def get(self, request):
        warehouse_id = request.GET.get('warehouse_id')
        if not warehouse_id:
            return Response({'error': 'warehouse_id required'}, status=400)
        try:
            template = (
                StockOpnameTemplate.objects
                .prefetch_related(
                    Prefetch(
                        'template_items',
                        queryset=StockOpnameTemplateItem.objects
                            .select_related('item')
                            .order_by('section_order', 'item_order'),
                    )
                )
                .get(warehouse_id=warehouse_id)
            )
        except StockOpnameTemplate.DoesNotExist:
            return Response({'sections': [], 'exists': False})

        sections: dict[str, list] = {}
        section_order_map: dict[str, int] = {}
        for ti in template.template_items.all():
            if ti.section not in sections:
                sections[ti.section] = []
                section_order_map[ti.section] = ti.section_order
            sections[ti.section].append({
                'item_id': ti.item_id,
                'item_code': ti.item.code,
                'item_name': ti.item.name,
            })

        return Response({
            'warehouse_id': template.warehouse_id,
            'sections': [
                {'name': name, 'items': items}
                for name, items in sorted(
                    sections.items(), key=lambda x: section_order_map[x[0]]
                )
            ],
            'updated_at': template.updated_at.isoformat(),
        })

    def post(self, request):
        warehouse_id = request.data.get('warehouse_id')
        file = request.FILES.get('file')
        if not warehouse_id or not file:
            return Response({'error': 'warehouse_id and file required.'}, status=400)

        try:
            warehouse = Warehouse.objects.get(pk=warehouse_id)
        except Warehouse.DoesNotExist:
            return Response({'error': 'Warehouse not found.'}, status=404)

        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
        except Exception:
            return Response({'error': 'Could not parse file. Upload a valid .xlsx file.'}, status=400)

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'error': 'File is empty.'}, status=400)

        headers = [str(h).strip() if h else '' for h in rows[0]]
        sections = [(i, h) for i, h in enumerate(headers) if h]
        if not sections:
            return Response({'error': 'No section headers found in row 1.'}, status=400)

        all_codes: set[str] = set()
        for row in rows[1:]:
            for col_idx, _ in sections:
                val = row[col_idx] if col_idx < len(row) else None
                if val:
                    all_codes.add(str(val).strip().upper())

        item_map = {
            item.code.upper(): item
            for item in InventoryItem.objects.filter(code__in=all_codes)
        }

        errors = []
        template_entries = []
        seen_items: set[int] = set()

        for section_order, (col_idx, section_name) in enumerate(sections):
            item_order = 0
            for row in rows[1:]:
                val = row[col_idx] if col_idx < len(row) else None
                if not val:
                    continue
                code = str(val).strip().upper()
                item = item_map.get(code)
                if not item:
                    errors.append(f'Kode "{code}" di seksi "{section_name}" tidak ditemukan.')
                    continue
                if item.id in seen_items:
                    continue
                seen_items.add(item.id)
                template_entries.append((section_order, section_name, item_order, item))
                item_order += 1

        if not template_entries:
            return Response({'error': 'Tidak ada barang valid.\n' + '\n'.join(errors)}, status=400)

        with transaction.atomic():
            template_obj, _ = StockOpnameTemplate.objects.update_or_create(
                warehouse=warehouse,
                defaults={'created_by': _actor(request)},
            )
            StockOpnameTemplateItem.objects.filter(template=template_obj).delete()
            StockOpnameTemplateItem.objects.bulk_create([
                StockOpnameTemplateItem(
                    template=template_obj,
                    item=item,
                    section=section_name,
                    section_order=section_order,
                    item_order=item_order,
                )
                for section_order, section_name, item_order, item in template_entries
            ])

        return Response({
            'message': (
                f'Template disimpan: {len(template_entries)} barang '
                f'di {len(sections)} seksi.'
            ),
            'warnings': errors,
        }, status=201)

    def patch(self, request):
        """Reorder an existing template from a drag-and-drop edit.

        Body: { warehouse_id, sections: [{ name, item_ids: [..] }, ..] }
        The incoming order of sections and of item_ids within each section
        defines section_order / item_order. Items are rewritten wholesale.
        """
        warehouse_id = request.data.get('warehouse_id')
        sections = request.data.get('sections')
        if not warehouse_id or sections is None:
            return Response({'error': 'warehouse_id dan sections wajib diisi.'}, status=400)

        try:
            template_obj = StockOpnameTemplate.objects.get(warehouse_id=warehouse_id)
        except StockOpnameTemplate.DoesNotExist:
            return Response({'error': 'Template tidak ditemukan untuk gudang ini.'}, status=404)

        # Only keep item_ids that still exist, preserving the submitted order.
        requested_ids = [
            int(iid)
            for sec in sections
            for iid in sec.get('item_ids', [])
        ]
        valid_ids = set(
            InventoryItem.objects.filter(pk__in=requested_ids).values_list('id', flat=True)
        )

        entries = []
        seen: set[int] = set()
        for section_order, sec in enumerate(sections):
            name = (sec.get('name') or '').strip()
            item_order = 0
            for iid in sec.get('item_ids', []):
                iid = int(iid)
                if iid in seen or iid not in valid_ids:
                    continue
                seen.add(iid)
                entries.append(
                    StockOpnameTemplateItem(
                        template=template_obj,
                        item_id=iid,
                        section=name,
                        section_order=section_order,
                        item_order=item_order,
                    )
                )
                item_order += 1

        if not entries:
            return Response({'error': 'Tidak ada barang valid untuk disimpan.'}, status=400)

        with transaction.atomic():
            StockOpnameTemplateItem.objects.filter(template=template_obj).delete()
            StockOpnameTemplateItem.objects.bulk_create(entries)
            template_obj.save(update_fields=['updated_at'])

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='StockOpnameTemplate',
            entity_id=str(template_obj.id),
            description=(
                f'Urutan template stok opname gudang #{warehouse_id} '
                f'diperbarui ({len(entries)} barang).'
            ),
        )
        return Response({'message': f'Urutan template disimpan ({len(entries)} barang).'})


class StockOpnameTemplateSampleView(APIView):
    def get(self, request):
        warehouse_id = request.GET.get('warehouse_id')

        if warehouse_id:
            item_ids = (
                InventoryBatch.objects
                .filter(warehouse_id=warehouse_id, quantity_remaining__gt=0,
                        item__is_active=True, item__is_service=False)
                .values('item_id')
                .annotate(total=Sum('quantity_remaining'))
                .filter(total__gt=0)
                .values_list('item_id', flat=True)
            )
            items = list(
                InventoryItem.objects
                .filter(pk__in=item_ids, is_service=False)
                .order_by('code')
            )
        else:
            items = list(
                InventoryItem.objects
                .filter(is_active=True, is_service=False)
                .order_by('code')[:50]
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'SO Template'

        blue_fill = PatternFill(fill_type='solid', fgColor='2563EB')
        white_bold = Font(bold=True, color='FFFFFF')
        center = Alignment(horizontal='center')

        sample_sections = ['Seksi 1', 'Seksi 2', 'Seksi 3']
        for col_idx, name in enumerate(sample_sections, start=1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = white_bold
            cell.fill = blue_fill
            cell.alignment = center
            ws.column_dimensions[cell.column_letter].width = 28

        for row_idx, item in enumerate(items, start=2):
            ws.cell(row=row_idx, column=1, value=item.code)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="so_template_sample.xlsx"'
        return response
