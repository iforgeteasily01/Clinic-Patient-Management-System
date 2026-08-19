import io
import logging
from decimal import Decimal, InvalidOperation

import openpyxl
from django.db import transaction
from django.db.models import F
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView

from ..api.serializers import (
    InvoiceCreateSerializer, InvoiceReadSerializer, InvoiceUpdateSerializer,
    LedgerEntrySerializer,
)
from ..models import (
    AppUser, AuditLog, ChartOfAccounts, InventoryItem, Invoice, InvoiceItem,
    InvoicePayment,
    LedgerEntry, Patient, PatientPackage, PatientPackageRedemption, PaymentMethod,
    PromotionUsage, Treatment, TreatmentPackage, Warehouse,
)
from ..services.cash_accounts import cash_bank_account_ids
from ..services.journal_engine import LegSet, _apply_balance, _le, _post_legs, _revenue_legs
from .crm_page import refresh_crm_profile
from .inventory_page import (
    FifoSimulation, _fifo_deduct, _fifo_restock,
)
from .promotion_page import validate_promotion

logger = logging.getLogger(__name__)

# ── Shared Excel column layout ────────────────────────────────────────────────
# Sheet "Invoices": invoice_number | datetime | patient_no | payment_method |
#                   discount | tax | additional_charges | grand_total |
#                   cashier_id | warehouse_id
# Sheet "Items":    invoice_number | item_code | item_name | quantity | price
# ─────────────────────────────────────────────────────────────────────────────

INV_HEADERS  = ['invoice_number','datetime','patient_no','payment_method',
                'discount','tax','additional_charges','grand_total',
                'cashier_id','warehouse_id']
ITEM_HEADERS = ['invoice_number','item_code','item_name','quantity','price']


def _actor(request):
    return request.user if isinstance(request.user, AppUser) else None


class PaymentSplitError(Exception):
    """Raised by _resolve_payment_splits with the DRF error body to return."""

    def __init__(self, detail):
        super().__init__(detail)
        self.detail = detail


def _resolve_payment_splits(payments, grand_total):
    """Validate split-payment input and resolve it to unsaved InvoicePayment rows.

    Returns every tender in order. The caller takes the first for the invoice's
    own ``payment_method`` / ``payment_account`` fields, and only *persists* the
    rows when there are two or more — a single tender is fully described by
    those scalars, and storing it as a row would make every one-method invoice
    look like a split to any reader that checks for one.

    Raises ``PaymentSplitError`` on anything that would post an unbalanced or
    unroutable entry — an unknown method, a non-cash account, or amounts that do
    not sum to ``grand_total``. The last one is the important one: these rows are
    the debit side of the journal entry, so if the POS sends cash tendered
    instead of cash applied, the entry silently lands the change in Sales
    Discount. Better to reject it.
    """
    if not payments:
        return []

    cash_ids = cash_bank_account_ids()
    rows = []
    total = Decimal('0')

    for idx, p in enumerate(payments):
        method = None
        if p.get('payment_method_id'):
            method = PaymentMethod.objects.select_related('linked_account').filter(
                pk=p['payment_method_id']).first()
            if method is None:
                raise PaymentSplitError(
                    {'payments': [f'Payment method {p["payment_method_id"]} not found.']})

        account = None
        if p.get('payment_account_id'):
            if p['payment_account_id'] not in cash_ids:
                raise PaymentSplitError(
                    {'payments': [f'Account {p["payment_account_id"]} is not a cash/bank account.']})
            account = ChartOfAccounts.objects.filter(pk=p['payment_account_id']).first()
        elif method is not None:
            account = method.linked_account

        amount = p['amount']
        total += amount
        rows.append(InvoicePayment(
            payment_method=method, payment_account=account,
            amount=amount, sort_order=idx,
        ))

    if abs(total - grand_total) > Decimal('0.01'):
        raise PaymentSplitError({'payments': [
            f'Payment amounts total {total}, which does not match grand_total '
            f'{grand_total}. Send the amount each method settles, not the cash tendered.'
        ]})

    return rows


# _le / _apply_balance / _revenue_legs / _post_legs and the ACC_* constants now
# live in managementsys/services/journal_engine.py (Phase 2) — imported above
# so every call site below is unchanged. This module still owns _post_accounting
# / _reverse_accounting_instances (FIFO + package side effects), which were not
# part of the extraction.


def _lines_from_dicts(line_items, items_by_id):
    return [
        (items_by_id.get(l['item_id']) if l.get('item_id') else None,
         l.get('item_name', ''), l['quantity'], l['price'])
        for l in line_items
    ]


def _lines_from_instances(item_instances):
    return [
        (inst.item if inst.item_id else None, inst.item_name, inst.quantity, inst.price)
        for inst in item_instances
    ]


def build_invoice_legs(invoice, line_items, items_by_id, *, deduct=True, sim=None):
    """Every journal leg a completed invoice produces, as a LegSet:
      - Cash/payment account  += grand_total  (DEBIT asset)
      - Per service line: revenue for the treatment's category  (CREDIT revenue)
      - Per physical line (is_service=False): FIFO deduct stock, then
        inventory asset -= COGS  (CREDIT asset),  COGS account += COGS  (DEBIT COGS)
    Routing:
      - Services (treatments) post to their item_category's revenue accounts.
      - Physical inventory items are treated as a single entity — all post to the
        shared product accounts (4200000 Product Sales Revenue / 5100000 Cost of
        Products Sold), regardless of item_category.
    These product accounts also serve as the fallback when a service has no category.

    ``deduct`` is the one thing that makes this *not* a pure function. The COGS
    legs are a result of FIFO consumption, not an input to it, so a real posting
    must actually consume stock to know what to post. With ``deduct=False`` the
    FIFO walk runs against a ``FifoSimulation`` instead — no rows are touched,
    and the resulting COGS legs are flagged ``is_estimated`` because stock can
    move between a preview and its commit. Pass one ``sim`` for the whole
    preview so repeated draws on the same batch stay honest.
    """
    inv_no = invoice.invoice_number
    legset = LegSet(memo=f'Invoice {inv_no}')

    # Cash, revenue, tax, additional charges and the discount plug, as one
    # self-balancing block covering every line — including lines with no item FK.
    for account, entry_type, amount, description in _revenue_legs(
        invoice, _lines_from_dicts(line_items, items_by_id)
    ):
        legset.add(account, entry_type, amount, description)

    fallback_cogs = ChartOfAccounts.objects.filter(account_number=5100000).first()
    inventory_asset = ChartOfAccounts.objects.filter(account_number=1300000).first()

    def cogs_pair(amount, label_asset, label_cogs):
        """The balanced inventory/COGS pair for one consumption event."""
        if inventory_asset:
            legset.add(inventory_asset, 'credit', amount, label_asset, is_estimated=not deduct)
        if fallback_cogs:
            legset.add(fallback_cogs, 'debit', amount, label_cogs, is_estimated=not deduct)
        if not inventory_asset or not fallback_cogs:
            legset.warnings.append('Akun persediaan/HPP tidak lengkap — HPP tidak dijurnal.')

    for line in line_items:
        if not line.get('item_id'):
            continue
        item = items_by_id[line['item_id']]

        if not item.is_service and invoice.warehouse_id:
            if line['quantity'] > 0:
                _shortfall, cogs_amount = _fifo_deduct(
                    item.id, invoice.warehouse_id, line['quantity'],
                    commit=deduct, sim=sim,
                )
                if cogs_amount > 0:
                    cogs_pair(cogs_amount,
                              f'Invoice {inv_no} – FIFO deduction: {item.name}',
                              f'Invoice {inv_no} – COGS: {item.name}')

        # Service lines post no cost. A treatment used to consume a fixed
        # bill-of-materials (TreatmentMaterial) and book the FIFO result as
        # COGS; that model was removed because a fixed recipe does not survive
        # contact with real usage — the actual quantity of cleanser, anaesthetic
        # or filler used varies per patient and per operator, so the posted cost
        # was precise and wrong. Cost of sales for treatments is now entered by
        # hand as a periodic journal (see migration 0107).

    return legset


def _post_accounting(invoice, line_items, items_by_id, memo_date=None, source_type='invoice'):
    """Write the invoice's journal legs and roll every affected balance.

    The arithmetic — and the FIFO consumption it depends on — lives in
    ``build_invoice_legs``. This is the thin writer, kept so every existing
    create/edit/void call site is unchanged.
    """
    legset = build_invoice_legs(invoice, line_items, items_by_id, deduct=True)
    _post_legs(
        invoice,
        [(l.account, l.entry_type, l.amount, l.description) for l in legset.legs],
        date=memo_date, source_type=source_type,
    )


def _snapshot_scalars(invoice):
    """Capture the invoice fields the GL block is derived from.

    A PUT writes the new scalars onto the instance before the reversal runs, so
    the reversal must be handed the pre-edit values or it will unwind amounts
    that were never posted.
    """
    return {
        'grand_total': invoice.grand_total,
        'tax': invoice.tax,
        'additional_charges': invoice.additional_charges,
        'discount': invoice.discount,
        'datetime': invoice.datetime,
        'payment_method_id': invoice.payment_method_id,
        'payment_account_id': invoice.payment_account_id,
    }


def _reverse_accounting_instances(payment_method_id, grand_total, item_instances, warehouse_id,
                                  invoice=None, old_scalars=None, memo_date=None, source_type='invoice'):
    """
    Reverse every accounting entry that _post_accounting originally made,
    using InvoiceItem model instances (with item relations already prefetched).
    Called at the start of a PUT/PATCH so the edit can be re-applied cleanly.
    Each reversal is recorded as a LedgerEntry (opposite entry_type) when invoice is supplied.

    ``memo_date``/``source_type`` let the void/edit-memo path (Phase 2) stamp
    these reversal rows as today-dated, tagged 'void_memo'/'edit_memo' instead
    of the historic 'invoice' rows on the original transaction date.
    """
    inv_no = invoice.invoice_number if invoice else '?'

    if invoice is not None:
        # Undo the exact block _post_accounting wrote. Rebuilding it from the
        # same inputs and flipping every side keeps the two functions mirror
        # images by construction rather than by manual upkeep.
        old = dict(old_scalars or {})
        old.setdefault('grand_total', grand_total)
        old.setdefault('payment_method_id', payment_method_id)
        # Same treatment as payment_method_id: an edit that changes the bank
        # account must reverse against the account the *original* posting
        # actually debited, not the one the invoice has been updated to by
        # the time this reversal runs — otherwise the reversal leg lands on
        # the wrong account and both balances end up wrong.
        old.setdefault('payment_account_id', invoice.payment_account_id)
        saved = _snapshot_scalars(invoice)
        try:
            invoice.grand_total        = old['grand_total']
            invoice.tax                = old.get('tax', invoice.tax)
            invoice.additional_charges = old.get('additional_charges', invoice.additional_charges)
            invoice.discount           = old.get('discount', invoice.discount)
            invoice.datetime           = old.get('datetime', invoice.datetime)
            if old['payment_method_id'] != invoice.payment_method_id:
                invoice.payment_method = PaymentMethod.objects.filter(
                    pk=old['payment_method_id']).first()
            if old['payment_account_id'] != invoice.payment_account_id:
                invoice.payment_account = ChartOfAccounts.objects.filter(
                    pk=old['payment_account_id']).first()
            legs = _revenue_legs(invoice, _lines_from_instances(item_instances))
            _post_legs(invoice, legs, reverse=True, date=memo_date, source_type=source_type)
        finally:
            invoice.grand_total        = saved['grand_total']
            invoice.tax                = saved['tax']
            invoice.additional_charges = saved['additional_charges']
            invoice.discount           = saved['discount']
            invoice.datetime           = saved['datetime']
            if invoice.payment_method_id != saved['payment_method_id']:
                invoice.payment_method = PaymentMethod.objects.filter(
                    pk=saved['payment_method_id']).first()
            if invoice.payment_account_id != saved['payment_account_id']:
                invoice.payment_account = ChartOfAccounts.objects.filter(
                    pk=saved['payment_account_id']).first()
    elif payment_method_id:
        method = PaymentMethod.objects.select_related('linked_account').filter(pk=payment_method_id).first()
        if method:
            _apply_balance(method.linked_account, 'credit', grand_total)
    # (no LedgerEntry row for the payment-method-only branch above — matches
    # historic behaviour; it only runs when no invoice was supplied at all)

    fallback_cogs    = ChartOfAccounts.objects.filter(account_number=5100000).first()
    inventory_asset  = ChartOfAccounts.objects.filter(account_number=1300000).first()

    for inst in item_instances:
        if not inst.item_id:
            continue
        item = inst.item
        # Mirror of _post_accounting: COGS always posts to the shared fallback.
        cogs_acct    = fallback_cogs

        if not item.is_service and warehouse_id:
            if inst.quantity > 0:
                cogs_amount = _fifo_restock(item.id, warehouse_id, inst.quantity)
                if cogs_amount > 0:
                    if inventory_asset:
                        ChartOfAccounts.objects.filter(pk=inventory_asset.pk).update(
                            balance=F('balance') + cogs_amount
                        )
                        if invoice:
                            _le(inventory_asset, 'debit', cogs_amount, invoice,
                                f'Invoice {inv_no} – FIFO restock: {item.name}',
                                date=memo_date, source_type=source_type)
                    if cogs_acct:
                        ChartOfAccounts.objects.filter(pk=cogs_acct.pk).update(
                            balance=F('balance') - cogs_amount
                        )
                        if invoice:
                            _le(cogs_acct, 'credit', cogs_amount, invoice,
                                f'Invoice {inv_no} – COGS correction: {item.name}',
                                date=memo_date, source_type=source_type)

        # Service lines consume nothing, so there is nothing to give back.
        # The TreatmentMaterial bill-of-materials this used to mirror was removed
        # (migration 0107) — see the matching note in build_invoice_legs. Keeping
        # a reversal for a deduction that no longer happens would credit stock
        # the sale never took.


def _handle_packages(invoice, patient, line_items, items_by_id):
    """Process package sales and redemptions for a saved invoice.

    Silently skips redemptions whose package has no remaining sessions for the
    requested treatment — the line still posts at the price the cashier sent
    (cashier sees remaining count in the UI, so this is a defensive fallback,
    not an authorization check).
    """
    # ── Sales ──
    if patient is not None:
        package_by_catalog = {
            tp.catalog_item_id: tp
            for tp in TreatmentPackage.objects.filter(
                catalog_item_id__in=[i for i in items_by_id]
            )
        }
        for line in line_items:
            if not line.get('item_id'):
                continue
            tp = package_by_catalog.get(line['item_id'])
            if not tp:
                continue
            qty = int(line['quantity'].to_integral_value())
            for _ in range(max(qty, 0)):
                PatientPackage.objects.create(
                    patient=patient,
                    package=tp,
                    purchased_invoice=invoice,
                )

    # ── Redemptions ──
    touched = set()
    for line in line_items:
        pp_id = line.get('redeem_patient_package_id')
        treatment_id = line.get('treatment_id')
        if not pp_id or not treatment_id:
            continue
        try:
            pp = PatientPackage.objects.select_related('package').get(pk=pp_id)
        except PatientPackage.DoesNotExist:
            continue
        if pp.remaining_for(treatment_id) <= 0:
            continue
        try:
            treatment = Treatment.objects.get(pk=treatment_id)
        except Treatment.DoesNotExist:
            continue
        qty = max(int(line['quantity'].to_integral_value()), 1)
        for _ in range(qty):
            if pp.remaining_for(treatment_id) <= 0:
                break
            PatientPackageRedemption.objects.create(
                patient_package=pp,
                treatment=treatment,
                invoice=invoice,
            )
        touched.add(pp.id)

    for pp_id in touched:
        try:
            PatientPackage.objects.get(pk=pp_id).refresh_status()
        except PatientPackage.DoesNotExist:
            pass


def _foreign_redemptions_exist(invoice):
    """
    True when a package sold by this invoice has been redeemed on a *different*
    invoice. Reversing the sale would cascade-delete those redemptions, so the
    edit is refused instead.
    """
    return (
        PatientPackageRedemption.objects
        .filter(patient_package__purchased_invoice=invoice)
        .exclude(invoice=invoice)
        .exists()
    )


def _reverse_packages(invoice):
    """
    Undo the package sales and redemptions _handle_packages made for this
    invoice, so they can be re-applied against the edited lines.
    Caller must have checked _foreign_redemptions_exist first.
    """
    touched = set(
        PatientPackageRedemption.objects
        .filter(invoice=invoice)
        .values_list('patient_package_id', flat=True)
    )
    PatientPackageRedemption.objects.filter(invoice=invoice).delete()
    PatientPackage.objects.filter(purchased_invoice=invoice).delete()

    for pp in PatientPackage.objects.filter(pk__in=touched):
        pp.refresh_status()


def resolve_line_item_ids(lines):
    """Fill in ``item_id`` for lines that named a treatment but did not link it.

    POSPage's "transfer from billing queue" and package-redemption paths build
    their lines from a Treatment, send ``treatment_id`` faithfully, and then set
    ``item_id`` to null — so the line arrived with everything needed to link it
    and no link. 767 InvoiceItem rows were written that way, 673 of them in June
    and July 2026 alone, and every single one of their treatments had a
    ``catalog_item`` sitting there available.

    An unlinked line is not cosmetic. It has no FK to route revenue by, so
    ``journal_engine._line_revenue_account`` falls back to matching the free-text
    name against Treatment — which works, but silently fails on a typo (see
    migration 0077) — and no ``TreatmentMaterial`` is consumed, so the treatment
    costs nothing.

    Resolution order, first hit wins:

    1. ``treatment_id`` → that Treatment's ``catalog_item``. Authoritative: the
       client told us exactly which treatment this is.
    2. ``item_name`` → a Treatment of that name (case- and space-insensitive)
       → its ``catalog_item``. Covers clients that only send a name, including
       the Medya-Cashier POS.

    Mutates ``lines`` in place and returns the number of lines linked. Lines that
    already carry an ``item_id``, and names that match nothing, are left exactly
    as they are — a genuine ad-hoc line must stay ad-hoc.
    """
    pending = [
        line for line in lines
        if not line.get('item_id') and (line.get('treatment_id') or line.get('item_name'))
    ]
    if not pending:
        return 0

    treatment_ids = {line['treatment_id'] for line in pending if line.get('treatment_id')}
    by_id = {}
    if treatment_ids:
        by_id = dict(
            Treatment.objects
            .filter(id__in=treatment_ids, catalog_item__isnull=False)
            .values_list('id', 'catalog_item_id')
        )

    names = {
        (line.get('item_name') or '').strip().lower()
        for line in pending
        if not line.get('treatment_id') and line.get('item_name')
    }
    names.discard('')
    by_name = {}
    if names:
        # Name matching is done in Python, not with a case-insensitive IN
        # lookup, so that trailing spaces and mixed case both normalise the same
        # way they do in journal_engine._line_revenue_account.
        by_name = {
            t.name.strip().lower(): t.catalog_item_id
            for t in Treatment.objects.filter(catalog_item__isnull=False).only(
                'id', 'name', 'catalog_item_id')
            if t.name.strip().lower() in names
        }

    linked = 0
    for line in pending:
        catalog_item_id = None
        if line.get('treatment_id'):
            catalog_item_id = by_id.get(line['treatment_id'])
        if catalog_item_id is None and line.get('item_name'):
            catalog_item_id = by_name.get(line['item_name'].strip().lower())
        if catalog_item_id is not None:
            line['item_id'] = catalog_item_id
            linked += 1
    return linked


class InvoiceCreateView(APIView):
    permission_classes = [AllowAny]

    @transaction.atomic
    def post(self, request):
        serializer = InvoiceCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data

        # ── Resolve FKs ───────────────────────────────────────────────────────

        patient = None
        if data.get('patient_no'):
            try:
                patient = Patient.objects.get(patient_no=data['patient_no'])
            except Patient.DoesNotExist:
                return Response(
                    {'patient_no': 'Patient not found.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        cashier = None
        if data.get('cashier_id'):
            try:
                cashier = AppUser.objects.get(id=data['cashier_id'])
            except AppUser.DoesNotExist:
                return Response(
                    {'cashier_id': 'Cashier not found.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        warehouse = None
        if data.get('warehouse_id'):
            try:
                warehouse = Warehouse.objects.get(id=data['warehouse_id'])
            except Warehouse.DoesNotExist:
                return Response(
                    {'warehouse_id': 'Warehouse not found.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        payment_method_obj = None
        if data.get('payment_method_id'):
            try:
                payment_method_obj = PaymentMethod.objects.get(
                    id=data['payment_method_id'],
                    is_active=True,
                )
            except PaymentMethod.DoesNotExist:
                return Response(
                    {'payment_method_id': 'Payment method not found.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # ── Payment account (design doc §3) ────────────────────────────────────
        payment_account_obj = None
        payment_account_id = data.get('payment_account_id')
        if payment_account_id is not None:
            if payment_account_id not in cash_bank_account_ids():
                return Response(
                    {'payment_account_id': ['Not a cash/bank account.']},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                payment_account_obj = ChartOfAccounts.objects.get(pk=payment_account_id)
            except ChartOfAccounts.DoesNotExist:
                return Response(
                    {'payment_account_id': ['Not a cash/bank account.']},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        elif payment_method_obj is not None:
            # A caller that only sends payment_method_id (Medya-Cashier POS,
            # still legacy) still gets payment_account resolved from it, so a
            # freshly created invoice is never write-only-legacy.
            payment_account_obj = payment_method_obj.linked_account

        # ── Split payment (optional) ──────────────────────────────────────────
        try:
            tenders = _resolve_payment_splits(data.get('payments'), data['grand_total'])
        except PaymentSplitError as exc:
            return Response(exc.detail, status=status.HTTP_400_BAD_REQUEST)

        # Only a genuine split is stored as rows; see _resolve_payment_splits.
        split_rows = tenders if len(tenders) > 1 else []

        if tenders:
            # The invoice-level fields keep describing the first tender, so
            # every existing reader (list filters, export, receipt, Vercel push)
            # still sees a payment method. The split rows carry the rest.
            first = tenders[0]
            payment_method_obj  = payment_method_obj or first.payment_method
            payment_account_obj = payment_account_obj or first.payment_account

        # ── Link treatment lines the client sent unlinked ─────────────────────
        # Must run before item_ids is collected: a line linked here has to be
        # validated, stock-checked and material-consumed like any other.
        resolve_line_item_ids(data['items'])

        # ── Validate all inventory items exist before writing anything ────────

        item_ids = [i['item_id'] for i in data['items'] if i.get('item_id')]
        items_by_id = {
            obj.id: obj
            for obj in InventoryItem.objects.filter(id__in=item_ids).select_related(
                'item_category__revenue_account',
            ).prefetch_related('treatment')
        }
        missing = [iid for iid in item_ids if iid not in items_by_id]
        if missing:
            return Response(
                {'items': f"Item IDs not found: {missing}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Create Invoice ────────────────────────────────────────────────────

        invoice = Invoice.objects.create(
            datetime=data.get('datetime') or timezone.now(),
            patient_no=patient,
            payment_method=payment_method_obj,
            payment_account=payment_account_obj,
            discount=data['discount'],
            cashier=cashier,
            warehouse=warehouse,
            tax=data['tax'],
            additional_charges=data['additional_charges'],
            grand_total=data['grand_total'],
            notes=data.get('notes', ''),
            promotion_code=(data.get('promotion_code') or '').strip(),
        )

        # ── Create InvoiceItems ───────────────────────────────────────────────

        InvoiceItem.objects.bulk_create([
            InvoiceItem(
                invoice=invoice,
                item=items_by_id[i['item_id']] if i.get('item_id') else None,
                item_name=i.get('item_name', '') if not i.get('item_id') else '',
                quantity=i['quantity'],
                price=i['price'],
                discount_pct=i.get('discount_pct', 0),
            )
            for i in data['items']
        ])

        # ── Split payment rows ────────────────────────────────────────────────

        if split_rows:
            for row in split_rows:
                row.invoice = invoice
            InvoicePayment.objects.bulk_create(split_rows)

        # ── Accounting + stock deduction ──────────────────────────────────────
        # Phase 2: journal posting is deferred. A brand-new invoice always
        # starts posting_status='unposted' (the model default) and carries
        # zero LedgerEntry rows until POST /api/accounting/journal/run/ sweeps
        # its transaction date — see managementsys/services/journal_engine.py.

        # ── Treatment Packages: sales + redemptions ───────────────────────────
        # Sale: any line whose item is a TreatmentPackage.catalog_item creates
        #   one PatientPackage per quantity (rounded down).
        # Redemption: any line with redeem_patient_package_id consumes one
        #   session from that PatientPackage for the given treatment.
        _handle_packages(invoice, patient, data['items'], items_by_id)

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='CREATE',
            entity_type='Invoice',
            entity_id=str(invoice.id),
            description=f'Invoice {invoice.invoice_number} created, {len(data["items"])} item(s), total {invoice.grand_total}',
        )

        # ── Promotion ─────────────────────────────────────────────────────────

        promotion_code = (request.data.get('promotion_code') or '').strip()
        if promotion_code:
            promo, discount_amount, error = validate_promotion(
                promotion_code, patient, data['grand_total']
            )
            if promo:
                PromotionUsage.objects.create(
                    promotion=promo,
                    patient_no=patient,
                    invoice=invoice,
                    discount_applied=discount_amount,
                )
                invoice.promotion = promo
                invoice.save(update_fields=['promotion'])
            else:
                logger.warning(
                    'Promotion code %r skipped on invoice %s: %s',
                    promotion_code, invoice.invoice_number, error,
                )

        # ── CRM refresh ───────────────────────────────────────────────────────

        if patient is not None:
            refresh_crm_profile(patient)

        invoice = (
            Invoice.objects
            .select_related('patient_no', 'cashier', 'warehouse', 'payment_method', 'payment_account')
            .prefetch_related('items__item', 'payments__payment_method', 'payments__payment_account')
            .get(pk=invoice.pk)
        )
        return Response(
            InvoiceReadSerializer(invoice).data,
            status=status.HTTP_201_CREATED,
        )


class InvoiceListView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        qs = (
            Invoice.objects
            .select_related('patient_no', 'cashier', 'warehouse', 'payment_method', 'payment_account')
            .prefetch_related('items__item', 'payments__payment_method', 'payments__payment_account')
            .order_by('-datetime')
        )
        if inv_no := request.GET.get('invoice_number', '').strip():
            qs = qs.filter(invoice_number=inv_no)
        elif q := request.GET.get('q', '').strip():
            from django.db.models import Q
            qs = qs.filter(
                Q(invoice_number__icontains=q) |
                Q(patient_no__name__icontains=q) |
                Q(patient_no__patient_no__icontains=q)
            )
        if method := request.GET.get('payment_method', '').strip():
            qs = qs.filter(payment_method_id=method)
        if account := request.GET.get('payment_account', '').strip():
            qs = qs.filter(payment_account_id=account)

        # Hide voided invoices unless the caller explicitly asks to include them.
        if request.GET.get('include_voided', '').strip().lower() not in ('1', 'true', 'yes'):
            qs = qs.filter(is_voided=False)

        total = qs.count()

        try:
            page_size = int(request.GET.get('page_size', 50))
        except (ValueError, TypeError):
            page_size = 50
        page_size = page_size if page_size in (10, 50, 100) else 50

        try:
            page = max(int(request.GET.get('page', 1)), 1)
        except (ValueError, TypeError):
            page = 1

        offset = (page - 1) * page_size
        qs = qs[offset:offset + page_size]

        return Response({
            'count': total,
            'results': InvoiceReadSerializer(qs, many=True).data,
        })


class InvoiceDetailView(APIView):
    permission_classes = [AllowAny]

    def _get(self, pk):
        try:
            return (
                Invoice.objects
                .select_related('patient_no', 'cashier', 'warehouse', 'payment_method', 'payment_account')
                .prefetch_related('items__item', 'payments__payment_method', 'payments__payment_account')
                .get(pk=pk)
            )
        except Invoice.DoesNotExist:
            return None

    def get(self, request, pk):
        invoice = self._get(pk)
        if invoice is None:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        data = InvoiceReadSerializer(invoice).data
        # The GL postings this invoice produced — including the reversal legs
        # written by a later edit or void, which is why they are ordered
        # oldest-first and not filtered. Only the detail endpoint carries them;
        # the list would be one query per row.
        entries = (
            LedgerEntry.objects
            .filter(invoice=invoice)
            .select_related('account')
            .order_by('date', 'created_at', 'id')
        )
        data['ledger_entries'] = LedgerEntrySerializer(entries, many=True).data
        return Response(data)

    @transaction.atomic
    def patch(self, request, pk):
        return self.put(request, pk)

    @transaction.atomic
    def put(self, request, pk):
        invoice = self._get(pk)
        if invoice is None:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if invoice.is_voided:
            return Response(
                {'error': 'Cannot edit a voided invoice.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = InvoiceUpdateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        changes = []

        # ── Resolve and validate the new lines before touching anything ───────
        # Everything below this point mutates; a 400 returned after a write
        # would still commit, since only an exception rolls the atomic block back.

        replacing_items = 'items' in data
        new_items_by_id: dict = {}

        if replacing_items:
            resolve_line_item_ids(data['items'])
            item_ids = [i['item_id'] for i in data['items'] if i.get('item_id')]
            new_items_by_id = {
                obj.id: obj
                for obj in InventoryItem.objects.filter(id__in=item_ids).select_related(
                    'item_category__revenue_account',
                ).prefetch_related('treatment')
            }
            missing = [iid for iid in item_ids if iid not in new_items_by_id]
            if missing:
                return Response({'items': f'Item IDs not found: {missing}'}, status=status.HTTP_400_BAD_REQUEST)

            # Replacing the lines reverses this invoice's package sales. If another
            # invoice already redeemed against one, reversing would destroy it.
            if _foreign_redemptions_exist(invoice):
                return Response(
                    {'items': 'Cannot edit items: a treatment package sold by this '
                              'invoice has already been redeemed on another invoice. '
                              'Void the redeeming invoice first.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # ── Split payment (optional) ──────────────────────────────────────────
        # Resolved up-front so a bad split is a 400 before anything is written,
        # but the rows are only swapped in *after* the reversal below — the
        # reversal rebuilds the old legs off invoice.payments and would unwind
        # the new split instead of the posted one.
        replacing_payments = 'payments' in data
        new_tenders = []
        if replacing_payments:
            effective_total = data.get('grand_total', invoice.grand_total)
            try:
                new_tenders = _resolve_payment_splits(data['payments'], effective_total)
            except PaymentSplitError as exc:
                return Response(exc.detail, status=status.HTTP_400_BAD_REQUEST)
        new_split_rows = new_tenders if len(new_tenders) > 1 else []

        # ── Capture old state before any modifications ────────────────────────
        # The GL fingerprint is taken here, against the rows still in the
        # database, and again after every mutation below. Equal fingerprints mean
        # the edit touched nothing the posting is derived from, so the edit-memo
        # reversal + repost is skipped — see the accounting block.
        from ..services.journal_preview import invoice_gl_fingerprint  # local: cycle
        gl_before             = invoice_gl_fingerprint(invoice)
        old_scalars           = _snapshot_scalars(invoice)
        old_grand_total       = invoice.grand_total
        old_payment_method_id = invoice.payment_method_id
        old_warehouse_id      = invoice.warehouse_id
        old_patient           = invoice.patient_no
        old_item_instances    = list(
            invoice.items
            .select_related(
                'item__item_category__revenue_account',
            )
            .prefetch_related('item__treatment')
            .all()
        )

        # ── Scalar fields ─────────────────────────────────────────────────────

        if 'datetime' in data:
            invoice.datetime = data['datetime']
            changes.append('datetime')

        if 'payment_method_id' in data:
            if data['payment_method_id']:
                try:
                    invoice.payment_method = PaymentMethod.objects.get(
                        id=data['payment_method_id'],
                        is_active=True,
                    )
                except PaymentMethod.DoesNotExist:
                    return Response({'payment_method_id': 'Payment method not found.'}, status=status.HTTP_400_BAD_REQUEST)
                if 'payment_account_id' not in data:
                    # Mirrors InvoiceCreateView: a caller that only sends
                    # payment_method_id still gets payment_account kept in step.
                    invoice.payment_account = invoice.payment_method.linked_account
                    changes.append('payment_account')
            else:
                invoice.payment_method = None
            changes.append('payment_method')

        if 'payment_account_id' in data:
            if data['payment_account_id']:
                if data['payment_account_id'] not in cash_bank_account_ids():
                    return Response({'payment_account_id': ['Not a cash/bank account.']}, status=status.HTTP_400_BAD_REQUEST)
                try:
                    invoice.payment_account = ChartOfAccounts.objects.get(pk=data['payment_account_id'])
                except ChartOfAccounts.DoesNotExist:
                    return Response({'payment_account_id': ['Not a cash/bank account.']}, status=status.HTTP_400_BAD_REQUEST)
            else:
                invoice.payment_account = None
            changes.append('payment_account')

        for field in ('discount', 'tax', 'additional_charges', 'grand_total'):
            if field in data:
                setattr(invoice, field, data[field])
                changes.append(field)

        # ── Patient ───────────────────────────────────────────────────────────

        if 'patient_no' in data:
            if data['patient_no']:
                try:
                    invoice.patient_no = Patient.objects.get(patient_no=data['patient_no'])
                except Patient.DoesNotExist:
                    return Response({'patient_no': 'Patient not found.'}, status=status.HTTP_400_BAD_REQUEST)
            else:
                invoice.patient_no = None
            changes.append('patient_no')

        # ── Cashier ───────────────────────────────────────────────────────────

        if 'cashier_id' in data:
            if data['cashier_id']:
                try:
                    invoice.cashier = AppUser.objects.get(id=data['cashier_id'])
                except AppUser.DoesNotExist:
                    return Response({'cashier_id': 'Cashier not found.'}, status=status.HTTP_400_BAD_REQUEST)
            else:
                invoice.cashier = None
            changes.append('cashier')

        # ── Warehouse ─────────────────────────────────────────────────────────

        if 'warehouse_id' in data:
            if data['warehouse_id']:
                try:
                    invoice.warehouse = Warehouse.objects.get(id=data['warehouse_id'])
                except Warehouse.DoesNotExist:
                    return Response({'warehouse_id': 'Warehouse not found.'}, status=status.HTTP_400_BAD_REQUEST)
            else:
                invoice.warehouse = None
            changes.append('warehouse')

        invoice.save()

        # ── Line items (replace-all strategy) ────────────────────────────────

        new_line_items_data = None  # None = items unchanged

        if replacing_items:
            invoice.items.all().delete()
            InvoiceItem.objects.bulk_create([
                InvoiceItem(
                    invoice=invoice,
                    item=new_items_by_id.get(i['item_id']) if i.get('item_id') else None,
                    item_name=i.get('item_name', '') if not i.get('item_id') else '',
                    quantity=i['quantity'],
                    price=i['price'],
                    discount_pct=i.get('discount_pct', 0),
                )
                for i in data['items']
            ])
            new_line_items_data = data['items']
            changes.append('items')

        # ── Accounting: reverse old entries, re-apply for new state ──────────
        # Phase 2: an invoice whose transaction date has never been posted
        # (posting_status='unposted') carries zero LedgerEntry rows — editing
        # it just updates the record; the next journal run will post the new
        # values. An already-posted invoice (its date's JournalDayLog.is_posted
        # =True) instead gets a same-day "edit memo": a reversal of the OLD
        # posted values plus a fresh posting of the NEW values, both dated
        # today and tagged source_type='edit_memo'. The original transaction
        # date's LedgerEntry rows and JournalDayLog are left untouched.
        #
        # The memo pair is written only when the edit actually moved something
        # the posting reads. Without this check every PATCH appended a full
        # reversal + repost — eight-odd rows that net to zero but pile up on the
        # invoice's journal view, so a handful of re-saves made one sale look
        # like it had been journalled half a dozen times. `payments` is compared
        # against the rows about to replace it rather than the ones still in the
        # table, since the swap below deliberately happens after the reversal.
        after_splits = (
            [(r.payment_account_id, r.payment_method_id, r.amount, r.sort_order)
             for r in new_split_rows]
            if replacing_payments else None
        )
        gl_changed = invoice_gl_fingerprint(invoice, splits=after_splits) != gl_before

        was_posted = invoice.posting_status == 'posted' and gl_changed
        memo_date = timezone.now().date() if was_posted else None
        memo_source = 'edit_memo' if was_posted else 'invoice'

        if was_posted:
            _reverse_accounting_instances(
                old_payment_method_id, old_grand_total, old_item_instances, old_warehouse_id,
                invoice=invoice, old_scalars=old_scalars,
                memo_date=memo_date, source_type=memo_source,
            )

        # Old legs are unwound; the new split can go in now. An edit that leaves
        # `payments` out keeps whatever rows the invoice already had — if it also
        # changed grand_total those rows no longer sum to it, and the engine puts
        # the difference on the invoice-level payment account rather than
        # silently misstating the split.
        if replacing_payments:
            invoice.payments.all().delete()
            for row in new_split_rows:
                row.invoice = invoice
            InvoicePayment.objects.bulk_create(new_split_rows)
            if new_tenders:
                first = new_tenders[0]
                if 'payment_method_id' not in data and first.payment_method_id:
                    invoice.payment_method = first.payment_method
                if 'payment_account_id' not in data and first.payment_account_id:
                    invoice.payment_account = first.payment_account
                invoice.save(update_fields=['payment_method', 'payment_account'])
            changes.append('payments')

        if new_line_items_data is not None:
            _reverse_packages(invoice)
            if was_posted:
                _post_accounting(invoice, new_line_items_data, new_items_by_id,
                                  memo_date=memo_date, source_type=memo_source)
            _handle_packages(invoice, invoice.patient_no, new_line_items_data, new_items_by_id)
        else:
            if was_posted:
                # Items unchanged — re-apply using old instances converted to dict format
                old_lines_as_dicts = [
                    {'item_id': inst.item_id, 'item_name': inst.item_name,
                     'quantity': inst.quantity, 'price': inst.price}
                    for inst in old_item_instances
                ]
                old_items_by_id = {
                    inst.item_id: inst.item
                    for inst in old_item_instances
                    if inst.item_id
                }
                _post_accounting(invoice, old_lines_as_dicts, old_items_by_id,
                                  memo_date=memo_date, source_type=memo_source)

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='Invoice',
            entity_id=str(invoice.id),
            description=f'Invoice {invoice.invoice_number} updated — fields changed: {", ".join(changes)}',
        )

        # ── CRM refresh ───────────────────────────────────────────────────────
        # Both patients when the invoice moved between them.
        for p in {old_patient, invoice.patient_no}:
            if p is not None:
                refresh_crm_profile(p)

        invoice.refresh_from_db()
        return Response(InvoiceReadSerializer(
            Invoice.objects
            .select_related('patient_no', 'cashier', 'warehouse', 'payment_method', 'payment_account')
            .prefetch_related('items__item', 'payments__payment_method', 'payments__payment_account')
            .get(pk=pk)
        ).data)

    @transaction.atomic
    def delete(self, request, pk):
        """Void (soft-delete) an invoice: reverse its accounting but keep the record."""
        invoice = self._get(pk)
        if invoice is None:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if invoice.is_voided:
            return Response(
                {'error': 'Invoice is already voided.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        old_item_instances = list(
            invoice.items
            .select_related(
                'item__item_category__revenue_account',
            )
            .prefetch_related('item__treatment')
            .all()
        )

        inv_no = invoice.invoice_number
        patient = invoice.patient_no

        # Phase 2: an unposted invoice has no LedgerEntry rows to reverse — void
        # it with zero journal impact (the next journal run will simply skip it,
        # see the sweep query's posting_status filter). A posted invoice gets a
        # same-day reversing "void memo" instead of touching its original
        # transaction date: dated today, tagged source_type='void_memo',
        # referencing this invoice. The original day's JournalDayLog is left
        # untouched either way.
        was_posted = invoice.posting_status == 'posted'
        if was_posted:
            _reverse_accounting_instances(
                invoice.payment_method_id,
                invoice.grand_total,
                old_item_instances,
                invoice.warehouse_id,
                invoice=invoice,
                memo_date=timezone.now().date(),
                source_type='void_memo',
            )

        invoice.is_voided = True
        invoice.voided_at = timezone.now()
        invoice.voided_by = _actor(request)
        invoice.save(update_fields=['is_voided', 'voided_at', 'voided_by'])

        if patient is not None:
            refresh_crm_profile(patient)

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='VOID',
            entity_type='Invoice',
            entity_id=str(pk),
            description=f'Invoice {inv_no} voided — accounting reversed',
        )

        return Response(InvoiceReadSerializer(self._get(pk)).data)


class InvoiceExportView(APIView):
    """GET /api/invoices/export/  — download filtered invoices as .xlsx"""
    permission_classes = [AllowAny]

    def get(self, request):
        qs = (
            Invoice.objects
            .select_related('patient_no', 'cashier', 'warehouse', 'payment_method', 'payment_account')
            .prefetch_related('items__item', 'payments__payment_method', 'payments__payment_account')
            .order_by('-datetime')
        )
        if q := request.GET.get('q', '').strip():
            from django.db.models import Q
            qs = qs.filter(
                Q(invoice_number__icontains=q) |
                Q(patient_no__name__icontains=q) |
                Q(patient_no__patient_no__icontains=q)
            )
        if method := request.GET.get('payment_method', '').strip():
            qs = qs.filter(payment_method_id=method)
        if account := request.GET.get('payment_account', '').strip():
            qs = qs.filter(payment_account_id=account)

        # Hide voided invoices unless explicitly requested (mirrors the list view).
        if request.GET.get('include_voided', '').strip().lower() not in ('1', 'true', 'yes'):
            qs = qs.filter(is_voided=False)

        wb = openpyxl.Workbook()
        ws_inv  = wb.active
        ws_inv.title = 'Invoices'
        ws_item = wb.create_sheet('Items')

        ws_inv.append(INV_HEADERS)
        ws_item.append(ITEM_HEADERS)

        for inv in qs:
            ws_inv.append([
                inv.invoice_number,
                inv.datetime.strftime('%Y-%m-%dT%H:%M:%S') if inv.datetime else '',
                inv.patient_no_id or '',
                inv.payment_method.name if inv.payment_method else '',
                float(inv.discount),
                float(inv.tax),
                float(inv.additional_charges),
                float(inv.grand_total),
                inv.cashier_id or '',
                inv.warehouse.name if inv.warehouse else '',
            ])
            for item in inv.items.all():
                ws_item.append([
                    inv.invoice_number,
                    item.item.code,
                    item.item.name,
                    float(item.quantity),
                    float(item.price),
                ])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"invoices_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class InvoiceImportView(APIView):
    """POST /api/invoices/import/  — upload .xlsx, create invoices that don't exist yet"""
    permission_classes = [AllowAny]

    @transaction.atomic
    def post(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(upload, read_only=True, data_only=True)
        except Exception as e:
            return Response({'error': f'Cannot read workbook: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        if 'Invoices' not in wb.sheetnames or 'Items' not in wb.sheetnames:
            return Response(
                {'error': 'Workbook must contain sheets named "Invoices" and "Items".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws_inv  = wb['Invoices']
        ws_item = wb['Items']

        # Parse Invoices sheet (skip header row 1)
        inv_rows = []
        for row in ws_inv.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            inv_rows.append({
                'invoice_number':    str(row[0]).strip(),
                'datetime':          str(row[1]).strip() if row[1] else '',
                'patient_no':        str(row[2]).strip() if row[2] else '',
                'payment_method':    str(row[3]).strip() if row[3] else '',
                'discount':          _to_dec(row[4]),
                'tax':               _to_dec(row[5]),
                'additional_charges':_to_dec(row[6]),
                'grand_total':       _to_dec(row[7]),
                'cashier_id':        str(row[8]).strip() if row[8] else '',
                'warehouse_id':      str(row[9]).strip() if row[9] else '',
            })

        # Parse Items sheet keyed by invoice_number
        items_by_inv: dict[str, list] = {}
        for row in ws_item.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            inv_no = str(row[0]).strip()
            items_by_inv.setdefault(inv_no, []).append({
                'item_code': str(row[1]).strip() if row[1] else '',
                'quantity':  _to_dec(row[3]),
                'price':     _to_dec(row[4]),
            })

        created = []
        skipped = []
        errors  = []

        for row in inv_rows:
            inv_no = row['invoice_number']

            if Invoice.objects.filter(invoice_number=inv_no).exists():
                skipped.append(inv_no)
                continue

            # Resolve optional FKs — soft-fail: skip unknown references
            patient = None
            if row['patient_no']:
                patient = Patient.objects.filter(patient_no=row['patient_no']).first()

            cashier = None
            if row['cashier_id'] and str(row['cashier_id']).isdigit():
                cashier = AppUser.objects.filter(id=int(row['cashier_id'])).first()

            warehouse = None
            if row['warehouse_id']:
                warehouse = Warehouse.objects.filter(name=row['warehouse_id']).first()

            payment = PaymentMethod.objects.filter(name=row['payment_method']).first()

            try:
                dt = timezone.datetime.fromisoformat(row['datetime']) if row['datetime'] else timezone.now()
                if timezone.is_naive(dt):
                    dt = timezone.make_aware(dt)
            except ValueError:
                dt = timezone.now()

            line_dicts = items_by_inv.get(inv_no, [])
            if not line_dicts:
                errors.append(f'{inv_no}: no items found in Items sheet')
                continue

            # Resolve item codes to InventoryItem objects
            resolved_lines = []
            row_error = False
            for ld in line_dicts:
                item_obj = InventoryItem.objects.filter(code=ld['item_code']).first()
                if item_obj is None:
                    errors.append(f'{inv_no}: item code "{ld["item_code"]}" not found — row skipped')
                    row_error = True
                    break
                resolved_lines.append((item_obj, ld['quantity'], ld['price']))

            if row_error:
                continue

            invoice = Invoice(
                datetime=dt,
                patient_no=patient,
                payment_method=payment,
                discount=row['discount'],
                cashier=cashier,
                warehouse=warehouse,
                tax=row['tax'],
                additional_charges=row['additional_charges'],
                grand_total=row['grand_total'],
            )
            invoice.invoice_number = inv_no  # preserve original number
            invoice.save()

            InvoiceItem.objects.bulk_create([
                InvoiceItem(invoice=invoice, item=item_obj, quantity=qty, price=price)
                for item_obj, qty, price in resolved_lines
            ])

            AuditLog.objects.create(
                performed_by=_actor(request),
                action='IMPORT',
                entity_type='Invoice',
                entity_id=str(invoice.id),
                description=f'Invoice {inv_no} imported via Excel upload',
            )
            created.append(inv_no)

        return Response({
            'created': len(created),
            'skipped': len(skipped),
            'errors':  errors,
            'created_numbers': created,
            'skipped_numbers': skipped,
        }, status=status.HTTP_200_OK)


def _to_dec(val) -> Decimal:
    try:
        return Decimal(str(val)).quantize(Decimal('0.01'))
    except (InvalidOperation, TypeError):
        return Decimal('0.00')
