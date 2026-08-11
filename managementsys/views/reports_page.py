import datetime
import io
import os
from decimal import Decimal
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from django.db.models import Count, DecimalField, F, Sum, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.negotiation import DefaultContentNegotiation
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView


class IgnoreFormatQueryNegotiation(DefaultContentNegotiation):
    """
    Content negotiation that ignores the reserved ``?format=`` query parameter.

    DRF's default negotiation treats ``?format=xxx`` as a renderer selector and
    raises 404 when no renderer matches (e.g. ``?format=xlsx``). This view uses
    ``?output=`` to pick json vs. xlsx itself, so we always return the first
    (JSON) renderer and let the view decide the real output.
    """

    def select_renderer(self, request, renderers, format_suffix=None):
        renderer = renderers[0]
        return renderer, renderer.media_type

from ..models import (
    InventoryBatch,
    InventoryItem,
    Invoice,
    InvoiceItem,
    InvoicePayment,
    PurchaseInvoice,
    SiteConfig,
    StockOutLog,
)

_JAKARTA = ZoneInfo('Asia/Jakarta')


def payment_method_breakdown(invoices):
    """Sales per payment method over ``invoices``, honouring split payments.

    Grouping straight on ``Invoice.payment_method`` credits a split invoice's
    whole grand total to whichever method the cashier picked first — the cash
    drawer then reads high by the part that went to the bank, which is exactly
    what the cashier is trying to reconcile. So: invoices with no split rows are
    summed on their own method as before, and split invoices are summed from
    their ``InvoicePayment`` rows instead.

    ``invoice_count`` counts an invoice once per method it touched, so the counts
    across methods can exceed the number of invoices in the range.

    Returns ``[{'payment_method_id', 'method', 'account_number', 'total',
    'invoice_count'}]``, largest first, with ``total`` as a Decimal.
    """
    split_invoice_ids = set(
        InvoicePayment.objects
        .filter(invoice__in=invoices)
        .values_list('invoice_id', flat=True)
    )

    rows = {}

    def add(method_id, name, account_number, total, count):
        row = rows.setdefault(method_id, {
            'payment_method_id': method_id,
            'method': name or 'Tidak Diketahui',
            'account_number': account_number,
            'total': Decimal('0'),
            'invoice_count': 0,
        })
        row['total'] += total or Decimal('0')
        row['invoice_count'] += count

    unsplit = (
        invoices.exclude(id__in=split_invoice_ids)
        .values('payment_method_id', 'payment_method__name',
                'payment_method__linked_account__account_number')
        .annotate(total=Sum('grand_total'), invoice_count=Count('id'))
    )
    for r in unsplit:
        add(r['payment_method_id'], r['payment_method__name'],
            r['payment_method__linked_account__account_number'],
            r['total'], r['invoice_count'])

    split = (
        InvoicePayment.objects
        .filter(invoice__in=invoices)
        .values('payment_method_id', 'payment_method__name',
                'payment_method__linked_account__account_number')
        .annotate(total=Sum('amount'), invoice_count=Count('invoice_id', distinct=True))
    )
    for r in split:
        add(r['payment_method_id'], r['payment_method__name'],
            r['payment_method__linked_account__account_number'],
            r['total'], r['invoice_count'])

    return sorted(rows.values(), key=lambda r: r['total'], reverse=True)


class DashboardReportView(APIView):
    def get(self, request):
        now = timezone.now()
        # Use Jakarta local time so "today" and "this month" match clinic hours.
        local_now = now.astimezone(_JAKARTA)

        today_start = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
        month_start = local_now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        if local_now.month == 1:
            last_month_start = local_now.replace(
                year=local_now.year - 1, month=12, day=1,
                hour=0, minute=0, second=0, microsecond=0,
            )
        else:
            last_month_start = month_start.replace(month=month_start.month - 1)

        inv_today = Invoice.objects.filter(datetime__gte=today_start)
        inv_month = Invoice.objects.filter(datetime__gte=month_start)
        inv_last_month = Invoice.objects.filter(
            datetime__gte=last_month_start, datetime__lt=month_start
        )

        today_agg = inv_today.aggregate(total=Sum('grand_total'), count=Count('id'))
        month_agg = inv_month.aggregate(total=Sum('grand_total'), count=Count('id'))
        last_month_agg = inv_last_month.aggregate(total=Sum('grand_total'), count=Count('id'))

        payment_breakdown = payment_method_breakdown(inv_month)

        items_qs = InventoryItem.objects.filter(is_service=False, is_active=True).annotate(
            total_stock=Coalesce(Sum('batches__quantity_remaining'), Value(0, output_field=DecimalField()))
        )
        low_stock_items = list(
            items_qs
            .filter(total_stock__lt=F('min_stock'))
            .values('id', 'code', 'name', 'unit_small', 'min_stock', 'total_stock')
            .order_by('name')
        )

        recent = Invoice.objects.select_related('patient_no', 'payment_method').order_by('-datetime')[:10]
        recent_data = [
            {
                'invoice_number': inv.invoice_number,
                'datetime': inv.datetime,
                'patient_name': inv.patient_no.name if inv.patient_no else None,
                'grand_total': str(inv.grand_total),
                'payment_method_id': inv.payment_method_id,
                'payment_method_name': inv.payment_method.name if inv.payment_method_id else None,
            }
            for inv in recent
        ]

        return Response({
            'revenue': {
                'today_total': str(today_agg['total'] or 0),
                'today_count': today_agg['count'] or 0,
                'this_month_total': str(month_agg['total'] or 0),
                'this_month_count': month_agg['count'] or 0,
                'last_month_total': str(last_month_agg['total'] or 0),
                'last_month_count': last_month_agg['count'] or 0,
                'by_payment_method': [
                    {
                        'payment_method_id': r['payment_method_id'],
                        'method': r['method'],
                        'total': str(r['total'] or 0),
                        'count': r['invoice_count'],
                    }
                    for r in payment_breakdown
                ],
            },
            'inventory': {
                'total_active_items': items_qs.count(),
                'low_stock_count': len(low_stock_items),
                'low_stock_items': low_stock_items,
            },
            'recent_invoices': recent_data,
        })


class SalesRangeReportView(APIView):
    """
    GET /api/reports/sales/?start=YYYY-MM-DD&end=YYYY-MM-DD
    Total sales and breakdown per cash (payment) account over an inclusive date
    range. Defaults to the current month-to-date in Jakarta time.
    """

    def get(self, request):
        today = timezone.now().astimezone(_JAKARTA).date()

        def parse(param, default):
            raw = request.query_params.get(param, '').strip()
            if not raw:
                return default, None
            try:
                return datetime.date.fromisoformat(raw), None
            except ValueError:
                return None, f'Tanggal {param} tidak valid. Gunakan format YYYY-MM-DD.'

        start, err = parse('start', today.replace(day=1))
        if err:
            return Response({'error': err}, status=status.HTTP_400_BAD_REQUEST)
        end, err = parse('end', today)
        if err:
            return Response({'error': err}, status=status.HTTP_400_BAD_REQUEST)

        if start > end:
            return Response(
                {'error': 'Tanggal mulai tidak boleh setelah tanggal akhir.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # `datetime` is stored in UTC; bound the range by Jakarta-local midnights
        # so a day belongs to the range the clinic actually worked it.
        range_start = datetime.datetime(start.year, start.month, start.day, tzinfo=_JAKARTA)
        range_end = datetime.datetime(end.year, end.month, end.day, tzinfo=_JAKARTA) + datetime.timedelta(days=1)

        invoices = Invoice.objects.filter(datetime__gte=range_start, datetime__lt=range_end)
        agg = invoices.aggregate(total=Sum('grand_total'), count=Count('id'))

        breakdown = payment_method_breakdown(invoices)

        return Response({
            'start': str(start),
            'end': str(end),
            'total': str(agg['total'] or 0),
            'invoice_count': agg['count'] or 0,
            'by_account': [
                {
                    'account_id': r['payment_method_id'],
                    'account_number': r['account_number'],
                    'account_name': r['method'],
                    'total': str(r['total'] or 0),
                    'invoice_count': r['invoice_count'],
                }
                for r in breakdown
            ],
        })


_ID_MONTHS = [
    '', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
    'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember',
]


def _month_label(year, month):
    return f'{_ID_MONTHS[month]} {year}'


class SalesItemsBreakdownView(APIView):
    """
    GET /api/reports/sales-items/?start=YYYY-MM-DD&end=YYYY-MM-DD

    Item-level sales analytics over an inclusive date range, grouped by calendar
    month (Jakarta time). For each month and for the whole range it reports the
    revenue/quantity split per category, the top treatment, the top physical
    item, and the leading category. Defaults to the current month-to-date.
    """

    def get(self, request):
        today = timezone.now().astimezone(_JAKARTA).date()

        def parse(param, default):
            raw = request.query_params.get(param, '').strip()
            if not raw:
                return default, None
            try:
                return datetime.date.fromisoformat(raw), None
            except ValueError:
                return None, f'Tanggal {param} tidak valid. Gunakan format YYYY-MM-DD.'

        start, err = parse('start', today.replace(day=1))
        if err:
            return Response({'error': err}, status=status.HTTP_400_BAD_REQUEST)
        end, err = parse('end', today)
        if err:
            return Response({'error': err}, status=status.HTTP_400_BAD_REQUEST)
        if start > end:
            return Response(
                {'error': 'Tanggal mulai tidak boleh setelah tanggal akhir.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        range_start = datetime.datetime(start.year, start.month, start.day, tzinfo=_JAKARTA)
        range_end = datetime.datetime(end.year, end.month, end.day, tzinfo=_JAKARTA) + datetime.timedelta(days=1)

        lines = (
            InvoiceItem.objects
            .select_related('item', 'item__item_category', 'invoice')
            .filter(
                invoice__datetime__gte=range_start,
                invoice__datetime__lt=range_end,
                invoice__is_voided=False,
            )
        )

        # month key -> aggregation buckets
        months = {}

        def new_month(year, month):
            return {
                'month': f'{year:04d}-{month:02d}',
                'label': _month_label(year, month),
                'revenue': 0.0,
                'qty': 0.0,
                'line_count': 0,
                'categories': {},    # name -> {revenue, qty, line_count}
                'treatments': {},    # key  -> {name, revenue, qty}
                'items': {},         # key  -> {name, revenue, qty}
            }

        def accumulate(bucket, ii, is_service, cat_name, name, key, revenue, qty):
            bucket['revenue'] += revenue
            bucket['qty'] += qty
            bucket['line_count'] += 1

            cat = bucket['categories'].get(cat_name)
            if cat is None:
                cat = {'revenue': 0.0, 'qty': 0.0, 'line_count': 0}
                bucket['categories'][cat_name] = cat
            cat['revenue'] += revenue
            cat['qty'] += qty
            cat['line_count'] += 1

            target = bucket['treatments'] if is_service else bucket['items']
            entry = target.get(key)
            if entry is None:
                entry = {'name': name, 'category': cat_name, 'revenue': 0.0, 'qty': 0.0}
                target[key] = entry
            entry['revenue'] += revenue
            entry['qty'] += qty

        overall = new_month(0, 0)
        overall['month'] = None
        overall['label'] = 'Total'

        for ii in lines:
            revenue = float(ii.price) * float(ii.quantity) * (1 - float(ii.discount_pct) / 100)
            qty = float(ii.quantity)

            if ii.item_id:
                is_service = bool(ii.item.is_service)
                name = ii.item.name
                key = ii.item_id
                if ii.item.item_category_id:
                    cat_name = ii.item.item_category.name
                else:
                    cat_name = ii.item.category or 'Tanpa Kategori'
            else:
                is_service = False
                name = ii.item_name or 'Item Lain'
                key = f'name:{name}'
                cat_name = 'Tanpa Kategori'

            local_dt = ii.invoice.datetime.astimezone(_JAKARTA)
            mkey = (local_dt.year, local_dt.month)
            bucket = months.get(mkey)
            if bucket is None:
                bucket = new_month(local_dt.year, local_dt.month)
                months[mkey] = bucket

            accumulate(bucket, ii, is_service, cat_name, name, key, revenue, qty)
            accumulate(overall, ii, is_service, cat_name, name, key, revenue, qty)

        def finalize(bucket):
            categories = sorted(
                (
                    {
                        'name': cname,
                        'revenue': round(c['revenue'], 2),
                        'qty': round(c['qty'], 3),
                        'line_count': c['line_count'],
                    }
                    for cname, c in bucket['categories'].items()
                ),
                key=lambda c: c['revenue'],
                reverse=True,
            )

            def top(entries, metric):
                if not entries:
                    return None
                best = max(entries.values(), key=lambda e: e[metric])
                return {
                    'name': best['name'],
                    'category': best.get('category', ''),
                    'qty': round(best['qty'], 3),
                    'revenue': round(best['revenue'], 2),
                }

            return {
                'month': bucket['month'],
                'label': bucket['label'],
                'total_revenue': round(bucket['revenue'], 2),
                'total_qty': round(bucket['qty'], 3),
                'line_count': bucket['line_count'],
                'by_category': categories,
                'top_category': categories[0] if categories else None,
                'top_treatment': top(bucket['treatments'], 'qty'),
                'top_item': top(bucket['items'], 'qty'),
            }

        months_out = [finalize(months[k]) for k in sorted(months.keys())]

        return Response({
            'start': str(start),
            'end': str(end),
            'overall': finalize(overall),
            'months': months_out,
        })


# ── Generate Report (parametrized, JSON preview + .xlsx download) ─────────────

_REPORT_TITLES = {
    'stock_in':      'Laporan Daftar Item Masuk',
    'stock_out':     'Laporan Daftar Item Keluar',
    'purchases':     'Laporan Pembelian Detail',
    'sales_by_item': 'Daftar Penjualan Per Item Per Jenis',
}

_STOCK_COLUMNS = [
    'No Transaksi', 'Dept.', 'Tanggal', 'Kode Item',
    'Nama Item', 'Jumlah', 'Satuan', 'Keterangan',
]
_PURCHASE_LINE_COLUMNS = [
    'No.', 'Kd. Item', 'Nama Item', 'Jml', 'Satuan', 'Harga', 'Pot. %', 'Total',
]
_SALES_COLUMNS = [
    'Kode Item', 'Nama Item', 'Jenis', 'Jumlah', 'Satuan', 'Total Harga',
]


def _parse_report_range(request):
    """Return (start_date, end_date, error_str). Defaults to month-to-date (Jakarta)."""
    today = timezone.now().astimezone(_JAKARTA).date()

    def parse(param, default):
        raw = request.query_params.get(param, '').strip()
        if not raw:
            return default, None
        try:
            return datetime.date.fromisoformat(raw), None
        except ValueError:
            return None, f'Tanggal {param} tidak valid. Gunakan format YYYY-MM-DD.'

    start, err = parse('start', today.replace(day=1))
    if err:
        return None, None, err
    end, err = parse('end', today)
    if err:
        return None, None, err
    if start > end:
        return None, None, 'Tanggal mulai tidak boleh setelah tanggal akhir.'
    return start, end, None


def _build_stock_in(start, end):
    qs = (
        InventoryBatch.objects
        .select_related('item', 'warehouse', 'purchase_invoice')
        .filter(input_date__gte=start, input_date__lte=end)
        .order_by('input_date', 'id')
    )
    rows = []
    for b in qs:
        if b.purchase_invoice_id:
            no_trx = b.purchase_invoice.internal_id
            keterangan = b.purchase_invoice.notes or ''
        else:
            no_trx = f'IB-{b.id}'
            keterangan = ''
        rows.append([
            no_trx,
            b.warehouse.code if b.warehouse_id else '',
            b.input_date.isoformat(),
            b.item.code if b.item_id else '',
            b.item.name if b.item_id else '',
            float(b.quantity_initial),
            b.item.unit_small if b.item_id else '',
            keterangan,
        ])
    return {'kind': 'flat', 'columns': _STOCK_COLUMNS, 'rows': rows}


def _build_stock_out(start, end):
    qs = (
        StockOutLog.objects
        .select_related('item', 'warehouse')
        .filter(out_date__gte=start, out_date__lte=end)
        .order_by('out_date', 'id')
    )
    rows = []
    for log in qs:
        rows.append([
            f'SO-{log.id}',
            log.warehouse.code if log.warehouse_id else '',
            log.out_date.isoformat(),
            log.item.code if log.item_id else '',
            log.item.name if log.item_id else '',
            float(log.quantity),
            log.item.unit_small if log.item_id else '',
            log.notes or '',
        ])
    return {'kind': 'flat', 'columns': _STOCK_COLUMNS, 'rows': rows}


def _build_purchases(start, end):
    qs = (
        PurchaseInvoice.objects
        .select_related('supplier', 'warehouse')
        .prefetch_related('items__item')
        .filter(purchase_date__gte=start, purchase_date__lte=end, is_voided=False)
        .order_by('purchase_date', 'id')
    )
    groups = []
    for inv in qs:
        lines = []
        for idx, it in enumerate(inv.items.all(), start=1):
            qty = float(it.quantity)
            unit_cost = float(it.unit_cost)
            discount = float(it.total_discount)
            gross = qty * unit_cost
            total = gross - discount
            pot_pct = (discount / gross * 100) if gross else 0.0
            lines.append([
                idx,
                it.item.code if it.item_id else '',
                it.item_name,
                qty,
                it.unit or '',
                unit_cost,
                round(pot_pct, 2),
                round(total, 2),
            ])
        groups.append({
            'header': {
                'No Transaksi': inv.internal_id,
                'Tanggal': inv.purchase_date.isoformat(),
                'Dept.': inv.warehouse.code if inv.warehouse_id else '',
                'Kode Supp.': inv.supplier_id,
                'Nama Supplier': inv.supplier.name if inv.supplier_id else '',
            },
            'lines': lines,
        })
    return {'kind': 'grouped', 'line_columns': _PURCHASE_LINE_COLUMNS, 'groups': groups}


def _build_sales_by_item(start, end):
    range_start = datetime.datetime(start.year, start.month, start.day, tzinfo=_JAKARTA)
    range_end = datetime.datetime(end.year, end.month, end.day, tzinfo=_JAKARTA) + datetime.timedelta(days=1)

    items = (
        InvoiceItem.objects
        .select_related('item', 'item__item_category', 'invoice')
        .filter(
            invoice__datetime__gte=range_start,
            invoice__datetime__lt=range_end,
            invoice__is_voided=False,
        )
    )

    agg = {}
    for ii in items:
        line_total = float(ii.price) * float(ii.quantity) * (1 - float(ii.discount_pct) / 100)
        if ii.item_id:
            key = ii.item_id
            code = ii.item.code
            name = ii.item.name
            if ii.item.item_category_id:
                jenis = ii.item.item_category.name
            else:
                jenis = ii.item.category or ''
            unit = ii.item.unit_small
        else:
            key = f'name:{ii.item_name}'
            code = ''
            name = ii.item_name
            jenis = ''
            unit = ''
        bucket = agg.get(key)
        if bucket is None:
            bucket = {'code': code, 'name': name, 'jenis': jenis, 'unit': unit, 'qty': 0.0, 'total': 0.0}
            agg[key] = bucket
        bucket['qty'] += float(ii.quantity)
        bucket['total'] += line_total

    rows = [
        [b['code'], b['name'], b['jenis'], b['qty'], b['unit'], round(b['total'], 2)]
        for b in sorted(agg.values(), key=lambda b: (b['jenis'].lower(), b['name'].lower()))
    ]
    return {'kind': 'flat', 'columns': _SALES_COLUMNS, 'rows': rows}


_REPORT_BUILDERS = {
    'stock_in':      _build_stock_in,
    'stock_out':     _build_stock_out,
    'purchases':     _build_purchases,
    'sales_by_item': _build_sales_by_item,
}


_CENTER = Alignment(horizontal='center', vertical='center')
_RIGHT = Alignment(horizontal='right', vertical='center')
_LEFT = Alignment(horizontal='left', vertical='center')
_THIN = Side(style='thin', color='D0D5DD')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FILL = PatternFill('solid', fgColor='EEF2F6')
_HDR_FONT = Font(bold=True, size=10)


# Logo is bounded to this box (px) and scaled with its aspect ratio preserved.
_LOGO_MAX_W = 130
_LOGO_MAX_H = 78
# Approx px width of one indent step (Calibri 11 "0"), used to push the header
# text just past the logo without depending on data-driven column widths.
_INDENT_PX = 7


def _write_banner(ws, title, period_label, ncols):
    """
    Write the logo + clinic identity + title + period block, left-aligned.

    The text is one full-width merged cell per line, indented just past the logo,
    so it sits tight against the logo regardless of how wide the data columns are.
    Returns (next_free_row, has_logo).
    """
    cfg = SiteConfig.get_solo()
    last_col = get_column_letter(max(ncols, 1))

    # Place + size the logo first so we know how far to indent the text.
    indent = 0
    # `.path` raises ValueError on an ImageField with no file (and on non-local
    # storage), so a bare getattr() default is not enough of a guard here.
    logo = getattr(cfg, 'logo', None)
    logo_path = None
    if logo:
        try:
            logo_path = logo.path
        except (ValueError, NotImplementedError):
            logo_path = None
    has_logo = bool(logo_path and os.path.exists(logo_path))
    if has_logo:
        try:
            img = XLImage(logo_path)
            if img.width and img.height:
                scale = min(_LOGO_MAX_W / img.width, _LOGO_MAX_H / img.height)
                img.width = int(img.width * scale)
                img.height = int(img.height * scale)
            ws.add_image(img, 'A1')
            indent = round(img.width / _INDENT_PX) + 1  # clear the logo, tight
        except Exception:
            has_logo = False  # bad logo: fall back to full-width text

    def line(text, font, row):
        ws.merge_cells(f'A{row}:{last_col}{row}')
        cell = ws.cell(row, 1, text)
        cell.font = font
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=indent)
        return row + 1

    r = 1
    r = line(title, Font(size=14, bold=True), r)
    if cfg.clinic_name:
        r = line(cfg.clinic_name.upper(), Font(size=12, bold=True), r)
    address = ', '.join(p for p in (cfg.address_line1, cfg.address_line2) if p)
    if address:
        r = line(address, Font(size=10), r)
    if cfg.phone_fax:
        r = line(cfg.phone_fax, Font(size=10), r)
    r = line(f'PERIODE : {period_label}', Font(size=10, italic=True), r)

    return r + 1, has_logo  # blank spacer row after the banner


def _write_table_header(ws, row, columns):
    for i, label in enumerate(columns, start=1):
        cell = ws.cell(row, i, label)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.border = _BORDER
        cell.alignment = _CENTER
    return row + 1


def _write_data_row(ws, row, values):
    for i, v in enumerate(values, start=1):
        cell = ws.cell(row, i, v)
        cell.border = _BORDER
        cell.alignment = _RIGHT if isinstance(v, (int, float)) else _LEFT
    return row + 1


def _autosize(ws, columns, rows):
    """Fit each column to the widest of its header label and data cells."""
    for i, label in enumerate(columns):
        letter = get_column_letter(i + 1)
        longest = len(str(label))
        for row in rows:
            if i >= len(row):
                continue
            v = row[i]
            if v is None:
                continue
            if isinstance(v, float):
                text = f'{v:,.2f}'.rstrip('0').rstrip('.') if v % 1 else f'{int(v):,}'
            elif isinstance(v, int):
                text = f'{v:,}'
            else:
                text = str(v)
            longest = max(longest, len(text))
        ws.column_dimensions[letter].width = min(max(longest + 2, 9), 48)


def _report_to_xlsx(report, title, period_label, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Laporan'
    ws.sheet_view.showGridLines = False

    columns = report['columns'] if report['kind'] == 'flat' else report['line_columns']
    ncols = len(columns)
    row, has_logo = _write_banner(ws, title, period_label, ncols)

    if report['kind'] == 'flat':
        row = _write_table_header(ws, row, columns)
        for data_row in report['rows']:
            row = _write_data_row(ws, row, data_row)
        _autosize(ws, columns, report['rows'])
    else:  # grouped (purchases)
        all_lines = []
        for grp in report['groups']:
            h = grp['header']
            info = (
                f"No Transaksi: {h['No Transaksi']}    Tanggal: {h['Tanggal']}    "
                f"Dept.: {h['Dept.']}    Supplier: {h['Nama Supplier']}"
            )
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(ncols, 1))
            cell = ws.cell(row, 1, info)
            cell.font = Font(bold=True, size=10)
            cell.alignment = _LEFT
            row += 1
            row = _write_table_header(ws, row, columns)
            for line in grp['lines']:
                row = _write_data_row(ws, row, line)
                all_lines.append(line)
            row += 1  # spacer between invoices
        _autosize(ws, columns, all_lines)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class GenerateReportView(APIView):
    """
    GET /api/reports/generate/?type=<slug>&start=YYYY-MM-DD&end=YYYY-MM-DD&format=json|xlsx

    type ∈ {stock_in, stock_out, purchases, sales_by_item}. Defaults: range =
    month-to-date (Jakarta), format = json (on-screen preview). format=xlsx
    returns a downloadable spreadsheet.
    """
    permission_classes = [AllowAny]  # reached via anchor download (no auth header), mirrors InvoiceExportView
    content_negotiation_class = IgnoreFormatQueryNegotiation

    def get(self, request):
        report_type = request.query_params.get('type', '').strip()
        if report_type not in _REPORT_BUILDERS:
            return Response(
                {'error': 'Jenis laporan tidak dikenal.', 'valid_types': list(_REPORT_BUILDERS)},
                status=status.HTTP_400_BAD_REQUEST,
            )

        start, end, err = _parse_report_range(request)
        if err:
            return Response({'error': err}, status=status.HTTP_400_BAD_REQUEST)

        report = _REPORT_BUILDERS[report_type](start, end)
        title = _REPORT_TITLES[report_type]
        period_label = f'{start.strftime("%d/%m/%y")} - {end.strftime("%d/%m/%y")}'

        # Accept both ?output= and the legacy ?format=. DRF normally 404s on an
        # unknown ?format value, but IgnoreFormatQueryNegotiation disables that.
        fmt = (
            request.query_params.get('output')
            or request.query_params.get('format')
            or 'json'
        ).strip().lower()
        if fmt == 'xlsx':
            filename = f'{report_type}_{start}_{end}.xlsx'
            return _report_to_xlsx(report, title, period_label, filename)

        meta = {
            'type': report_type,
            'title': title,
            'start': str(start),
            'end': str(end),
            'period_label': period_label,
        }
        if report['kind'] == 'flat':
            meta['total_rows'] = len(report['rows'])
            return Response({'meta': meta, **report})
        # grouped
        meta['total_rows'] = sum(len(g['lines']) for g in report['groups'])
        meta['group_count'] = len(report['groups'])
        return Response({'meta': meta, **report})
