"""
Financial statement endpoints, all under /api/accounting/reports/.

    trial-balance/    Neraca Saldo     ?as_of=YYYY-MM-DD
    profit-loss/      Laba Rugi        ?date_from=&date_to=
    balance-sheet/    Neraca           ?as_of=YYYY-MM-DD
    general-ledger/   Buku Besar       ?account=<id|all>&date_from=&date_to=
    cash-flow/        Arus Kas         ?date_from=&date_to=

Every endpoint returns JSON by default, or a styled .xlsx workbook when
`?format=xlsx` is supplied. All figures are computed from the LedgerEntry
journal via financial_reports_utils.
"""

import datetime
import io
from decimal import Decimal
from zoneinfo import ZoneInfo

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from ..models import ChartOfAccounts, LedgerEntry
from .financial_reports_utils import (
    ACTIVITY_LABELS,
    ACTIVITY_ORDER,
    ZERO,
    account_movements,
    accounts_by_id,
    cash_report_account_ids,
    classify_activity,
    earliest_ledger_date,
    earnings_through,
    opening_balances,
    signed_balance,
    unposted_dates_in_range,
)

WIB = ZoneInfo('Asia/Jakarta')

TYPE_LABELS = {
    'asset':         'Aset',
    'liability':     'Kewajiban',
    'equity':        'Ekuitas',
    'revenue':       'Pendapatan',
    'cogs':          'Harga Pokok Penjualan',
    'expense':       'Beban Operasional',
    'other_income':  'Pendapatan Lain-lain',
    'other_expense': 'Beban Lain-lain',
}

SOURCE_LABELS = {
    'invoice':    'Penjualan',
    'purchase':   'Pembelian',
    'transfer':   'Transfer',
    'adjustment': 'Penyesuaian',
    'stock':      'Mutasi Stok',
    'opname':     'Stock Opname',
    'manual':     'Manual',
    '':           'Lainnya',
}


# ── shared helpers ───────────────────────────────────────────────────────────

def _today_wib():
    return timezone.now().astimezone(WIB).date()


def _parse_date(value, field):
    """Return (date, error_response). Blank → (None, None)."""
    value = (value or '').strip()
    if not value:
        return None, None
    try:
        return datetime.date.fromisoformat(value), None
    except ValueError:
        return None, Response(
            {'error': f'{field} tidak valid. Gunakan format YYYY-MM-DD.'},
            status=status.HTTP_400_BAD_REQUEST,
        )


def _parse_range(request):
    """Returns (date_from, date_to, error_response). Both required."""
    d_from, err = _parse_date(request.query_params.get('date_from'), 'date_from')
    if err:
        return None, None, err
    d_to, err = _parse_date(request.query_params.get('date_to'), 'date_to')
    if err:
        return None, None, err
    if not d_from or not d_to:
        return None, None, Response(
            {'error': 'date_from dan date_to wajib diisi (YYYY-MM-DD).'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if d_from > d_to:
        return None, None, Response(
            {'error': 'date_from tidak boleh setelah date_to.'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    return d_from, d_to, None


def _unposted_gate(date_from, date_to):
    """Return a 400 Response listing unposted_dates when any day in
    [date_from, date_to] hasn't been posted by a journal run yet, else None.
    Every report view calls this before computing so a partially-posted
    period never gets silently reported as complete."""
    unposted = unposted_dates_in_range(date_from, date_to)
    if not unposted:
        return None
    return Response(
        {'unposted_dates': [d.isoformat() for d in unposted]},
        status=status.HTTP_400_BAD_REQUEST,
    )


def _unposted_gate_as_of(as_of):
    """Same as _unposted_gate but for as_of-style reports (Trial Balance,
    Balance Sheet), which have no explicit date_from — the range checked runs
    from the oldest LedgerEntry date through as_of, since those reports are
    cumulative from the start of the journal. An empty journal has nothing to
    gate."""
    start = earliest_ledger_date()
    if start is None:
        return None
    return _unposted_gate(start, as_of)


_CENTS = Decimal('0.01')


def _s(val):
    """Serialize a money Decimal as a string fixed to 2 decimal places, so the
    representation is identical regardless of DB backend / aggregate path."""
    return str((val if val is not None else ZERO).quantize(_CENTS))


def _bump(mapping, key, delta):
    """Add a signed money delta into ``mapping[key]``'s inflow/outflow pair,
    creating the slot on first use. Every bucket in the cash-flow statement
    keeps gross inflow and outflow separately, so a single signed number can
    never be split by hand at each call site."""
    slot = mapping.setdefault(key, {'inflow': ZERO, 'outflow': ZERO})
    if delta > 0:
        slot['inflow'] += delta
    else:
        slot['outflow'] += -delta


def _attribute(cash_delta, counterparts):
    """Attribute one journal entry's net cash movement to its counterpart accounts.

    ``cash_delta`` is the entry's net cash change (debit positive).
    ``counterparts`` are its non-cash lines.

    Attribution is by *signed* side, not raw amount. In a balanced entry the
    counterparts sum to exactly ``-cash_delta``, so each account's share is
    simply the negation of its own signed amount — which is the only way a
    contra account lands on the correct side. A cash sale posting
    ``Dr Cash 100 / Dr Sales Discount 10 / Cr Revenue 110`` must report revenue
    as a 110 inflow and the discount as a 10 outflow; weighting by raw amount
    would instead show the discount as an 8.33 *inflow*.

    When the entry does not balance the shares are scaled to foot to
    ``cash_delta`` anyway, so the statement always ties to the cash accounts.

    Yields ``(account_id, share)``, share positive for inflow.
    """
    signed = {}
    for line in counterparts:
        delta = line.amount if line.entry_type == 'debit' else -line.amount
        signed[line.account_id] = signed.get(line.account_id, ZERO) - delta

    total = sum(signed.values(), ZERO)
    if total == cash_delta or total == ZERO:
        shares = {k: v.quantize(_CENTS) for k, v in signed.items()}
    else:
        shares = {k: (v * cash_delta / total).quantize(_CENTS) for k, v in signed.items()}

    drift = cash_delta - sum(shares.values(), ZERO)
    if drift and shares:
        biggest = max(shares, key=lambda k: abs(shares[k]))
        shares[biggest] += drift
    return list(shares.items())


def _wants_xlsx(request):
    # NOTE: `format` is reserved by DRF content negotiation (returns 404 for an
    # unknown format), so the export trigger uses `export=xlsx` instead.
    return request.query_params.get('export', '').strip().lower() == 'xlsx'


# ── openpyxl styling primitives (mirrors PaymentPlanExportView) ──────────────

ACCT_FMT = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'
_thin = Side(style='thin')
_all_thin = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_double_top = Border(top=_thin, bottom=Side(style='double'))
_f_normal = Font(name='Calibri', size=11)
_f_bold = Font(name='Calibri', size=11, bold=True)
_f_title = Font(name='Calibri', size=14, bold=True)
_f_section = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
_section_fill = PatternFill(patternType='solid', fgColor='4F46E5')
_header_fill = PatternFill(patternType='solid', fgColor='E5E7EB')
_center = Alignment(horizontal='center', vertical='center')
_right = Alignment(horizontal='right', vertical='center')
_left = Alignment(horizontal='left', vertical='center')


def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _title_row(ws, text, span, subtitle=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = _f_title
    c.alignment = Alignment(horizontal='center')
    row = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
        s = ws.cell(row=2, column=1, value=subtitle)
        s.font = _f_normal
        s.alignment = Alignment(horizontal='center')
        row = 3
    return row + 1  # first free row after a spacer


# ── Neraca Saldo (Trial Balance) ─────────────────────────────────────────────

class TrialBalanceView(APIView):
    """GET /api/accounting/reports/trial-balance/?as_of=YYYY-MM-DD"""

    def get(self, request):
        as_of, err = _parse_date(request.query_params.get('as_of'), 'as_of')
        if err:
            return err
        if not as_of:
            as_of = _today_wib()

        gate = _unposted_gate_as_of(as_of)
        if gate:
            return gate

        mv = account_movements(date_to=as_of)
        accts = accounts_by_id()

        rows = []
        total_debit = total_credit = ZERO
        for acc_id, m in mv.items():
            acc = accts.get(acc_id)
            if not acc:
                continue
            net = m['net']
            if net == ZERO:
                continue
            debit = net if net > 0 else ZERO
            credit = -net if net < 0 else ZERO
            total_debit += debit
            total_credit += credit
            rows.append({
                'account_id':     acc_id,
                'account_number': acc.account_number,
                'name':           acc.name,
                'account_type':   acc.account_type,
                'debit':          _s(debit),
                'credit':         _s(credit),
            })

        rows.sort(key=lambda r: r['account_number'])
        payload = {
            'as_of':        as_of.isoformat(),
            'rows':         rows,
            'total_debit':  _s(total_debit),
            'total_credit': _s(total_credit),
            'is_balanced':  total_debit == total_credit,
        }

        if _wants_xlsx(request):
            return self._xlsx(payload)
        return Response(payload)

    def _xlsx(self, p):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Neraca Saldo'
        for col, w in zip('ABCD', [14, 42, 18, 18]):
            ws.column_dimensions[col].width = w
        r = _title_row(ws, 'Neraca Saldo', 4, f"Per {p['as_of']}")

        for idx, label in enumerate(['No. Akun', 'Nama Akun', 'Debit', 'Kredit']):
            c = ws.cell(row=r, column=idx + 1, value=label)
            c.font = _f_bold; c.alignment = _center; c.border = _all_thin; c.fill = _header_fill
        r += 1

        for row in p['rows']:
            ws.cell(row=r, column=1, value=row['account_number']).border = _all_thin
            ws.cell(row=r, column=2, value=row['name']).border = _all_thin
            for col, key in ((3, 'debit'), (4, 'credit')):
                cell = ws.cell(row=r, column=col, value=float(row[key]))
                cell.number_format = ACCT_FMT; cell.border = _all_thin; cell.alignment = _right
            r += 1

        tl = ws.cell(row=r, column=2, value='TOTAL')
        tl.font = _f_bold; tl.alignment = _right; tl.border = _double_top
        for col, key in ((3, 'total_debit'), (4, 'total_credit')):
            cell = ws.cell(row=r, column=col, value=float(p[key]))
            cell.font = _f_bold; cell.number_format = ACCT_FMT; cell.border = _double_top; cell.alignment = _right
        return _xlsx_response(wb, f"neraca-saldo-{p['as_of']}.xlsx")


# ── Laba Rugi (Profit & Loss) ────────────────────────────────────────────────

class ProfitLossView(APIView):
    """GET /api/accounting/reports/profit-loss/?date_from=&date_to=&compare=prev"""

    def get(self, request):
        d_from, d_to, err = _parse_range(request)
        if err:
            return err

        gate = _unposted_gate(d_from, d_to)
        if gate:
            return gate

        current = self._compute(d_from, d_to)
        payload = {
            'date_from': d_from.isoformat(),
            'date_to':   d_to.isoformat(),
            **current,
        }

        if request.query_params.get('compare', '').strip() == 'prev':
            span = (d_to - d_from).days + 1
            p_to = d_from - datetime.timedelta(days=1)
            p_from = p_to - datetime.timedelta(days=span - 1)
            payload['comparison'] = {
                'date_from': p_from.isoformat(),
                'date_to':   p_to.isoformat(),
                **self._compute(p_from, p_to),
            }

        if _wants_xlsx(request):
            return self._xlsx(payload)
        return Response(payload)

    def _section(self, mv, accts, types):
        """Return (rows, total) for the given account types, natural sign."""
        rows, total = [], ZERO
        for acc_id, m in mv.items():
            acc = accts.get(acc_id)
            if not acc or acc.account_type not in types:
                continue
            amt = signed_balance(acc.account_type, m['net'])
            if amt == ZERO:
                continue
            total += amt
            rows.append({
                'account_number': acc.account_number,
                'name':           acc.name,
                'amount':         _s(amt),
            })
        rows.sort(key=lambda r: r['account_number'])
        return rows, total

    def _compute(self, d_from, d_to):
        mv = account_movements(date_from=d_from, date_to=d_to)
        accts = accounts_by_id()

        revenue, rev_total = self._section(mv, accts, ('revenue',))
        cogs, cogs_total = self._section(mv, accts, ('cogs',))
        gross_profit = rev_total - cogs_total
        opex, opex_total = self._section(mv, accts, ('expense',))
        operating_income = gross_profit - opex_total
        other_inc, oi_total = self._section(mv, accts, ('other_income',))
        other_exp, oe_total = self._section(mv, accts, ('other_expense',))
        net_profit = operating_income + oi_total - oe_total

        return {
            'revenue':          {'rows': revenue, 'total': _s(rev_total)},
            'cogs':             {'rows': cogs, 'total': _s(cogs_total)},
            'gross_profit':     _s(gross_profit),
            'operating_expenses': {'rows': opex, 'total': _s(opex_total)},
            'operating_income': _s(operating_income),
            'other_income':     {'rows': other_inc, 'total': _s(oi_total)},
            'other_expenses':   {'rows': other_exp, 'total': _s(oe_total)},
            'net_profit':       _s(net_profit),
        }

    def _xlsx(self, p):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Laba Rugi'
        for col, w in zip('ABC', [14, 44, 20]):
            ws.column_dimensions[col].width = w
        r = _title_row(ws, 'Laporan Laba Rugi', 3, f"Periode {p['date_from']} s/d {p['date_to']}")

        def section(title, block, sign=''):
            nonlocal r
            c = ws.cell(row=r, column=1, value=title)
            c.font = _f_section; c.fill = _section_fill
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 1
            for row in block['rows']:
                ws.cell(row=r, column=1, value=row['account_number'])
                ws.cell(row=r, column=2, value=row['name'])
                cell = ws.cell(row=r, column=3, value=float(row['amount']))
                cell.number_format = ACCT_FMT; cell.alignment = _right
                r += 1
            tl = ws.cell(row=r, column=2, value=f'Total {title}')
            tl.font = _f_bold; tl.alignment = _right
            tc = ws.cell(row=r, column=3, value=float(block['total']))
            tc.font = _f_bold; tc.number_format = ACCT_FMT; tc.alignment = _right
            r += 2

        def subtotal(label, value):
            nonlocal r
            tl = ws.cell(row=r, column=2, value=label)
            tl.font = _f_bold; tl.alignment = _right; tl.border = _double_top
            tc = ws.cell(row=r, column=3, value=float(value))
            tc.font = _f_bold; tc.number_format = ACCT_FMT; tc.alignment = _right; tc.border = _double_top
            r += 2

        section('Pendapatan', p['revenue'])
        section('Harga Pokok Penjualan', p['cogs'])
        subtotal('LABA KOTOR', p['gross_profit'])
        section('Beban Operasional', p['operating_expenses'])
        subtotal('LABA OPERASIONAL', p['operating_income'])
        section('Pendapatan Lain-lain', p['other_income'])
        section('Beban Lain-lain', p['other_expenses'])
        subtotal('LABA BERSIH', p['net_profit'])
        return _xlsx_response(wb, f"laba-rugi-{p['date_from']}_{p['date_to']}.xlsx")


# ── Neraca (Balance Sheet) ───────────────────────────────────────────────────

class BalanceSheetView(APIView):
    """GET /api/accounting/reports/balance-sheet/?as_of=YYYY-MM-DD"""

    def get(self, request):
        as_of, err = _parse_date(request.query_params.get('as_of'), 'as_of')
        if err:
            return err
        if not as_of:
            as_of = _today_wib()

        gate = _unposted_gate_as_of(as_of)
        if gate:
            return gate

        mv = account_movements(date_to=as_of)
        accts = accounts_by_id()

        def section(types):
            rows, total = [], ZERO
            for acc_id, m in mv.items():
                acc = accts.get(acc_id)
                if not acc or acc.account_type not in types:
                    continue
                amt = signed_balance(acc.account_type, m['net'])
                if amt == ZERO:
                    continue
                total += amt
                rows.append({
                    'account_number': acc.account_number,
                    'name':           acc.name,
                    'amount':         _s(amt),
                })
            rows.sort(key=lambda r: r['account_number'])
            return rows, total

        assets, assets_total = section(('asset',))
        liabilities, liab_total = section(('liability',))
        equity, equity_total = section(('equity',))

        earnings = earnings_through(as_of)
        equity_rows = list(equity)
        equity_rows.append({
            'account_number': None,
            'name':           'Laba Berjalan / Ditahan (dihitung)',
            'amount':         _s(earnings),
        })
        equity_with_earnings = equity_total + earnings
        liab_and_equity = liab_total + equity_with_earnings

        payload = {
            'as_of':       as_of.isoformat(),
            'assets':      {'rows': assets, 'total': _s(assets_total)},
            'liabilities': {'rows': liabilities, 'total': _s(liab_total)},
            'equity':      {'rows': equity_rows, 'total': _s(equity_with_earnings),
                            'retained_earnings': _s(earnings)},
            'total_assets':                 _s(assets_total),
            'total_liabilities_and_equity': _s(liab_and_equity),
            'is_balanced':                  assets_total == liab_and_equity,
        }

        if _wants_xlsx(request):
            return self._xlsx(payload)
        return Response(payload)

    def _xlsx(self, p):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Neraca'
        for col, w in zip('ABC', [14, 44, 20]):
            ws.column_dimensions[col].width = w
        r = _title_row(ws, 'Neraca', 3, f"Per {p['as_of']}")

        def section(title, block):
            nonlocal r
            c = ws.cell(row=r, column=1, value=title)
            c.font = _f_section; c.fill = _section_fill
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            r += 1
            for row in block['rows']:
                if row['account_number'] is not None:
                    ws.cell(row=r, column=1, value=row['account_number'])
                ws.cell(row=r, column=2, value=row['name'])
                cell = ws.cell(row=r, column=3, value=float(row['amount']))
                cell.number_format = ACCT_FMT; cell.alignment = _right
                r += 1
            tl = ws.cell(row=r, column=2, value=f'Total {title}')
            tl.font = _f_bold; tl.alignment = _right; tl.border = _double_top
            tc = ws.cell(row=r, column=3, value=float(block['total']))
            tc.font = _f_bold; tc.number_format = ACCT_FMT; tc.alignment = _right; tc.border = _double_top
            r += 2

        section('Aset', p['assets'])
        section('Kewajiban', p['liabilities'])
        section('Ekuitas', p['equity'])

        tl = ws.cell(row=r, column=2, value='TOTAL KEWAJIBAN + EKUITAS')
        tl.font = _f_bold; tl.alignment = _right
        tc = ws.cell(row=r, column=3, value=float(p['total_liabilities_and_equity']))
        tc.font = _f_bold; tc.number_format = ACCT_FMT; tc.alignment = _right
        return _xlsx_response(wb, f"neraca-{p['as_of']}.xlsx")


# ── Buku Besar (General Ledger) ──────────────────────────────────────────────

class GeneralLedgerView(APIView):
    """GET /api/accounting/reports/general-ledger/?account=<id|all>&date_from=&date_to="""

    def get(self, request):
        d_from, d_to, err = _parse_range(request)
        if err:
            return err

        gate = _unposted_gate(d_from, d_to)
        if gate:
            return gate

        account_param = request.query_params.get('account', 'all').strip()
        accts = accounts_by_id()

        if account_param and account_param != 'all':
            try:
                target_ids = [int(account_param)]
            except ValueError:
                return Response({'error': 'account tidak valid.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            target_ids = None  # all accounts that have entries

        opening = opening_balances(d_from, account_ids=target_ids)

        entries_qs = (
            LedgerEntry.objects
            .filter(date__gte=d_from, date__lte=d_to)
            .select_related('account')
            .order_by('account__account_number', 'date', 'created_at')
        )
        if target_ids is not None:
            entries_qs = entries_qs.filter(account_id__in=target_ids)

        # Group entries by account, carrying a running balance from opening.
        per_account = {}
        for e in entries_qs:
            acc = e.account
            bucket = per_account.setdefault(acc.id, {
                'account_number': acc.account_number,
                'name':           acc.name,
                'account_type':   acc.account_type,
                'opening':        opening.get(acc.id, ZERO),
                'running':        opening.get(acc.id, ZERO),
                'lines':          [],
            })
            delta = e.amount if e.entry_type == 'debit' else -e.amount
            bucket['running'] += delta
            bucket['lines'].append({
                'date':        e.date.isoformat(),
                'description': e.description,
                'source_type': e.source_type,
                'debit':       _s(e.amount if e.entry_type == 'debit' else ZERO),
                'credit':      _s(e.amount if e.entry_type == 'credit' else ZERO),
                'balance':     _s(bucket['running']),
            })

        # Include accounts that only have an opening balance (no in-range movement)
        if target_ids is not None:
            for acc_id in target_ids:
                if acc_id not in per_account and acc_id in accts:
                    acc = accts[acc_id]
                    per_account[acc_id] = {
                        'account_number': acc.account_number,
                        'name':           acc.name,
                        'account_type':   acc.account_type,
                        'opening':        opening.get(acc_id, ZERO),
                        'running':        opening.get(acc_id, ZERO),
                        'lines':          [],
                    }

        accounts_out = []
        for bucket in sorted(per_account.values(), key=lambda b: b['account_number']):
            accounts_out.append({
                'account_number':  bucket['account_number'],
                'name':            bucket['name'],
                'account_type':    bucket['account_type'],
                'opening_balance': _s(bucket['opening']),
                'closing_balance': _s(bucket['running']),
                'lines':           bucket['lines'],
            })

        payload = {
            'date_from': d_from.isoformat(),
            'date_to':   d_to.isoformat(),
            'accounts':  accounts_out,
        }

        if _wants_xlsx(request):
            return self._xlsx(payload)
        return Response(payload)

    def _xlsx(self, p):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Buku Besar'
        for col, w in zip('ABCDE', [12, 40, 16, 16, 18]):
            ws.column_dimensions[col].width = w
        r = _title_row(ws, 'Buku Besar', 5, f"Periode {p['date_from']} s/d {p['date_to']}")

        for acc in p['accounts']:
            c = ws.cell(row=r, column=1, value=f"{acc['account_number']} — {acc['name']}")
            c.font = _f_section; c.fill = _section_fill
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            r += 1
            for idx, label in enumerate(['Tanggal', 'Keterangan', 'Debit', 'Kredit', 'Saldo']):
                hc = ws.cell(row=r, column=idx + 1, value=label)
                hc.font = _f_bold; hc.fill = _header_fill; hc.border = _all_thin
            r += 1
            ob = ws.cell(row=r, column=2, value='Saldo Awal')
            ob.font = _f_bold; ob.alignment = _right
            obv = ws.cell(row=r, column=5, value=float(acc['opening_balance']))
            obv.font = _f_bold; obv.number_format = ACCT_FMT; obv.alignment = _right
            r += 1
            for ln in acc['lines']:
                ws.cell(row=r, column=1, value=ln['date'])
                ws.cell(row=r, column=2, value=ln['description'])
                for col, key in ((3, 'debit'), (4, 'credit'), (5, 'balance')):
                    cell = ws.cell(row=r, column=col, value=float(ln[key]))
                    cell.number_format = ACCT_FMT; cell.alignment = _right
                r += 1
            cb = ws.cell(row=r, column=2, value='Saldo Akhir')
            cb.font = _f_bold; cb.alignment = _right; cb.border = _double_top
            cbv = ws.cell(row=r, column=5, value=float(acc['closing_balance']))
            cbv.font = _f_bold; cbv.number_format = ACCT_FMT; cbv.alignment = _right; cbv.border = _double_top
            r += 2
        return _xlsx_response(wb, f"buku-besar-{p['date_from']}_{p['date_to']}.xlsx")


# ── Arus Kas (direct-method Cash Flow) ───────────────────────────────────────

class CashFlowView(APIView):
    """GET /api/accounting/reports/cash-flow/?date_from=&date_to=

    Direct-method cash flow built from the cash side of the journal.

    Three things make this more than a ``GROUP BY source_type`` over cash lines:

    * **Scope.** "Cash" is ``cash_report_account_ids()`` — every real cash/bank
      account including ones retired from the payment pickers, excluding the
      iPos history clearing account. Reporting on the ``1100000`` *head* instead
      (as this view used to) yields nothing at all: heads never carry entries.

    * **Internal transfers are netted out.** Moving money from Bank BCA to Cash
      is not an inflow and not an outflow; counting both legs inflates gross
      cash movement without changing the net. Those legs are reported separately
      as a memo line.

    * **Activity classification.** Each cash line is attributed through the
      *counterpart* accounts in its own journal entry, by signed side so a
      contra account lands as a reduction. That is what separates operating
      from investing and financing — ``source_type`` cannot, since a purchase
      of stock and a purchase of equipment are both ``purchase``.

    Anything that cannot be attributed (a journal entry whose only line is the
    cash line — i.e. an unbalanced entry) is reported in its own bucket rather
    than folded into operating, so a posting bug shows up on the report instead
    of hiding inside a subtotal.
    """

    def get(self, request):
        d_from, d_to, err = _parse_range(request)
        if err:
            return err

        gate = _unposted_gate(d_from, d_to)
        if gate:
            return gate

        cash_ids = cash_report_account_ids()
        accts = accounts_by_id()
        cash_accounts = sorted(
            (accts[i] for i in cash_ids if i in accts),
            key=lambda a: a.account_number,
        )

        opening_map = opening_balances(d_from, account_ids=cash_ids)
        opening_cash = sum((opening_map.get(i, ZERO) for i in cash_ids), ZERO)

        cash_lines = list(
            LedgerEntry.objects
            .filter(date__gte=d_from, date__lte=d_to, account_id__in=cash_ids)
            .only('id', 'account_id', 'amount', 'entry_type', 'source_type', 'journal_entry')
        )

        # One extra query pulls every line of every journal entry that touched
        # cash, so counterparts are resolved in memory rather than per row.
        je_ids = {e.journal_entry_id for e in cash_lines if e.journal_entry_id}
        siblings = {}
        if je_ids:
            for line in LedgerEntry.objects.filter(journal_entry_id__in=je_ids).only(
                'id', 'account_id', 'amount', 'entry_type', 'journal_entry'
            ):
                siblings.setdefault(line.journal_entry_id, []).append(line)

        # activity -> counterpart account_id -> {'inflow', 'outflow'}
        buckets = {a: {} for a in ACTIVITY_ORDER}
        by_source = {}
        internal_in = internal_out = ZERO
        unclassified = {}
        per_account = {
            i: {'inflow': ZERO, 'outflow': ZERO, 'opening': opening_map.get(i, ZERO)}
            for i in cash_ids
        }
        net_change = ZERO

        # Per-line first: net change and the per-account reconciliation are
        # purely mechanical and must include internal transfers, since those do
        # move an individual account's balance.
        cash_by_je = {}
        for e in cash_lines:
            delta = e.amount if e.entry_type == 'debit' else -e.amount
            net_change += delta

            pa = per_account[e.account_id]
            if e.entry_type == 'debit':
                pa['inflow'] += e.amount
            else:
                pa['outflow'] += e.amount

            cash_by_je.setdefault(e.journal_entry_id, []).append((e, delta))

        # Then per journal entry: attribution is only meaningful against a whole
        # entry, because that is where a cash line's counterparts live.
        for je_id, members in cash_by_je.items():
            group = siblings.get(je_id, [])
            counterparts = [line for line in group if line.account_id not in cash_ids]
            je_delta = sum((d for _, d in members), ZERO)

            if not counterparts:
                # Only cash lines in this entry: a genuine internal transfer if
                # there is more than one, otherwise the entry is one-sided and
                # cannot be attributed at all. je_id None is the pre-migration-0098
                # backfill gap — those lines share no header, so they are never
                # each other's transfer counterpart no matter how many there are.
                internal = je_id is not None and len(members) > 1
                for e, delta in members:
                    if internal:
                        if delta > 0:
                            internal_in += delta
                        else:
                            internal_out += -delta
                    else:
                        _bump(unclassified, e.source_type, delta)
                continue

            for e, delta in members:
                _bump(by_source, e.source_type, delta)

            source_type = members[0][0].source_type
            for cp_id, share in _attribute(je_delta, counterparts):
                cp = accts.get(cp_id)
                activity = classify_activity(cp) if cp else None
                if activity:
                    _bump(buckets[activity], cp_id, share)
                else:
                    _bump(unclassified, source_type, share)

        closing_cash = opening_cash + net_change

        activities = []
        for name in ACTIVITY_ORDER:
            rows = []
            a_in = a_out = ZERO
            for acc_id, v in buckets[name].items():
                acc = accts.get(acc_id)
                a_in += v['inflow']
                a_out += v['outflow']
                rows.append({
                    'account_number': acc.account_number if acc else None,
                    'account_name':   acc.name if acc else 'Akun tidak dikenal',
                    'inflow':         _s(v['inflow']),
                    'outflow':        _s(v['outflow']),
                    'net':            _s(v['inflow'] - v['outflow']),
                })
            rows.sort(key=lambda r: (r['account_number'] is None, r['account_number'] or 0))
            activities.append({
                'activity': name,
                'label':    ACTIVITY_LABELS[name],
                'lines':    rows,
                'inflow':   _s(a_in),
                'outflow':  _s(a_out),
                'net':      _s(a_in - a_out),
            })

        flows = []
        for source, g in sorted(by_source.items(), key=lambda kv: str(kv[0])):
            flows.append({
                'source_type': source,
                'label':       SOURCE_LABELS.get(source, source or 'Lainnya'),
                'inflow':      _s(g['inflow']),
                'outflow':     _s(g['outflow']),
                'net':         _s(g['inflow'] - g['outflow']),
            })

        unclassified_rows = []
        u_in = u_out = ZERO
        for source, g in sorted(unclassified.items(), key=lambda kv: str(kv[0])):
            u_in += g['inflow']
            u_out += g['outflow']
            unclassified_rows.append({
                'source_type': source,
                'label':       SOURCE_LABELS.get(source, source or 'Lainnya'),
                'inflow':      _s(g['inflow']),
                'outflow':     _s(g['outflow']),
                'net':         _s(g['inflow'] - g['outflow']),
            })

        payload = {
            'date_from':    d_from.isoformat(),
            'date_to':      d_to.isoformat(),
            'opening_cash': _s(opening_cash),
            'activities':   activities,
            'flows':        flows,
            'internal_transfers': {
                'inflow':  _s(internal_in),
                'outflow': _s(internal_out),
                'net':     _s(internal_in - internal_out),
            },
            'unclassified': {
                'lines':   unclassified_rows,
                'inflow':  _s(u_in),
                'outflow': _s(u_out),
                'net':     _s(u_in - u_out),
            },
            'net_change':   _s(net_change),
            'closing_cash': _s(closing_cash),
            'accounts': [
                {
                    'account_number': a.account_number,
                    'name':           a.name,
                    'opening':        _s(per_account[a.id]['opening']),
                    'inflow':         _s(per_account[a.id]['inflow']),
                    'outflow':        _s(per_account[a.id]['outflow']),
                    'closing':        _s(
                        per_account[a.id]['opening']
                        + per_account[a.id]['inflow']
                        - per_account[a.id]['outflow']
                    ),
                }
                for a in cash_accounts
            ],
        }

        if _wants_xlsx(request):
            return self._xlsx(payload)
        return Response(payload)

    def _xlsx(self, p):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Arus Kas'
        for col, w in zip('ABCDE', [46, 18, 18, 18, 18]):
            ws.column_dimensions[col].width = w
        r = _title_row(ws, 'Laporan Arus Kas (Metode Langsung)', 4,
                       f"Periode {p['date_from']} s/d {p['date_to']}")

        ob = ws.cell(row=r, column=1, value='Saldo Kas Awal')
        ob.font = _f_bold
        obv = ws.cell(row=r, column=4, value=float(p['opening_cash']))
        obv.font = _f_bold; obv.number_format = ACCT_FMT; obv.alignment = _right
        r += 2

        for idx, label in enumerate(['Keterangan', 'Kas Masuk', 'Kas Keluar', 'Bersih']):
            c = ws.cell(row=r, column=idx + 1, value=label)
            c.font = _f_bold; c.fill = _header_fill; c.border = _all_thin
        r += 1

        def _money_row(row, label, block, bold=False, indent=0):
            c = ws.cell(row=row, column=1, value=('    ' * indent) + label)
            if bold:
                c.font = _f_bold
            for col, key in ((2, 'inflow'), (3, 'outflow'), (4, 'net')):
                cell = ws.cell(row=row, column=col, value=float(block[key]))
                cell.number_format = ACCT_FMT; cell.alignment = _right
                if bold:
                    cell.font = _f_bold
            return row + 1

        for act in p['activities']:
            sec = ws.cell(row=r, column=1, value=act['label'])
            sec.font = _f_section; sec.fill = _section_fill
            r += 1
            for line in act['lines']:
                r = _money_row(
                    r, f"{line['account_number']} — {line['account_name']}", line, indent=1)
            r = _money_row(r, f"Jumlah — {act['label']}", act, bold=True)
            r += 1

        if p['unclassified']['lines']:
            sec = ws.cell(row=r, column=1, value='Belum Terklasifikasi (jurnal tidak seimbang)')
            sec.font = _f_section; sec.fill = _section_fill
            r += 1
            for line in p['unclassified']['lines']:
                r = _money_row(r, line['label'], line, indent=1)
            r = _money_row(r, 'Jumlah — Belum Terklasifikasi', p['unclassified'], bold=True)
            r += 1

        nl = ws.cell(row=r, column=1, value='Perubahan Kas Bersih')
        nl.font = _f_bold; nl.border = _double_top
        nv = ws.cell(row=r, column=4, value=float(p['net_change']))
        nv.font = _f_bold; nv.number_format = ACCT_FMT; nv.alignment = _right; nv.border = _double_top
        r += 1
        cl = ws.cell(row=r, column=1, value='Saldo Kas Akhir')
        cl.font = _f_bold
        cv = ws.cell(row=r, column=4, value=float(p['closing_cash']))
        cv.font = _f_bold; cv.number_format = ACCT_FMT; cv.alignment = _right
        r += 2

        it = p['internal_transfers']
        if float(it['inflow']) or float(it['outflow']):
            memo = ws.cell(
                row=r, column=1,
                value='Memo — mutasi antar rekening kas '
                      f"(tidak memengaruhi kas bersih): {float(it['inflow']):,.0f}")
            memo.font = _f_normal
            r += 2

        # Per-account reconciliation, so the operator can tie each cash/bank
        # account's closing balance back to the statement total.
        sec = ws.cell(row=r, column=1, value='Rincian per Rekening Kas')
        sec.font = _f_section; sec.fill = _section_fill
        r += 1
        for idx, label in enumerate(['Rekening', 'Saldo Awal', 'Masuk', 'Keluar', 'Saldo Akhir']):
            c = ws.cell(row=r, column=idx + 1, value=label)
            c.font = _f_bold; c.fill = _header_fill; c.border = _all_thin
        r += 1
        for a in p['accounts']:
            ws.cell(row=r, column=1, value=f"{a['account_number']} — {a['name']}")
            for col, key in ((2, 'opening'), (3, 'inflow'), (4, 'outflow'), (5, 'closing')):
                cell = ws.cell(row=r, column=col, value=float(a[key]))
                cell.number_format = ACCT_FMT; cell.alignment = _right
            r += 1

        return _xlsx_response(wb, f"arus-kas-{p['date_from']}_{p['date_to']}.xlsx")
