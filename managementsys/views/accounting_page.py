import datetime
import io
import json
from decimal import Decimal, InvalidOperation

import openpyxl
from django.db import transaction
from django.db.models import Count, DecimalField, ExpressionWrapper, F, Max, Q, Sum
from django.http import HttpResponse, StreamingHttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from ..api.serializers import (
    AccountTransferSerializer,
    ExpenseSerializer,
    JournalEntryDetailSerializer,
    JournalEntryListSerializer,
    JournalEntryRefSerializer,
    JournalStagingBatchSerializer,
    LedgerEntrySerializer,
    PurchaseInvoiceDetailSerializer,
    PurchaseInvoiceItemSerializer,
    PurchaseInvoiceListSerializer,
    StagedJournalEntryDetailSerializer,
    StagedJournalEntryListSerializer,
    SupplierSerializer,
)
from ..services.journal_sweep import run_journal_sweep
from ..models import (
    AccountTransfer, AppUser, AuditLog, ChartOfAccounts,
    Expense, ExpenseItem,
    InventoryBatch, InventoryItem, Invoice, JournalBatch, JournalDayLog,
    JournalEntry, JournalStagingBatch, LedgerEntry,
    PurchaseAdditionalCost, PurchaseInvoice, PurchaseInvoiceItem, PurchasePayment,
    PaymentMethod, StagedJournalEntry, Supplier, Warehouse,
)


def _actor(request):
    return request.user if isinstance(request.user, AppUser) else None


def _safe_decimal(val) -> Decimal:
    try:
        return Decimal(str(val))
    except (InvalidOperation, TypeError):
        return Decimal('0')


# ── Double-entry posting helpers ─────────────────────────────────────────────
# Moved to managementsys/services/journal_engine.py in Phase 2 (same math, same
# account resolution) so the journal run and the same-day void/edit memo path
# can both call them. Re-imported here under their original names so every
# call site below (_post_le, _post_purchase_accrual, _unpost_purchase,
# _post_purchase_price_variance) is unchanged.
from ..services.branches import filter_by_branch, write_branch
from ..services.cash_accounts import cash_bank_account_ids
from ..services.sales_returns import build_sales_return_legs
from ..services.expense_create import create_expense
from ..services import journal_preview
from ..services.journal_engine import (
    CENT, INVENTORY_ASSET_NUMBER, NORMAL_BALANCE, PRICE_VARIANCE_NUMBER,
    LegSet,
    _apply_purchase_balance as _apply_balance,
    _ensure_price_variance_account, _post_expense_accrual, _post_expense_le, _post_le,
    _post_purchase_accrual, _post_purchase_price_variance, _unpost_expense,
    _unpost_purchase, build_stock_correction_legs, expense_credit_account,
    expense_credit_memo, expense_leg_memo, is_date_posted, legset_from_entry,
    post_account_transfer, reserve_entry_numbers, reverse_legset, write_legs,
)
from .reports_page import payment_method_breakdown


NOT_A_CASH_ACCOUNT = 'Rekening yang dipilih bukan rekening kas/bank.'


def _resolve_cash_account(data, *, current=None):
    """Resolve the credit-side COA for an expense from request data.

    Returns ``(account_or_None, error_response_or_None)``.

    Accepts the new ``cash_account`` key (a ChartOfAccounts pk) and, for
    back-compat, the old ``payment_account`` key (a PaymentMethod pk) whose
    ``linked_account`` is used when no ``cash_account`` was sent. ``current``
    is returned untouched when neither key is present at all.
    """
    if 'cash_account' in data:
        raw = data.get('cash_account')
        if raw in (None, '', 0, '0'):
            return None, None
        try:
            account = ChartOfAccounts.objects.get(pk=raw)
        except (ChartOfAccounts.DoesNotExist, TypeError, ValueError):
            return None, Response({'error': NOT_A_CASH_ACCOUNT},
                                  status=status.HTTP_400_BAD_REQUEST)
        if account.pk not in cash_bank_account_ids():
            return None, Response({'error': NOT_A_CASH_ACCOUNT},
                                  status=status.HTTP_400_BAD_REQUEST)
        return account, None

    return current, None


# ── Suppliers ──────────────────────────────────────────────────────────────────

class SupplierListCreateView(APIView):
    def get(self, request):
        qs = Supplier.objects.all()
        q = request.query_params.get('q', '').strip()
        if q:
            qs = qs.filter(name__icontains=q)
        active = request.query_params.get('active', '').strip()
        if active == 'true':
            qs = qs.filter(is_active=True)
        return Response(SupplierSerializer(qs, many=True).data)

    def post(self, request):
        ser = SupplierSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        instance = ser.save()
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='CREATE',
            entity_type='Supplier',
            entity_id=str(instance.id),
            description=f'Supplier created: {instance.name}',
        )
        return Response(SupplierSerializer(instance).data, status=status.HTTP_201_CREATED)


class SupplierDetailView(APIView):
    def _get(self, pk):
        try:
            return Supplier.objects.get(pk=pk)
        except Supplier.DoesNotExist:
            return None

    def get(self, request, pk):
        obj = self._get(pk)
        if not obj:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(SupplierSerializer(obj).data)

    def put(self, request, pk):
        obj = self._get(pk)
        if not obj:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        ser = SupplierSerializer(obj, data=request.data)
        ser.is_valid(raise_exception=True)
        instance = ser.save()
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='Supplier',
            entity_id=str(instance.id),
            description=f'Supplier updated: {instance.name}',
        )
        return Response(SupplierSerializer(instance).data)

    def delete(self, request, pk):
        obj = self._get(pk)
        if not obj:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        name = obj.name
        obj.delete()
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='DELETE',
            entity_type='Supplier',
            entity_id=str(pk),
            description=f'Supplier deleted: {name}',
        )
        return Response(status=status.HTTP_204_NO_CONTENT)


class SupplierAccountView(APIView):
    """GET /api/accounting/suppliers/<pk>/account/

    Everything needed for a vendor detail page in one call:
      - the supplier (incl. its AP account + cached AP balance)
      - outstanding payable total and invoice counts
      - every item ever purchased from this vendor, aggregated
    Purchase invoices themselves come from /purchases/?supplier=<pk>, and the
    per-line AP journal from /admin/accounts/<ap_account_id>/ledger/.
    """

    def get(self, request, pk):
        try:
            supplier = Supplier.objects.select_related('ap_account').get(pk=pk)
        except Supplier.DoesNotExist:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        live = PurchaseInvoice.objects.filter(supplier_id=pk, is_voided=False)
        outstanding = live.filter(status__in=['unpaid', 'partial']).aggregate(
            s=Sum('total_amount') - Sum('amount_paid')
        )['s'] or Decimal('0')

        items = list(
            PurchaseInvoiceItem.objects
            .filter(invoice__supplier_id=pk, invoice__is_voided=False)
            .values('item_id', 'item__code', 'item_name', 'line_type')
            .annotate(
                total_qty=Sum('quantity'),
                total_spend=Sum(ExpressionWrapper(
                    F('quantity') * F('actual_unit_cost'),
                    output_field=DecimalField(max_digits=20, decimal_places=4),
                )),
                last_date=Max('invoice__purchase_date'),
                purchase_count=Count('invoice', distinct=True),
            )
            .order_by('-total_spend')
        )
        for row in items:
            row['item_code'] = row.pop('item__code')
            row['total_qty'] = str(row['total_qty'] or 0)
            row['total_spend'] = str((row['total_spend'] or Decimal('0')).quantize(Decimal('0.01')))
            row['last_date'] = row['last_date'].isoformat() if row['last_date'] else None

        return Response({
            'supplier':          SupplierSerializer(supplier).data,
            'outstanding':       str(outstanding),
            'invoice_count':     live.count(),
            'outstanding_count': live.filter(status__in=['unpaid', 'partial']).count(),
            'items':             items,
        })


class SupplierTemplateView(APIView):
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Vendors'
        ws.append(['name', 'contact_name', 'phone', 'email', 'address'])
        ws.append(['PT Contoh Supplier', 'Budi Santoso', '0812-3456-7890', 'budi@example.com', 'Jl. Contoh No. 1'])
        ws.append(['CV Bahan Kecantikan', 'Sari Dewi', '0811-9876-5432', '', 'Jl. Raya Kosmetik 5'])
        for col, width in zip('ABCDE', [30, 22, 20, 28, 36]):
            ws.column_dimensions[col].width = width
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="vendors_template.xlsx"'},
        )


class SupplierImportView(APIView):
    def post(self, request):
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'No file provided.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
            ws = wb.active
        except Exception:
            return Response({'error': 'Could not read Excel file.'}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'error': 'File is empty.'}, status=status.HTTP_400_BAD_REQUEST)

        headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
        if 'name' not in headers:
            return Response({'error': 'Column "name" is required in the first row.'}, status=status.HTTP_400_BAD_REQUEST)

        def col(row, field):
            try:
                idx = headers.index(field)
                val = row[idx] if idx < len(row) else None
                return str(val).strip() if val is not None else ''
            except ValueError:
                return ''

        created = updated = 0
        errors = []
        for i, row in enumerate(rows[1:], start=2):
            name = col(row, 'name')
            if not name:
                continue
            data = {
                'name': name,
                'contact_name': col(row, 'contact_name'),
                'phone': col(row, 'phone'),
                'email': col(row, 'email'),
                'address': col(row, 'address'),
            }
            existing = Supplier.objects.filter(name__iexact=name).first()
            if existing:
                ser = SupplierSerializer(existing, data=data)
            else:
                ser = SupplierSerializer(data=data)
            if ser.is_valid():
                ser.save()
                if existing:
                    updated += 1
                else:
                    created += 1
            else:
                errors.append({'row': i, 'name': name, 'error': str(ser.errors)})

        return Response({'created': created, 'updated': updated, 'errors': errors})


# ── Purchase Invoices ──────────────────────────────────────────────────────────

def _resolve_payment_method(raw_id):
    """Look up a PaymentMethod by id, returning None when it is missing or
    unknown — callers turn that into a 400 with their own message."""
    if not raw_id:
        return None
    try:
        return PaymentMethod.objects.select_related('linked_account').get(pk=raw_id)
    except (PaymentMethod.DoesNotExist, TypeError, ValueError):
        return None


def _resolve_purchase_cash_account(data):
    """The cash/bank COA a purchase is settled from.

    Prefers the ``cash_account`` key (a ChartOfAccounts pk, what the picker
    sends now). Falls back to the legacy ``payment_account``/``payment_method``
    keys (a PaymentMethod pk) so older clients keep working, resolving them
    through ``linked_account``.

    Returns ``(account_or_None, legacy_method_or_None, error_response_or_None)``.
    The method is carried along only so the legacy mirror column stays
    populated. A *sent but unusable* account is an error rather than a silent
    ``None`` — otherwise callers that fall back to the invoice's last account
    would post the payment against the wrong account.
    """
    raw = data.get('cash_account')
    if raw not in (None, '', 0, '0'):
        try:
            account = ChartOfAccounts.objects.get(pk=raw)
        except (ChartOfAccounts.DoesNotExist, TypeError, ValueError):
            return None, None, Response({'error': NOT_A_CASH_ACCOUNT},
                                        status=status.HTTP_400_BAD_REQUEST)
        if account.pk not in cash_bank_account_ids():
            return None, None, Response({'error': NOT_A_CASH_ACCOUNT},
                                        status=status.HTTP_400_BAD_REQUEST)
        return account, None, None

    method = (_resolve_payment_method(data.get('payment_account'))
              or _resolve_payment_method(data.get('payment_method')))
    if method is not None and method.linked_account_id:
        return method.linked_account, method, None
    return None, method, None


def _parse_date(raw, fallback=None):
    """Accept 'YYYY-MM-DD' (or a date), falling back when empty/unparseable."""
    if isinstance(raw, datetime.date):
        return raw
    if raw:
        try:
            return datetime.date.fromisoformat(str(raw)[:10])
        except ValueError:
            pass
    return fallback


def _record_purchase_payment(invoice, amount, cash_account, pay_date, actor,
                             notes='', payment_method=None):
    """Record one settlement against a purchase invoice: a PurchasePayment row,
    the running amount_paid/status on the invoice, and the double entry
        Dr Accounts Payable — <vendor>   (the liability goes away)
        Cr <cash_account>                (funds leave)
    dated on the day the money actually moved, not the day it was keyed in.

    ``cash_account`` is a ChartOfAccounts row from ``cash_bank_account_ids()``.
    ``payment_method`` is the legacy PaymentMethod, stored only when a legacy
    client sent one — nothing reads it for posting.

    Caller must have validated the amount against the remaining balance and be
    inside a transaction.
    """
    payment = PurchasePayment.objects.create(
        invoice=invoice,
        payment_date=pay_date,
        payment_method=payment_method,
        payment_account=cash_account,
        amount=amount,
        notes=notes,
        created_by=actor,
    )

    invoice.amount_paid += amount
    # Mirror the latest account onto the invoice so list/detail views keep a
    # single "paid from" column without joining every payment.
    invoice.payment_account = cash_account
    invoice.payment_method = payment_method
    invoice.refresh_status()
    invoice.save(update_fields=['amount_paid', 'payment_account', 'payment_method', 'status'])

    ap_account = invoice.supplier.ensure_ap_account()
    _post_le(ap_account, 'debit', amount, invoice,
             f'Pembayaran utang {invoice.internal_id} — {invoice.supplier.name}', pay_date)
    _post_le(cash_account, 'credit', amount, invoice,
             f'Pembayaran pembelian {invoice.internal_id}', pay_date)

    return payment


class PurchaseInvoiceListCreateView(APIView):
    """
    GET  /api/accounting/purchases/
         ?status=unpaid|partial|paid
         ?supplier=<id>
         ?q=<search internal_id or external_invoice_no>
         ?overdue=true   (due_date < today and not paid)

    POST /api/accounting/purchases/
         { external_invoice_no, supplier, purchase_date, due_date?, notes?,
           payment_status: 'unpaid'|'paid',        (default 'unpaid')
           payment_account, payment_date,          (required when 'paid')
           items: [{item?, item_name, quantity, unit, unit_cost}] }

    A 'paid' invoice is settled in full the moment it is created, using the
    given payment method and payment date. An 'unpaid' one carries no payment
    method at all until a payment is recorded via .../pay/.
    """

    def get(self, request):
        qs = (
            PurchaseInvoice.objects
            .select_related('supplier', 'payment_account', 'payment_method', 'payment_method__linked_account')
            .prefetch_related('payments')   # last_payment_date, one query for the page
        )
        qs = filter_by_branch(qs, request)

        q = request.query_params.get('q', '').strip()
        if q:
            qs = qs.filter(
                Q(internal_id__icontains=q) |
                Q(external_invoice_no__icontains=q) |
                Q(supplier__name__icontains=q)
            )
        st = request.query_params.get('status', '').strip()
        if st in ('unpaid', 'partial', 'paid'):
            qs = qs.filter(status=st)
        supplier_id = request.query_params.get('supplier', '').strip()
        if supplier_id:
            qs = qs.filter(supplier_id=supplier_id)
        overdue = request.query_params.get('overdue', '').strip()
        if overdue == 'true':
            today = timezone.now().date()
            qs = qs.filter(due_date__lt=today).exclude(status='paid')

        # Hide voided invoices unless the caller explicitly asks to include them.
        if request.query_params.get('include_voided', '').strip().lower() not in ('1', 'true', 'yes'):
            qs = qs.filter(is_voided=False)

        return Response(PurchaseInvoiceListSerializer(qs, many=True).data)

    def post(self, request):
        data = request.data

        # Support multipart (with image) or plain JSON
        items_raw = data.get('items', [])
        if isinstance(items_raw, str):
            try:
                items = json.loads(items_raw)
            except (json.JSONDecodeError, ValueError):
                return Response({'error': 'Invalid items JSON.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            items = items_raw

        additional_costs_raw = data.get('additional_costs', [])
        if isinstance(additional_costs_raw, str):
            try:
                additional_costs = json.loads(additional_costs_raw)
            except (json.JSONDecodeError, ValueError):
                return Response({'error': 'Invalid additional_costs JSON.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            additional_costs = additional_costs_raw

        if not items:
            return Response({'error': 'At least one item is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            supplier = Supplier.objects.get(pk=data['supplier'])
        except (Supplier.DoesNotExist, KeyError):
            return Response({'error': 'Invalid supplier.'}, status=status.HTTP_400_BAD_REQUEST)

        # Paid-on-creation: the cash/bank account and the date the money moved
        # are both required, and the invoice is settled in full below once its
        # total is known. Unpaid: no payment account is stored — it is chosen at
        # payment time.
        pay_now = str(data.get('payment_status', 'unpaid')).strip().lower() == 'paid'
        cash_account, payment_method, cash_err = _resolve_purchase_cash_account(data)
        if cash_err is not None:
            return cash_err
        payment_date   = _parse_date(data.get('payment_date'))
        if pay_now:
            if cash_account is None:
                return Response(
                    {'error': 'Rekening kas/bank wajib dipilih untuk faktur yang langsung dibayar.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if payment_date is None:
                return Response(
                    {'error': 'Tanggal pembayaran wajib diisi untuk faktur yang langsung dibayar.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Resolve optional invoice-level warehouse
        warehouse_id = data.get('warehouse') or None
        warehouse_obj = None
        if warehouse_id:
            try:
                warehouse_obj = Warehouse.objects.get(pk=warehouse_id)
            except Warehouse.DoesNotExist:
                return Response({'error': 'Invalid warehouse.'}, status=status.HTTP_400_BAD_REQUEST)

        # Every expense line needs a target account so the debit side of the
        # accrual posting has somewhere to land.
        for row in items:
            if row.get('line_type', 'stock') == 'expense' and not (row.get('expense_account')):
                return Response(
                    {'error': 'Setiap baris beban harus memiliki akun beban.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        with transaction.atomic():
            invoice = PurchaseInvoice.objects.create(
                external_invoice_no=data.get('external_invoice_no', ''),
                supplier=supplier,
                payment_method=payment_method if pay_now else None,
                payment_account=cash_account if pay_now else None,
                warehouse=warehouse_obj,
                purchase_date=data['purchase_date'],
                due_date=data.get('due_date') or None,
                notes=data.get('notes', ''),
                created_by=_actor(request),
                branch=write_branch(request),
            )

            # Attach image if uploaded
            image_file = request.FILES.get('invoice_image')
            if image_file:
                invoice.invoice_image = image_file
                invoice.save(update_fields=['invoice_image'])

            # ── Step 1: parse items and compute per-item adjusted subtotals ──────
            parsed = []
            items_subtotal = Decimal('0')
            total_units = Decimal('0')

            for row in items:
                qty            = _safe_decimal(row.get('quantity', 0))
                cost           = _safe_decimal(row.get('unit_cost', 0))
                total_discount = _safe_decimal(row.get('total_discount', 0))
                line_type      = row.get('line_type', 'stock')
                item_id        = row.get('item') or None
                expense_acct_id = row.get('expense_account') or None

                # Row-level warehouse falls back to invoice-level warehouse
                row_wh_id = row.get('warehouse') or warehouse_id or None

                gross          = qty * cost
                # Ensure discount does not exceed gross
                discount_capped = min(total_discount, gross)
                adjusted_sub   = gross - discount_capped

                items_subtotal += adjusted_sub
                if line_type == 'stock' and qty > 0:
                    total_units += qty

                parsed.append({
                    'line_type':       line_type,
                    'item_id':         item_id,
                    'item_name':       row.get('item_name', ''),
                    'quantity':        qty,
                    'unit':            row.get('unit', ''),
                    'unit_cost':       cost,
                    'total_discount':  discount_capped,
                    'adjusted_sub':    adjusted_sub,
                    'expense_acct_id': expense_acct_id,
                    'warehouse_id':    row_wh_id,
                })

            # ── Step 2: compute net adjustment from additional costs ─────────────
            running_total = items_subtotal
            net_adjustment = Decimal('0')

            cost_objs = []
            for i, ac in enumerate(additional_costs):
                name        = str(ac.get('name', '')).strip()
                modifier    = ac.get('modifier', 'add')
                amount_type = ac.get('amount_type', 'cash')
                amount      = _safe_decimal(ac.get('amount', 0))

                if not name or amount <= 0:
                    continue

                if amount_type == 'percent':
                    adj = running_total * amount / Decimal('100')
                else:
                    adj = amount

                if modifier == 'subtract':
                    adj = -adj

                running_total  += adj
                net_adjustment += adj

                cost_objs.append(PurchaseAdditionalCost(
                    invoice=invoice,
                    name=name,
                    modifier=modifier,
                    amount_type=amount_type,
                    amount=amount,
                    sort_order=i,
                ))

            if cost_objs:
                PurchaseAdditionalCost.objects.bulk_create(cost_objs)

            grand_total = items_subtotal + net_adjustment

            # ── Step 3: distribute net_adjustment equally per stock unit ─────────
            per_unit_adj = Decimal('0')
            if total_units > 0:
                per_unit_adj = net_adjustment / total_units

            # ── Step 4: create item records and batches ───────────────────────────
            item_objs = []
            batches_to_create = []

            for p in parsed:
                qty = p['quantity']
                if qty > 0:
                    base_unit_cost = p['adjusted_sub'] / qty
                    actual = base_unit_cost + (per_unit_adj if p['line_type'] == 'stock' else Decimal('0'))
                else:
                    base_unit_cost = p['unit_cost']
                    actual = p['unit_cost']

                item_objs.append(PurchaseInvoiceItem(
                    invoice=invoice,
                    line_type=p['line_type'],
                    item_id=p['item_id'],
                    item_name=p['item_name'],
                    quantity=qty,
                    unit=p['unit'],
                    unit_cost=p['unit_cost'],
                    total_discount=p['total_discount'],
                    actual_unit_cost=actual,
                    expense_account_id=p['expense_acct_id'],
                    warehouse_id=p['warehouse_id'],
                ))

                if p['line_type'] == 'stock' and p['item_id'] and p['warehouse_id'] and qty > 0:
                    batches_to_create.append({
                        'item_id':     p['item_id'],
                        'warehouse_id': p['warehouse_id'],
                        'qty':         qty,
                        'cost':        actual,
                    })

            PurchaseInvoiceItem.objects.bulk_create(item_objs)

            # Create inventory batches (FIFO stock-in). value is the TOTAL batch
            # value (actual_unit_cost × qty); every consumer derives the per-unit
            # cost as value / quantity_initial.
            for b in batches_to_create:
                InventoryBatch.objects.create(
                    item_id=b['item_id'],
                    warehouse_id=b['warehouse_id'],
                    input_date=invoice.purchase_date,
                    quantity_initial=b['qty'],
                    quantity_remaining=b['qty'],
                    value=b['cost'] * b['qty'],
                    purchase_invoice=invoice,
                )

            grand_total = grand_total.quantize(Decimal('0.01'))
            invoice.total_amount = grand_total
            invoice.save(update_fields=['total_amount'])

            # Phase 2: the accrual double-entry (Dr Inventory/Expense, Cr
            # AP-vendor) is deferred. The invoice is created posting_status=
            # 'unposted' (model default) with zero LedgerEntry rows; a journal
            # run (POST /api/accounting/journal/run/) posts it later. Inventory
            # batches above are still created immediately — stock-on-hand is a
            # physical fact independent of when the journal catches up.

            # Settled at creation: record the payment straight away, exactly as
            # the pay endpoint would. Its Dr AP / Cr bank legs post now; the
            # accrual's Cr AP lands whenever the journal run sweeps the
            # purchase date, and the two net to zero AP for this vendor.
            if pay_now and grand_total > 0:
                _record_purchase_payment(
                    invoice, grand_total, cash_account, payment_date, _actor(request),
                    notes='Dibayar saat faktur dibuat',
                    payment_method=payment_method,
                )

            AuditLog.objects.create(
                performed_by=_actor(request),
                action='CREATE',
                entity_type='PurchaseInvoice',
                entity_id=str(invoice.id),
                description=f'Purchase invoice created: {invoice.internal_id}',
            )

        return Response(
            PurchaseInvoiceDetailSerializer(invoice, context={'request': request}).data,
            status=status.HTTP_201_CREATED,
        )


class PurchaseInvoiceDetailView(APIView):
    def _get(self, pk):
        try:
            return PurchaseInvoice.objects.select_related(
                'supplier', 'payment_account', 'payment_method', 'payment_method__linked_account',
                'created_by', 'warehouse',
            ).prefetch_related(
                'items__item', 'items__expense_account', 'additional_costs',
                'payments__payment_account', 'payments__payment_method__linked_account',
                'payments__created_by',
            ).get(pk=pk)
        except PurchaseInvoice.DoesNotExist:
            return None

    def get(self, request, pk):
        obj = self._get(pk)
        if not obj:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(PurchaseInvoiceDetailSerializer(obj, context={'request': request}).data)

    def patch(self, request, pk):
        """Update metadata only (notes, dates, invoice number). Items require PUT."""
        obj = self._get(pk)
        if not obj:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if obj.is_voided:
            return Response(
                {'error': 'Cannot edit a voided invoice.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        updatable = ['external_invoice_no', 'notes', 'due_date', 'purchase_date']
        for field in updatable:
            if field in request.data:
                setattr(obj, field, request.data[field] or (None if field == 'due_date' else ''))
        obj.save(update_fields=[f for f in updatable if f in request.data])

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='PurchaseInvoice',
            entity_id=str(obj.id),
            description=f'Purchase invoice metadata updated: {obj.internal_id}',
        )
        return Response(PurchaseInvoiceDetailSerializer(obj, context={'request': request}).data)

    @transaction.atomic
    def put(self, request, pk):
        """Full replacement of invoice items, costs, and header fields. Only allowed when amount_paid == 0."""
        obj = self._get(pk)
        if not obj:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if obj.is_voided:
            return Response(
                {'error': 'Cannot edit a voided invoice.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if obj.amount_paid > 0:
            return Response(
                {'error': 'Cannot edit an invoice that has payments recorded.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Phase 2: whether this invoice's transaction date has already been
        # posted by a journal run decides how the accounting side of this edit
        # is applied — see the branch after the new lines are built below.
        was_posted = obj.posting_status == 'posted'

        # Snapshot the units this invoice's batches have already sold, per (item,
        # warehouse). A batch's (initial − remaining) is how many of its units
        # sales consumed, and (value / initial) × sold is the exact cost those
        # sales posted (mirrors _fifo_deduct). We preserve that portion as a
        # fully-drawn "frozen" batch and post a price-variance for it rather than
        # reversing it, so the sale's COGS is never disturbed and the journal
        # stays balanced. (Replaces the old hard block on consumed inventory.)
        sold_by_key = {}   # (item_id, warehouse_id) -> {'qty', 'cost', 'date'}
        for b in obj.inventory_batches.all():
            sold = b.quantity_initial - b.quantity_remaining
            if sold <= 0:
                continue
            key = (b.item_id, b.warehouse_id)
            per_unit = (b.value / b.quantity_initial) if b.quantity_initial else Decimal('0')
            agg = sold_by_key.setdefault(
                key, {'qty': Decimal('0'), 'cost': Decimal('0'), 'date': b.input_date})
            agg['qty']  += sold
            agg['cost'] += per_unit * sold
            if b.input_date < agg['date']:
                agg['date'] = b.input_date

        data = request.data

        items_raw = data.get('items', [])
        if isinstance(items_raw, str):
            try:
                items = json.loads(items_raw)
            except (json.JSONDecodeError, ValueError):
                return Response({'error': 'Invalid items JSON.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            items = items_raw

        additional_costs_raw = data.get('additional_costs', [])
        if isinstance(additional_costs_raw, str):
            try:
                additional_costs = json.loads(additional_costs_raw)
            except (json.JSONDecodeError, ValueError):
                return Response({'error': 'Invalid additional_costs JSON.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            additional_costs = additional_costs_raw

        if not items:
            return Response({'error': 'At least one item is required.'}, status=status.HTTP_400_BAD_REQUEST)

        # Resolve supplier. An edit is only allowed while nothing has been paid,
        # so the invoice has no payment method to carry — one is only attached
        # when a payment is recorded.
        try:
            supplier = Supplier.objects.get(pk=data['supplier'])
        except (Supplier.DoesNotExist, KeyError):
            return Response({'error': 'Invalid supplier.'}, status=status.HTTP_400_BAD_REQUEST)

        warehouse_id = data.get('warehouse') or None
        warehouse_obj = None
        if warehouse_id:
            try:
                warehouse_obj = Warehouse.objects.get(pk=warehouse_id)
            except Warehouse.DoesNotExist:
                return Response({'error': 'Invalid warehouse.'}, status=status.HTTP_400_BAD_REQUEST)

        for row in items:
            if row.get('line_type', 'stock') == 'expense' and not (row.get('expense_account')):
                return Response(
                    {'error': 'Setiap baris beban harus memiliki akun beban.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Refuse to shrink a line below what sales have already consumed — those
        # units cannot be un-sold by editing the purchase. Checked before any
        # mutation so the 400 leaves the invoice untouched.
        new_qty_by_key = {}
        for row in items:
            if row.get('line_type', 'stock') != 'stock':
                continue
            raw_item = row.get('item')
            if not raw_item:
                continue
            try:
                k_item = int(raw_item)
                k_wh   = int(row.get('warehouse') or warehouse_id)
            except (TypeError, ValueError):
                continue
            q = _safe_decimal(row.get('quantity', 0))
            if q > 0:
                new_qty_by_key[(k_item, k_wh)] = new_qty_by_key.get((k_item, k_wh), Decimal('0')) + q

        for key, agg in sold_by_key.items():
            if new_qty_by_key.get(key, Decimal('0')) < agg['qty']:
                itm = InventoryItem.objects.filter(pk=key[0]).first()
                nm = itm.name if itm else f'#{key[0]}'
                return Response(
                    {'error': f'Tidak dapat mengurangi jumlah "{nm}" di bawah {agg["qty"]:g} unit '
                              f'yang sudah terjual. Batalkan penjualan terkait lebih dulu.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # Reverse and drop the prior accrual postings before re-posting, but only
        # when nothing has been posted yet for this invoice — _unpost_purchase
        # deletes the LedgerEntry rows it finds, which would destroy the
        # original transaction date's history for an already-posted invoice.
        # That case is handled below instead, via a same-day edit memo that
        # leaves the original rows alone.
        if not was_posted:
            _unpost_purchase(obj)

        # Remove old inventory batches and line items
        obj.inventory_batches.all().delete()
        obj.items.all().delete()
        obj.additional_costs.all().delete()

        # Update header fields
        obj.supplier = supplier
        obj.warehouse = warehouse_obj
        obj.purchase_date = data.get('purchase_date', obj.purchase_date)
        obj.due_date = data.get('due_date') or None
        obj.external_invoice_no = data.get('external_invoice_no', obj.external_invoice_no)
        obj.notes = data.get('notes', obj.notes)
        obj.save(update_fields=[
            'supplier', 'warehouse',
            'purchase_date', 'due_date', 'external_invoice_no', 'notes',
        ])

        # Update invoice image if provided
        image_file = request.FILES.get('invoice_image')
        if image_file:
            obj.invoice_image = image_file
            obj.save(update_fields=['invoice_image'])

        # Parse items and compute subtotals (same logic as POST)
        parsed = []
        items_subtotal = Decimal('0')
        total_units = Decimal('0')

        for row in items:
            qty            = _safe_decimal(row.get('quantity', 0))
            cost           = _safe_decimal(row.get('unit_cost', 0))
            total_discount = _safe_decimal(row.get('total_discount', 0))
            line_type      = row.get('line_type', 'stock')
            item_id        = row.get('item') or None
            expense_acct_id = row.get('expense_account') or None

            row_wh_id = row.get('warehouse') or warehouse_id or None

            gross           = qty * cost
            discount_capped = min(total_discount, gross)
            adjusted_sub    = gross - discount_capped

            items_subtotal += adjusted_sub
            if line_type == 'stock' and qty > 0:
                total_units += qty

            parsed.append({
                'line_type':       line_type,
                'item_id':         item_id,
                'item_name':       row.get('item_name', ''),
                'quantity':        qty,
                'unit':            row.get('unit', ''),
                'unit_cost':       cost,
                'total_discount':  discount_capped,
                'adjusted_sub':    adjusted_sub,
                'expense_acct_id': expense_acct_id,
                'warehouse_id':    row_wh_id,
            })

        # Compute net adjustment from additional costs
        running_total  = items_subtotal
        net_adjustment = Decimal('0')
        cost_objs      = []

        for i, ac in enumerate(additional_costs):
            name        = str(ac.get('name', '')).strip()
            modifier    = ac.get('modifier', 'add')
            amount_type = ac.get('amount_type', 'cash')
            amount      = _safe_decimal(ac.get('amount', 0))

            if not name or amount <= 0:
                continue

            if amount_type == 'percent':
                adj = running_total * amount / Decimal('100')
            else:
                adj = amount

            if modifier == 'subtract':
                adj = -adj

            running_total  += adj
            net_adjustment += adj

            cost_objs.append(PurchaseAdditionalCost(
                invoice=obj,
                name=name,
                modifier=modifier,
                amount_type=amount_type,
                amount=amount,
                sort_order=i,
            ))

        if cost_objs:
            PurchaseAdditionalCost.objects.bulk_create(cost_objs)

        grand_total  = items_subtotal + net_adjustment
        per_unit_adj = Decimal('0')
        if total_units > 0:
            per_unit_adj = net_adjustment / total_units

        # Create new item records and batches
        item_objs         = []
        batches_to_create = []

        for p in parsed:
            qty = p['quantity']
            if qty > 0:
                base_unit_cost = p['adjusted_sub'] / qty
                actual = base_unit_cost + (per_unit_adj if p['line_type'] == 'stock' else Decimal('0'))
            else:
                base_unit_cost = p['unit_cost']
                actual = p['unit_cost']

            item_objs.append(PurchaseInvoiceItem(
                invoice=obj,
                line_type=p['line_type'],
                item_id=p['item_id'],
                item_name=p['item_name'],
                quantity=qty,
                unit=p['unit'],
                unit_cost=p['unit_cost'],
                total_discount=p['total_discount'],
                actual_unit_cost=actual,
                expense_account_id=p['expense_acct_id'],
                warehouse_id=p['warehouse_id'],
            ))

            if p['line_type'] == 'stock' and p['item_id'] and p['warehouse_id'] and qty > 0:
                batches_to_create.append({
                    'item_id':      p['item_id'],
                    'warehouse_id': p['warehouse_id'],
                    'qty':          qty,
                    'cost':         actual,
                })

        PurchaseInvoiceItem.objects.bulk_create(item_objs)

        # Frozen batches: units already sold, re-created fully-drawn (remaining=0)
        # at the exact cost the sale expensed. FIFO never touches them; they only
        # carry the historical cost anchor forward across future edits. value is
        # the TOTAL, so value / quantity_initial recovers that anchor per unit.
        for key, agg in sold_by_key.items():
            InventoryBatch.objects.create(
                item_id=key[0],
                warehouse_id=key[1],
                input_date=agg['date'],
                quantity_initial=agg['qty'],
                quantity_remaining=Decimal('0'),
                value=agg['cost'],
                purchase_invoice=obj,
            )

        # On-hand batches for the remaining (unsold) units at the new cost. The
        # already-sold quantity per key is subtracted so we never re-stock units
        # that have left. value is the TOTAL batch value (per-unit × on-hand qty).
        alloc = {k: v['qty'] for k, v in sold_by_key.items()}
        for b in batches_to_create:
            key = (int(b['item_id']), int(b['warehouse_id']))
            take = min(alloc.get(key, Decimal('0')), b['qty'])
            alloc[key] = alloc.get(key, Decimal('0')) - take
            onhand = b['qty'] - take
            if onhand <= 0:
                continue
            InventoryBatch.objects.create(
                item_id=b['item_id'],
                warehouse_id=b['warehouse_id'],
                input_date=obj.purchase_date,
                quantity_initial=onhand,
                quantity_remaining=onhand,
                value=b['cost'] * onhand,
                purchase_invoice=obj,
            )

        grand_total = grand_total.quantize(Decimal('0.01'))
        obj.total_amount = grand_total
        obj.save(update_fields=['total_amount'])

        # Phase 2: an unposted invoice (posting_status='unposted', never swept
        # by a journal run) stays unposted — no LedgerEntry rows are written
        # here; the next run posts the edited values (including any
        # price-variance on units already FIFO-consumed from its batches,
        # since sold_by_key is derived straight from batch state, not from
        # posting_status). An already-posted invoice instead gets a same-day
        # "edit memo": every existing (non-memo) ledger row tied to it is
        # reversed today, tagged source_type='edit_memo', then the accrual +
        # any price-variance correction for the new lines is reposted today
        # under the same tag. The original purchase_date rows and
        # JournalDayLog are left untouched.
        if was_posted:
            memo_today = timezone.now().date()
            # Reverse *everything* currently posted for this invoice — the
            # original accrual AND any earlier edit-memo correction — so a
            # second (or third) edit re-derives the full state from scratch
            # instead of layering an incremental delta on top of a stale one.
            # *Every* row attached to the invoice, memos included: their sum is
            # its current posted state, so reversing all of them nets it to
            # zero. (A restored invoice carries a void_memo/restore_memo pair;
            # skipping the void_memo would leave its reversal stranded.)
            old_entries = list(
                LedgerEntry.objects.filter(purchase_invoice=obj)
                .select_related('account')
            )
            for e in old_entries:
                opp = 'credit' if e.entry_type == 'debit' else 'debit'
                _post_le(e.account, opp, e.amount, obj,
                         f'Koreksi edit {obj.internal_id}: {e.description}',
                         memo_today, source_type='edit_memo')
            _post_purchase_accrual(obj, parsed, item_objs, supplier, grand_total,
                                    post_date=memo_today, source_type='edit_memo')
            if sold_by_key:
                _post_purchase_price_variance(obj, sold_by_key, item_objs,
                                               post_date=memo_today, source_type='edit_memo')

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='PurchaseInvoice',
            entity_id=str(obj.id),
            description=f'Purchase invoice fully updated: {obj.internal_id}',
        )

        obj.refresh_from_db()
        return Response(PurchaseInvoiceDetailSerializer(
            self._get(pk), context={'request': request}
        ).data)

    @transaction.atomic
    def delete(self, request, pk):
        """Void (soft-delete) a purchase invoice: reverse stock and accounting but keep the record."""
        obj = self._get(pk)
        if not obj:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if obj.is_voided:
            return Response(
                {'error': 'Invoice is already voided.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if obj.amount_paid > 0:
            return Response(
                {'error': 'Cannot void an invoice that has payments recorded.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Block voiding if any inventory from this invoice has already been partially consumed
        consumed_batches = obj.inventory_batches.filter(
            quantity_remaining__lt=F('quantity_initial')
        )
        if consumed_batches.exists():
            return Response(
                {'error': 'Cannot void: some inventory from this invoice has already been used.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        today       = timezone.now().date()
        internal_id = obj.internal_id

        # Phase 2: an unposted invoice has no LedgerEntry rows at all (posting
        # is deferred until a journal run), so this loop is naturally a no-op
        # for it — void with zero journal impact, as required. A posted
        # invoice's rows (the original accrual AND any earlier edit-, void- or
        # restore-memo — everything currently in effect) are reversed here as a
        # same-day "void memo" (dated today, tagged source_type='void_memo'),
        # leaving the original transaction date's rows and JournalDayLog
        # untouched. Reversing the whole set is what makes the net zero, so an
        # invoice voided → restored → voided again lands back on nothing.
        # Safe either way: the guard above refused any paid invoice.
        for e in list(
            LedgerEntry.objects.filter(purchase_invoice=obj)
            .select_related('account')
        ):
            opp = 'credit' if e.entry_type == 'debit' else 'debit'
            _post_le(e.account, opp, e.amount, obj,
                     f'Pembatalan {internal_id}: {e.description}', today, source_type='void_memo')

        # Remove inventory batches created from this invoice (stock never came in)
        obj.inventory_batches.all().delete()

        obj.is_voided = True
        obj.voided_at = timezone.now()
        obj.voided_by = _actor(request)
        obj.save(update_fields=['is_voided', 'voided_at', 'voided_by'])

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='VOID',
            entity_type='PurchaseInvoice',
            entity_id=str(pk),
            description=f'Purchase invoice voided: {internal_id} — inventory reversed, journal updated',
        )
        return Response(PurchaseInvoiceDetailSerializer(obj, context={'request': request}).data)


class PurchaseInvoiceRestoreView(APIView):
    """
    POST /api/accounting/purchases/<pk>/restore/

    Un-void a cancelled purchase invoice — the inverse of ``DELETE`` on the
    detail view. The record itself was never deleted, so nothing is re-entered:
    the stored lines, costs and totals come back as they were, under the same
    ``internal_id``.

    Two effects to undo:

    * **Stock.** The void deleted this invoice's inventory batches, and it could
      only do so because none of them had been consumed. They are recreated
      here from the saved lines at their ``actual_unit_cost``, dated the
      original ``purchase_date`` so FIFO order is unchanged.
    * **Journal.** An invoice that was never posted has no ledger rows at all —
      restoring it is a pure no-op there, and the next journal run sweeps it
      (``is_voided=False`` is in the sweep filter). A posted one gets a same-day
      "restore memo": every row currently attached to the invoice is reversed
      (their sum *is* the current state, so this nets it to zero) and the
      accrual is re-posted, all dated today and tagged
      ``source_type='restore_memo'``. The original transaction date's rows and
      ``JournalDayLog`` are left untouched, exactly as the void/edit memos do.
    """

    @transaction.atomic
    def post(self, request, pk):
        try:
            invoice = PurchaseInvoice.objects.select_related(
                'supplier', 'supplier__ap_account',
            ).get(pk=pk)
        except PurchaseInvoice.DoesNotExist:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        if not invoice.is_voided:
            return Response(
                {'error': 'Faktur ini tidak dibatalkan.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        items = list(invoice.items.all())
        if not items:
            return Response(
                {'error': 'Faktur tanpa baris tidak dapat dipulihkan.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # A void is refused while any payment exists, so a voided invoice can
        # only be unpaid. If that ever stops holding, restoring would re-post
        # the accrual without the payment legs — refuse instead of guessing.
        if invoice.amount_paid > 0:
            return Response(
                {'error': 'Faktur dengan pembayaran tidak dapat dipulihkan.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        today = timezone.now().date()

        # Put the stock back. The void left no batches behind; the delete is
        # defensive so a half-finished earlier attempt cannot double the stock.
        invoice.inventory_batches.all().delete()
        for it in items:
            if (it.line_type != 'stock' or not it.item_id
                    or not it.warehouse_id or it.quantity <= 0):
                continue
            InventoryBatch.objects.create(
                item_id=it.item_id,
                warehouse_id=it.warehouse_id,
                input_date=invoice.purchase_date,
                quantity_initial=it.quantity,
                quantity_remaining=it.quantity,
                value=it.actual_unit_cost * it.quantity,
                purchase_invoice=invoice,
            )

        # Rebuild the (parsed, item_objs) pair the accrual builder expects,
        # straight from the stored lines — same order, so build_purchase_legs
        # can zip them. Nothing is recomputed: the totals were frozen when the
        # invoice was created or last edited.
        parsed = [{
            'line_type':       it.line_type,
            'item_id':         it.item_id,
            'item_name':       it.item_name,
            'quantity':        it.quantity,
            'unit':            it.unit,
            'unit_cost':       it.unit_cost,
            'total_discount':  it.total_discount,
            'adjusted_sub':    it.quantity * it.unit_cost - it.total_discount,
            'expense_acct_id': it.expense_account_id,
            'warehouse_id':    it.warehouse_id,
        } for it in items]

        existing = list(
            LedgerEntry.objects.filter(purchase_invoice=invoice).select_related('account')
        )
        if existing:
            for e in existing:
                opp = 'credit' if e.entry_type == 'debit' else 'debit'
                _post_le(e.account, opp, e.amount, invoice,
                         f'Pemulihan {invoice.internal_id}: {e.description}',
                         today, source_type='restore_memo')
            _post_purchase_accrual(
                invoice, parsed, items, invoice.supplier, invoice.total_amount,
                post_date=today, source_type='restore_memo',
            )

        invoice.is_voided = False
        invoice.voided_at = None
        invoice.voided_by = None
        invoice.save(update_fields=['is_voided', 'voided_at', 'voided_by'])

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='RESTORE',
            entity_type='PurchaseInvoice',
            entity_id=str(invoice.pk),
            description=f'Purchase invoice restored: {invoice.internal_id} — '
                        f'inventory returned, journal updated',
        )
        return Response(
            PurchaseInvoiceDetailSerializer(invoice, context={'request': request}).data
        )


class PurchaseInvoicePayView(APIView):
    """
    POST /api/accounting/purchases/<pk>/pay/
    Body: { amount, payment_date, cash_account, notes? }

    Records a payment against the purchase invoice: which bank/cash account it
    came out of and the date it left. Debits the vendor's Accounts Payable and
    credits the chosen cash/bank account. ``payment_account``/``payment_method``
    (a PaymentMethod pk) are still accepted for back-compat.
    """

    def post(self, request, pk):
        try:
            invoice = PurchaseInvoice.objects.select_related(
                'payment_account', 'payment_method', 'payment_method__linked_account',
                'supplier', 'supplier__ap_account',
            ).get(pk=pk)
        except PurchaseInvoice.DoesNotExist:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        if invoice.is_voided:
            return Response({'error': 'Cannot pay a voided invoice.'}, status=status.HTTP_400_BAD_REQUEST)

        if invoice.status == 'paid':
            return Response({'error': 'Invoice is already fully paid.'}, status=status.HTTP_400_BAD_REQUEST)

        amount = _safe_decimal(request.data.get('amount', 0))
        if amount <= 0:
            return Response({'error': 'Payment amount must be greater than zero.'}, status=status.HTTP_400_BAD_REQUEST)

        remaining = invoice.total_amount - invoice.amount_paid
        if amount > remaining:
            return Response(
                {'error': f'Payment exceeds remaining balance of Rp {remaining:,.0f}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Which account the money came from, and when it left. Both are asked
        # for at payment time; the account falls back to whatever settled this
        # invoice last so a repeat instalment needs no re-picking.
        cash_account, payment_method, cash_err = _resolve_purchase_cash_account(request.data)
        if cash_err is not None:
            return cash_err
        if cash_account is None:
            cash_account = invoice.payment_account
        if cash_account is None:
            return Response(
                {'error': 'Rekening kas/bank wajib dipilih.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        pay_date = _parse_date(request.data.get('payment_date'))
        if pay_date is None:
            return Response(
                {'error': 'Tanggal pembayaran wajib diisi.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if pay_date < invoice.purchase_date:
            return Response(
                {'error': 'Tanggal pembayaran tidak boleh sebelum tanggal pembelian.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            _record_purchase_payment(
                invoice, amount, cash_account, pay_date, _actor(request),
                notes=str(request.data.get('notes', '') or '')[:255],
                payment_method=payment_method,
            )

            AuditLog.objects.create(
                performed_by=_actor(request),
                action='UPDATE',
                entity_type='PurchaseInvoice',
                entity_id=str(invoice.id),
                description=(f'Payment Rp{amount:,.0f} from {cash_account.name} on {pay_date} '
                             f'recorded for {invoice.internal_id}, status → {invoice.status}'),
            )

        return Response(PurchaseInvoiceDetailSerializer(invoice, context={'request': request}).data)


class PurchaseLastPriceView(APIView):
    """GET /api/accounting/purchases/last-price/?item=<id>"""

    def get(self, request):
        item_id = request.query_params.get('item', '').strip()
        if not item_id:
            return Response({'error': 'item parameter is required.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            item_id = int(item_id)
        except ValueError:
            return Response({'error': 'Invalid item id.'}, status=status.HTTP_400_BAD_REQUEST)

        last = (
            PurchaseInvoiceItem.objects
            .filter(item_id=item_id, line_type='stock')
            .select_related('invoice')
            .order_by('-invoice__purchase_date', '-id')
            .first()
        )
        if last:
            return Response({
                'last_price': str(last.unit_cost),
                'last_date':  str(last.invoice.purchase_date),
                'invoice_id': last.invoice.internal_id,
            })
        return Response({'last_price': None, 'last_date': None, 'invoice_id': None})


# ── Cash / bank accounts ───────────────────────────────────────────────────────

class CashAccountListView(APIView):
    """
    GET /api/accounting/cash-accounts/
    → [{ id, name, account_number, balance }] sorted by name

    The curated set of COA rows that represent real cash/bank locations — see
    ``services.cash_accounts.cash_bank_account_ids``. ``account_number`` is
    returned for journal previews and printed documents; the picker itself
    shows the name only.
    """

    def get(self, request):
        accounts = (
            ChartOfAccounts.objects
            .filter(pk__in=cash_bank_account_ids())
            .order_by('name')
            .values('id', 'name', 'account_number', 'balance')
        )
        return Response([
            {
                'id':             a['id'],
                'name':           a['name'],
                'account_number': a['account_number'],
                'balance':        str(a['balance']),
            }
            for a in accounts
        ])


# ── Operating Expenses (Phase 3) ────────────────────────────────────────────────

class ExpenseListCreateView(APIView):
    """
    GET  /api/accounting/expenses/
         ?status=unpaid|partial|paid
         ?q=<search memo / notes / line description / line account name>
         ?date_from=&date_to=
         ?cash_account=<ChartOfAccounts pk>
         ?source=general|beautician  — separates hand-entered expenses from
             ones written through the beautician petty-cash flow (design doc
             §4). Every row already carries ``source``/``created_by_name``.

    POST /api/accounting/expenses/
         { expense_date, cash_account?, payment_memo?,
           payment_account?, amount_paid?, notes?,
           items: [{account, description?, amount}] }

    ``cash_account`` is the ChartOfAccounts pk the money leaves from and must
    be one of ``cash_bank_account_ids()``. ``payment_account`` is the legacy
    key (a PaymentMethod pk); when only it is sent, its ``linked_account``
    becomes the cash account. Each item's ``description`` is that leg's
    journal memo — leave it blank to inherit ``payment_memo``.
    """

    def get(self, request):
        qs = filter_by_branch(
            Expense.objects.select_related(
                'payment_method', 'payment_method__linked_account', 'payment_account',
            ),
            request,
        )

        q = request.query_params.get('q', '').strip()
        if q:
            # No payee field any more — search the strings that actually
            # identify an expense to a human: its memo, its notes, and the
            # description/account of any of its lines.
            qs = qs.filter(
                Q(payment_memo__icontains=q)
                | Q(notes__icontains=q)
                | Q(items__description__icontains=q)
                | Q(items__account__name__icontains=q)
            ).distinct()
        st = request.query_params.get('status', '').strip()
        if st in ('unpaid', 'partial', 'paid'):
            qs = qs.filter(status=st)
        date_from = request.query_params.get('date_from', '').strip()
        date_to = request.query_params.get('date_to', '').strip()
        if date_from:
            qs = qs.filter(expense_date__gte=date_from)
        if date_to:
            qs = qs.filter(expense_date__lte=date_to)
        cash_account = request.query_params.get('cash_account', '').strip()
        if cash_account:
            qs = qs.filter(payment_account_id=cash_account)
        src = request.query_params.get('source', '').strip()
        if src in ('general', 'beautician'):
            qs = qs.filter(source=src)

        return Response(
            ExpenseSerializer(
                qs.select_related('created_by').prefetch_related('items__account', 'items__alias'),
                many=True,
            ).data
        )

    def post(self, request):
        data = request.data
        items_raw = data.get('items', [])
        if isinstance(items_raw, str):
            try:
                items = json.loads(items_raw)
            except (json.JSONDecodeError, ValueError):
                return Response({'error': 'Invalid items JSON.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            items = items_raw

        if not items:
            return Response({'error': 'At least one item is required.'}, status=status.HTTP_400_BAD_REQUEST)

        expense_date = data.get('expense_date')
        if not expense_date:
            return Response({'error': 'Tanggal wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        payment_method = None
        payment_account_id = data.get('payment_account') or None
        if payment_account_id:
            try:
                payment_method = PaymentMethod.objects.select_related('linked_account').get(pk=payment_account_id)
            except PaymentMethod.DoesNotExist:
                return Response({'error': 'Invalid payment method.'}, status=status.HTTP_400_BAD_REQUEST)

        cash_account, err = _resolve_cash_account(data)
        if err:
            return err
        if cash_account is None and payment_method is not None:
            # Legacy caller sent only payment_account — resolve the COA behind it.
            cash_account = payment_method.linked_account

        for row in items:
            if not row.get('account'):
                return Response(
                    {'error': 'Setiap baris beban harus memiliki akun.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # The write itself lives in services.expense_create — the beautician
        # petty-cash flow (views/beautician_expense_page.py) calls the exact
        # same function, so there is one expense-creation code path, not two.
        expense = create_expense(
            expense_date=expense_date,
            payment_method=payment_method,
            payment_account=cash_account,
            payment_memo=data.get('payment_memo'),
            notes=data.get('notes', ''),
            amount_paid=data.get('amount_paid', 0),
            items=items,
            actor=_actor(request),
            branch=write_branch(request),
        )

        return Response(
            ExpenseSerializer(expense).data,
            status=status.HTTP_201_CREATED,
        )


class ExpenseDetailView(APIView):
    def _get(self, pk):
        try:
            return Expense.objects.select_related(
                'payment_method', 'payment_method__linked_account',
                'payment_account', 'created_by',
            ).prefetch_related('items__account', 'items__alias').get(pk=pk)
        except Expense.DoesNotExist:
            return None

    def get(self, request, pk):
        obj = self._get(pk)
        if not obj:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(ExpenseSerializer(obj).data)

    @transaction.atomic
    def put(self, request, pk):
        """Full replacement of expense items and header fields. Only allowed when amount_paid == 0."""
        obj = self._get(pk)
        if not obj:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if obj.amount_paid > 0:
            return Response(
                {'error': 'Cannot edit an expense that has payments recorded.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        was_posted = obj.posting_status == 'posted'

        data = request.data
        items_raw = data.get('items', [])
        if isinstance(items_raw, str):
            try:
                items = json.loads(items_raw)
            except (json.JSONDecodeError, ValueError):
                return Response({'error': 'Invalid items JSON.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            items = items_raw

        if not items:
            return Response({'error': 'At least one item is required.'}, status=status.HTTP_400_BAD_REQUEST)

        payment_method = obj.payment_method
        payment_method_sent = False
        payment_account_id = data.get('payment_account')
        if payment_account_id:
            try:
                payment_method = PaymentMethod.objects.select_related('linked_account').get(pk=payment_account_id)
            except PaymentMethod.DoesNotExist:
                return Response({'error': 'Invalid payment method.'}, status=status.HTTP_400_BAD_REQUEST)
            payment_method_sent = True
        elif 'payment_account' in data:
            payment_method = None
            payment_method_sent = True

        cash_account, err = _resolve_cash_account(data, current=obj.payment_account)
        if err:
            return err
        if 'cash_account' not in data and payment_method_sent:
            # Legacy caller sent only payment_account — keep the COA in step
            # with the method it just chose (or cleared).
            cash_account = payment_method.linked_account if payment_method else None

        for row in items:
            if not row.get('account'):
                return Response(
                    {'error': 'Setiap baris beban harus memiliki akun.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        if not was_posted:
            _unpost_expense(obj)

        obj.items.all().delete()

        obj.payment_method = payment_method
        obj.payment_account = cash_account
        if 'payment_memo' in data:
            obj.payment_memo = (data.get('payment_memo') or '')[:255]
        obj.expense_date = data.get('expense_date', obj.expense_date)
        obj.notes = data.get('notes', obj.notes)

        item_objs = []
        total = Decimal('0')
        for row in items:
            amt = _safe_decimal(row.get('amount', 0))
            if amt <= 0:
                continue
            total += amt
            item_objs.append(ExpenseItem(
                expense=obj,
                account_id=row['account'],
                description=row.get('description', ''),
                amount=amt,
            ))
        ExpenseItem.objects.bulk_create(item_objs)

        obj.total_amount = total.quantize(Decimal('0.01'))
        obj.refresh_status()
        obj.save(update_fields=[
            'payment_method', 'payment_account', 'payment_memo',
            'expense_date', 'notes', 'total_amount', 'status',
        ])

        # Same pattern as PurchaseInvoiceDetailView.put: an unposted expense
        # stays unposted (no LedgerEntry rows here; the next journal run posts
        # the edited values). An already-posted expense gets a same-day
        # "edit memo": every existing (non-memo) ledger row tied to it is
        # reversed today, tagged source_type='edit_memo', then the accrual is
        # reposted today under the same tag. The original expense_date rows
        # and JournalDayLog are left untouched.
        if was_posted:
            memo_today = timezone.now().date()
            old_entries = list(
                LedgerEntry.objects.filter(expense=obj)
                .exclude(source_type='void_memo')
                .select_related('account')
            )
            for e in old_entries:
                opp = 'credit' if e.entry_type == 'debit' else 'debit'
                _post_expense_le(e.account, opp, e.amount, obj,
                                  f'Koreksi edit beban #{obj.pk}: {e.description}',
                                  memo_today, source_type='edit_memo')
            _post_expense_accrual(obj, item_objs, post_date=memo_today, source_type='edit_memo')

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='Expense',
            entity_id=str(obj.id),
            description=f'Expense fully updated: #{obj.id}',
        )

        return Response(ExpenseSerializer(self._get(pk)).data)

    @transaction.atomic
    def delete(self, request, pk):
        """Void an expense: reverse accounting but keep the record."""
        obj = self._get(pk)
        if not obj:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if obj.amount_paid > 0:
            return Response(
                {'error': 'Cannot void an expense that has payments recorded.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        today = timezone.now().date()

        # Phase 2/3 pattern: an unposted expense has no LedgerEntry rows at all
        # (posting is deferred until a journal run), so this loop is a no-op
        # for it. A posted expense's rows (the original accrual and any
        # earlier edit-memo correction) are reversed here as a same-day
        # "void memo" (dated today, tagged source_type='void_memo'), leaving
        # the original expense_date rows and JournalDayLog untouched.
        for e in list(
            LedgerEntry.objects.filter(expense=obj)
            .exclude(source_type='void_memo')
            .select_related('account')
        ):
            opp = 'credit' if e.entry_type == 'debit' else 'debit'
            _post_expense_le(e.account, opp, e.amount, obj,
                              f'Pembatalan beban #{obj.pk}: {e.description}', today, source_type='void_memo')

        entity_id = str(obj.id)
        obj.delete()

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='DELETE',
            entity_type='Expense',
            entity_id=entity_id,
            description=f'Expense voided: #{entity_id}',
        )
        return Response(status=status.HTTP_204_NO_CONTENT)


class ExpensePayView(APIView):
    """
    POST /api/accounting/expenses/<pk>/pay/
    Body: { amount, payment_account?, cash_account? }

    Records a payment against the expense. Credits the expense's
    ``payment_account`` (falling back to the GL account behind its
    payment_method for legacy rows) and debits Accounts Payable, settling the
    liability the accrual posting created — mirrors PurchaseInvoicePayView.
    """

    def post(self, request, pk):
        try:
            expense = Expense.objects.select_related(
                'payment_method', 'payment_method__linked_account', 'payment_account',
            ).get(pk=pk)
        except Expense.DoesNotExist:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        if expense.status == 'paid':
            return Response({'error': 'Expense is already fully paid.'}, status=status.HTTP_400_BAD_REQUEST)

        amount = _safe_decimal(request.data.get('amount', 0))
        if amount <= 0:
            return Response({'error': 'Payment amount must be greater than zero.'}, status=status.HTTP_400_BAD_REQUEST)

        remaining = expense.total_amount - expense.amount_paid
        if amount > remaining:
            return Response(
                {'error': f'Payment exceeds remaining balance of Rp {remaining:,.0f}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # A caller may name the cash account outright (new) or hand over a
        # payment method (legacy). Either way the credit leg resolves to a COA,
        # preferring the one already recorded on the expense.
        cash_account, err = _resolve_cash_account(request.data)
        if err:
            return err

        payment_method = None
        payment_method_id = request.data.get('payment_account') or expense.payment_method_id
        if payment_method_id:
            try:
                payment_method = PaymentMethod.objects.select_related('linked_account').get(pk=payment_method_id)
            except (PaymentMethod.DoesNotExist, TypeError, ValueError):
                return Response({'error': 'Invalid payment method.'}, status=status.HTTP_400_BAD_REQUEST)

        if cash_account is None:
            if request.data.get('payment_account') and payment_method is not None:
                # Explicit legacy override — pay out of the method the caller named.
                cash_account = payment_method.linked_account
            else:
                cash_account = expense_credit_account(expense)
        if cash_account is None:
            return Response({'error': 'Invalid payment method.'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            expense.amount_paid += amount
            if not expense.payment_method_id and payment_method is not None:
                expense.payment_method = payment_method
            if not expense.payment_account_id:
                expense.payment_account = cash_account
            expense.refresh_status()
            expense.save(update_fields=[
                'amount_paid', 'status', 'payment_method', 'payment_account',
            ])

            pay_date = timezone.now().date()

            # Double entry: Dr Accounts Payable (settle the liability), Cr
            # cash/bank payment account (funds leave). Only meaningful once
            # the expense's accrual has actually been posted — an unposted
            # expense has no AP leg yet, so this payment simply waits to be
            # netted against the accrual whenever the journal run catches up.
            if expense.posting_status == 'posted':
                ap_account = Supplier._ensure_ap_control_account()
                _post_expense_le(ap_account, 'debit', amount, expense,
                                  f'Pembayaran utang beban #{expense.pk}', pay_date)
                _post_expense_le(cash_account, 'credit', amount, expense,
                                  expense_credit_memo(expense, paid=True), pay_date)

            AuditLog.objects.create(
                performed_by=_actor(request),
                action='UPDATE',
                entity_type='Expense',
                entity_id=str(expense.id),
                description=f'Payment Rp{amount:,.0f} recorded for expense #{expense.id}, status → {expense.status}',
            )

        return Response(ExpenseSerializer(expense).data)


# ── Account Transfers ──────────────────────────────────────────────────────────

class AccountTransferListCreateView(APIView):
    """
    GET  /api/accounting/transfers/?date_from=&date_to=&account=
    POST /api/accounting/transfers/
         { transfer_date, from_account, to_account, amount, description, reference? }

    Both accounts must be cash/bank accounts (``cash_bank_account_ids()``). A
    transfer is a move of funds between two places money physically sits —
    anything else (writing down inventory, reclassifying an expense) is a
    manual adjustment or a correction journal, not a transfer.
    """

    def get(self, request):
        qs = filter_by_branch(
            AccountTransfer.objects.select_related('from_account', 'to_account', 'created_by'),
            request,
        )

        date_from = request.query_params.get('date_from', '').strip()
        date_to   = request.query_params.get('date_to', '').strip()
        acct_id   = request.query_params.get('account', '').strip()

        if date_from:
            qs = qs.filter(transfer_date__gte=date_from)
        if date_to:
            qs = qs.filter(transfer_date__lte=date_to)
        if acct_id:
            qs = qs.filter(Q(from_account_id=acct_id) | Q(to_account_id=acct_id))

        return Response(AccountTransferSerializer(qs, many=True).data)

    def post(self, request):
        data = request.data

        try:
            from_acct = ChartOfAccounts.objects.get(pk=data['from_account'])
            to_acct   = ChartOfAccounts.objects.get(pk=data['to_account'])
        except (ChartOfAccounts.DoesNotExist, KeyError):
            return Response({'error': 'Invalid account(s).'}, status=status.HTTP_400_BAD_REQUEST)

        if from_acct.id == to_acct.id:
            return Response({'error': 'Rekening asal dan tujuan tidak boleh sama.'}, status=status.HTTP_400_BAD_REQUEST)

        cash_ids = cash_bank_account_ids()
        not_cash = [a.name for a in (from_acct, to_acct) if a.pk not in cash_ids]
        if not_cash:
            return Response(
                {'error': 'Transfer hanya diperbolehkan antar rekening kas/bank. '
                          f'Bukan rekening kas/bank: {", ".join(not_cash)}.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        amount = _safe_decimal(data.get('amount', 0))
        if amount <= 0:
            return Response({'error': 'Jumlah harus lebih dari nol.'}, status=status.HTTP_400_BAD_REQUEST)

        description = (data.get('description') or '').strip()
        if not description:
            return Response({'error': 'Keterangan wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        transfer_date = data.get('transfer_date')
        if not transfer_date:
            return Response({'error': 'Tanggal transfer wajib diisi.'}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            transfer = AccountTransfer.objects.create(
                transfer_date=transfer_date,
                from_account=from_acct,
                to_account=to_acct,
                amount=amount,
                description=description,
                reference=data.get('reference', ''),
                created_by=_actor(request),
                branch=write_branch(request),
            )

            # Phase 2: posting is deferred (posting_status='unposted', the model
            # default) — no balance update, no LedgerEntry rows here. A journal
            # run (POST /api/accounting/journal/run/) posts this transfer via
            # journal_engine.post_account_transfer() once it sweeps
            # transfer_date.

            AuditLog.objects.create(
                performed_by=_actor(request),
                action='CREATE',
                entity_type='AccountTransfer',
                entity_id=str(transfer.id),
                description=f'Transfer Rp{amount:,.0f} dari {from_acct} ke {to_acct}',
            )

        return Response(AccountTransferSerializer(transfer).data, status=status.HTTP_201_CREATED)


class AccountTransferDetailView(APIView):
    def _get(self, pk):
        try:
            return AccountTransfer.objects.select_related('from_account', 'to_account', 'created_by').get(pk=pk)
        except AccountTransfer.DoesNotExist:
            return None

    def get(self, request, pk):
        obj = self._get(pk)
        if not obj:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(AccountTransferSerializer(obj).data)

    def delete(self, request, pk):
        """Reverse a transfer: restore both account balances and remove ledger entries.

        Phase 2: an unposted transfer (the common case — see the deferred
        posting note in the POST handler) has no balance effect and no
        LedgerEntry rows, so this is a plain delete. A posted transfer (its
        transfer_date has been swept by a journal run) still has real
        balance/ledger effects, which are restored/removed here exactly as
        before — AccountTransfer has no memo/void-exception path (unlike
        Invoice/PurchaseInvoice, deleting a transfer is a hard delete, not a
        void), so this remains a direct reversal on the original rows.
        """
        obj = self._get(pk)
        if not obj:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Deleting a transfer hard-deletes its ledger rows (see the docstring),
        # so a row that a *completed* reconciliation cleared would vanish out
        # from under a closed period, silently un-matching a statement line
        # nobody will look at again. A closed reconciliation is evidence; reopen
        # it deliberately if this transfer really was wrong.
        locked = LedgerEntry.objects.filter(
            transfer=obj, reconciliation__status='completed',
        ).exists()
        if locked:
            return Response(
                {'error': 'Transfer ini sudah termasuk dalam rekonsiliasi bank '
                          'yang telah diselesaikan. Buka kembali rekonsiliasinya '
                          'terlebih dahulu.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            if obj.posting_status == 'posted':
                # Restore balances
                ChartOfAccounts.objects.filter(pk=obj.from_account_id).update(balance=obj.from_account.balance + obj.amount)
                ChartOfAccounts.objects.filter(pk=obj.to_account_id).update(balance=obj.to_account.balance - obj.amount)
                # Remove ledger entries for this transfer
                LedgerEntry.objects.filter(transfer=obj).delete()

            desc = str(obj)
            obj.delete()
            AuditLog.objects.create(
                performed_by=_actor(request),
                action='DELETE',
                entity_type='AccountTransfer',
                entity_id=str(pk),
                description=f'Transfer reversed: {desc}',
            )

        return Response(status=status.HTTP_204_NO_CONTENT)


# ── Manual Journal Adjustment ──────────────────────────────────────────────────

class JournalAdjustmentView(APIView):
    """
    POST /api/accounting/adjustments/  — **retired**, always 410.

    This wrote a single LedgerEntry row with no counterpart. It was wrong twice
    over: one row cannot balance, so every adjustment pushed the trial balance
    out by its own amount permanently; and it moved the cached balance by
    ``+amount`` for a debit and ``-amount`` for a credit regardless of the
    account's normal balance, so a credit to a revenue or liability account moved
    that account the wrong way.

    Manual entries now go to ``POST /api/accounting/manual-journal/``, which
    requires both sides, names the transaction, and posts through ``write_legs``
    like every other journal document. The URL is kept (rather than deleted) so
    an older client gets this explanation instead of a 404.
    """

    RETIRED_MESSAGE = (
        'Penyesuaian satu sisi sudah tidak digunakan karena membuat jurnal tidak '
        'seimbang. Gunakan Entri Jurnal Manual (/accounting/journal/manual), yang '
        'mewajibkan sisi debit dan kredit.'
    )

    def post(self, request):
        return Response(
            {'error': self.RETIRED_MESSAGE},
            status=status.HTTP_410_GONE,
        )


# ── Journal Run (Phase 2 batch posting) ────────────────────────────────────────

def _post_invoice_for_run(invoice):
    """Rebuild the (item, item_name, quantity, price) lines from the InvoiceItem
    rows already saved for this invoice and post it via the same
    _post_accounting used by the (now historic) synchronous create/edit path,
    with today's real transaction date — no date/source_type override, this is
    a normal posting, not a memo."""
    from .invoice_page import _post_accounting  # local import: avoids a cycle at module load

    items = list(
        invoice.items
        .select_related('item__item_category__revenue_account')
        .all()
    )
    items_by_id = {it.item_id: it.item for it in items if it.item_id}
    lines_as_dicts = [
        {'item_id': it.item_id, 'item_name': it.item_name, 'quantity': it.quantity, 'price': it.price}
        for it in items
    ]
    _post_accounting(invoice, lines_as_dicts, items_by_id)
    invoice.posting_status = 'posted'
    invoice.save(update_fields=['posting_status'])


def _post_purchase_for_run(pinv):
    """Rebuild the ``parsed``/``item_objs`` shape _post_purchase_accrual expects
    from the PurchaseInvoiceItem rows already saved for this invoice."""
    items = list(pinv.items.select_related('item', 'expense_account', 'warehouse').all())
    parsed = [{
        'line_type':       it.line_type,
        'item_id':         it.item_id,
        'item_name':       it.item_name,
        'quantity':        it.quantity,
        'unit':            it.unit,
        'unit_cost':       it.unit_cost,
        'total_discount':  it.total_discount,
        'adjusted_sub':    it.quantity * it.actual_unit_cost,
        'expense_acct_id': it.expense_account_id,
        'warehouse_id':    it.warehouse_id,
    } for it in items]
    _post_purchase_accrual(pinv, parsed, items, pinv.supplier, pinv.total_amount)
    pinv.posting_status = 'posted'
    pinv.save(update_fields=['posting_status'])


def _post_expense_for_run(expense):
    """Post an unposted Expense's accrual using the ExpenseItem rows already
    saved for it, exactly like _post_purchase_for_run does for PurchaseInvoice."""
    items = list(expense.items.select_related('account').all())
    _post_expense_accrual(expense, items)
    expense.posting_status = 'posted'
    expense.save(update_fields=['posting_status'])


def _build_journal_summary(date_from, date_to):
    """Aggregate what a journal run just posted, for JournalBatch.summary:
    total revenue, spend by expense/COGS account, outstanding AP total, and
    per-account debit/credit movement — all restricted to [date_from, date_to].
    """
    entries = (
        LedgerEntry.objects
        .filter(date__gte=date_from, date__lte=date_to)
        .select_related('account')
    )

    total_revenue = Decimal('0')
    spend_by_account = {}
    movement = {}

    for e in entries:
        acc = e.account
        mv = movement.setdefault(acc.id, {
            'account_id': acc.id, 'account_number': acc.account_number,
            'name': acc.name, 'account_type': acc.account_type,
            'debit': Decimal('0'), 'credit': Decimal('0'),
        })
        mv[e.entry_type] += e.amount

        if acc.account_type == 'revenue':
            total_revenue += e.amount if e.entry_type == 'credit' else -e.amount

        if acc.account_type in ('expense', 'cogs'):
            bucket = spend_by_account.setdefault(acc.id, {
                'account_id': acc.id, 'account_number': acc.account_number,
                'name': acc.name, 'amount': Decimal('0'),
            })
            bucket['amount'] += e.amount if e.entry_type == 'debit' else -e.amount

    # Accounts payable total — sum of unpaid/partial PurchaseInvoice balances
    # plus unpaid/partial Expense balances (Phase 3 — both post their unpaid
    # portion to the same AP control account).
    ap_total = (
        PurchaseInvoice.objects
        .filter(status__in=['unpaid', 'partial'], is_voided=False)
        .aggregate(s=Sum('total_amount') - Sum('amount_paid'))['s']
    ) or Decimal('0')
    expense_ap_total = (
        Expense.objects
        .filter(status__in=['unpaid', 'partial'])
        .aggregate(s=Sum('total_amount') - Sum('amount_paid'))['s']
    ) or Decimal('0')
    ap_total += expense_ap_total

    return {
        'total_revenue': str(total_revenue.quantize(CENT)),
        'accounts_payable_total': str(ap_total.quantize(CENT)),
        'spend_by_account': [
            {**v, 'amount': str(v['amount'].quantize(CENT))}
            for v in sorted(spend_by_account.values(), key=lambda m: m['account_number'])
        ],
        'account_movements': [
            {**v, 'debit': str(v['debit'].quantize(CENT)), 'credit': str(v['credit'].quantize(CENT))}
            for v in sorted(movement.values(), key=lambda m: m['account_number'])
        ],
    }


def _post_transfer_for_run(transfer):
    """Post an account transfer and flip it to 'posted'."""
    post_account_transfer(transfer)
    transfer.posting_status = 'posted'
    transfer.save(update_fields=['posting_status'])


def _post_stock_correction_for_run(log):
    """Post a stock correction and flip it to 'posted'.

    Charges the FIFO cost captured at deduction time to the account its reason
    maps to, crediting Inventory. ``build_stock_correction_legs`` returns an
    empty LegSet for anything with no P&L effect, and ``write_legs`` treats that
    as nothing to post — the row is still marked 'posted' so the sweep stops
    reconsidering it.
    """
    legset = build_stock_correction_legs(log)
    write_legs(legset, date=log.out_date, source_type='stock', document=log)
    log.posting_status = 'posted'
    log.save(update_fields=['posting_status'])


def _post_sales_return_for_run(sales_return):
    """Post a sales return and flip it to 'posted'.

    ``build_sales_return_legs(deduct=True)`` is what actually puts the stock
    back on the shelf — the FIFO cost the COGS reversal needs cannot be known
    without doing the restock, exactly as the invoice's FIFO deduction cannot be
    known without doing the deduction. So the physical movement happens here, at
    posting time, and not when the operator saved the return.
    """
    legset = build_sales_return_legs(sales_return, deduct=True)
    write_legs(legset, date=sales_return.datetime.date(),
               source_type='sales_return', document=sales_return)
    sales_return.posting_status = 'posted'
    sales_return.save(update_fields=['posting_status'])


# kind -> poster, injected into run_journal_sweep so the service module does not
# need to import this views module (which would be a cycle).
_RUN_POSTERS = {
    'invoice': _post_invoice_for_run,
    'purchase': _post_purchase_for_run,
    'expense': _post_expense_for_run,
    'transfer': _post_transfer_for_run,
    'stock': _post_stock_correction_for_run,
    'sales_return': _post_sales_return_for_run,
}


def _parse_date_to(request):
    """Shared validation for both run endpoints.

    Returns (date, None) or (None, error_response). Streaming callers must run
    this BEFORE opening the stream — once the first frame is sent the status
    code is committed to 200 and a 400 is no longer expressible.
    """
    raw = request.data.get('date_to')
    if not raw:
        return None, Response(
            {'error': 'date_to wajib diisi (YYYY-MM-DD).'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        return datetime.date.fromisoformat(str(raw)), None
    except ValueError:
        return None, Response(
            {'error': 'Format date_to tidak valid. Gunakan YYYY-MM-DD.'},
            status=status.HTTP_400_BAD_REQUEST,
        )


def _parse_date_from(request, date_to):
    """Optional window start. Returns (date_or_None, None) or (None, error).

    Absent or blank means the run reaches back to the oldest unposted document,
    which is what it has always done. A value narrows both ends of the window.
    """
    raw = request.data.get('date_from')
    if raw in (None, ''):
        return None, None
    try:
        parsed = datetime.date.fromisoformat(str(raw))
    except ValueError:
        return None, Response(
            {'error': 'Format date_from tidak valid. Gunakan YYYY-MM-DD.'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if parsed > date_to:
        return None, Response(
            {'error': 'date_from tidak boleh setelah date_to.'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    return parsed, None


class JournalRunView(APIView):
    """
    POST /api/accounting/journal/run/
    Body: { date_to: 'YYYY-MM-DD', date_from?: 'YYYY-MM-DD' }

    Finds every Invoice/PurchaseInvoice/AccountTransfer/Expense with
    posting_status='unposted' and a transaction date <= date_to — regardless
    of how old, which is what makes this a "sweep": a document from weeks ago
    that was never posted still gets caught here, not just today's. Pass
    date_from to bound that reach: documents older than it are left unposted
    for a later run. Same-day
    void/edit memo entries (source_type='void_memo'/'edit_memo') are never
    touched by this query — they are written directly by the void/edit
    endpoints and are already "posted" by construction.

    Upserts JournalDayLog (is_posted=True) for every calendar date from the
    earliest date actually posted through date_to, then records everything in
    a JournalBatch, whose summary is returned.

    The request/response contract is unchanged. What changed underneath, twice:

      * Phase 2 moved the transaction boundary — the sweep is no longer one
        atomic block over the whole run, it commits one day at a time, so a
        failure leaves earlier days posted. There is deliberately NO
        @transaction.atomic on this method: adding one back would nest every
        per-day transaction inside a single outer transaction and silently
        restore all-or-nothing behaviour.
      * Phase 4 made this a thin wrapper over preview-then-commit. It is kept
        for scripts and the POS, but it must not carry its own posting
        implementation — a second path writing LedgerEntry rows without a
        JournalEntry header would reopen the flat-ledger problem Phase 4
        closed. The web UI no longer calls this; it uses the two-phase flow.

    Side effect worth knowing about: like any preview, this supersedes an open
    draft. The documents in that draft are the ones this run is about to post,
    so the draft would have gone stale regardless.
    """

    def post(self, request):
        date_to, err = _parse_date_to(request)
        if err:
            return err
        date_from, err = _parse_date_from(request, date_to)
        if err:
            return err

        final = None
        for evt in journal_preview.run_and_commit(
            _actor(request), date_to, _build_journal_summary, date_from,
        ):
            if evt['type'] in ('done', 'error', 'stale'):
                final = evt

        if final is not None and final['type'] == 'stale':
            # Only reachable if a document changed between this run's own
            # preview and its commit — a genuine concurrent edit.
            return Response(
                {'error': final['message'], 'stale': final['stale']},
                status=status.HTTP_409_CONFLICT,
            )

        if final is None or final['type'] == 'error':
            message = (final or {}).get('message', 'Journal run failed.')
            return Response(
                {
                    'error': message,
                    'failed_on': (final or {}).get('date'),
                    'days_committed': (final or {}).get('days_committed', 0),
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # `variances` is a Phase 4 addition that belongs to the two-phase flow —
        # it reports where a committed FIFO amount drifted from what the
        # operator previewed. This endpoint previews and commits back to back,
        # so it can never be meaningfully populated, and callers of the legacy
        # contract must not suddenly receive a key they never had.
        payload = {k: v for k, v in final.items()
                   if k not in ('type', 'variances', 'skipped')}
        return Response(payload, status=status.HTTP_200_OK)


class JournalRunStreamView(APIView):
    """
    POST /api/accounting/journal/run/stream/
    Body: { date_to: 'YYYY-MM-DD', date_from?: 'YYYY-MM-DD' }

    Same sweep as JournalRunView, but streams progress as Server-Sent Events so
    the web UI can animate each day as it is committed:

        event: start   {batch_id, date_to, total, days:[{date, documents}]}
        event: day     {index, date, documents, status: posted|skipped}
        event: done    {...same payload as JournalRunView...}
        event: error   {date, message, days_committed}

    Clients must read this with fetch()+ReadableStream, not EventSource:
    EventSource is GET-only and cannot send the Authorization header.

    Phase 4: like JournalRunView, this is now preview-then-commit under the
    hood. The preview half is drained silently, so the events on the wire are
    the commit's — identical to what this endpoint always emitted.
    """

    def post(self, request):
        # Validate before opening the stream — once the first frame is written
        # the status code is committed to 200.
        date_to, err = _parse_date_to(request)
        if err:
            return err
        date_from, err = _parse_date_from(request, date_to)
        if err:
            return err

        # Resolve the actor up front; the generator must not touch `request`.
        actor = _actor(request)

        def frames():
            try:
                for evt in journal_preview.run_and_commit(
                    actor, date_to, _build_journal_summary, date_from,
                ):
                    yield _sse(evt)
            except Exception as exc:  # noqa: BLE001 — cannot become a 500 mid-stream
                yield _sse({
                    'type': 'error', 'date': None,
                    'message': str(exc) or exc.__class__.__name__,
                    'days_committed': 0,
                })

        return _stream(frames())


class JournalStatusView(APIView):
    """GET /api/accounting/journal/status/?date_from=&date_to=

    Returns {date: is_posted} for every calendar day in the range, for a
    calendar-style UI. A day with no JournalDayLog row at all counts as
    is_posted=False (never swept).
    """

    def get(self, request):
        date_from_raw = request.query_params.get('date_from', '').strip()
        date_to_raw = request.query_params.get('date_to', '').strip()
        if not date_from_raw or not date_to_raw:
            return Response({'error': 'date_from dan date_to wajib diisi (YYYY-MM-DD).'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            d_from = datetime.date.fromisoformat(date_from_raw)
            d_to = datetime.date.fromisoformat(date_to_raw)
        except ValueError:
            return Response({'error': 'Format tanggal tidak valid. Gunakan YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
        if d_from > d_to:
            return Response({'error': 'date_from tidak boleh setelah date_to.'}, status=status.HTTP_400_BAD_REQUEST)

        posted_dates = set(
            JournalDayLog.objects
            .filter(date__gte=d_from, date__lte=d_to, is_posted=True)
            .values_list('date', flat=True)
        )
        result = {}
        cur = d_from
        while cur <= d_to:
            result[cur.isoformat()] = cur in posted_dates
            cur += datetime.timedelta(days=1)
        return Response(result)


# ── Journal Preview (Phase 4 — stage, review, then commit) ────────────────────

def _sse(evt):
    return f"event: {evt['type']}\ndata: {json.dumps(evt, default=str)}\n\n"


def _stream(frames):
    resp = StreamingHttpResponse(frames, content_type='text/event-stream')
    resp['Cache-Control'] = 'no-cache'
    resp['X-Accel-Buffering'] = 'no'   # defeat nginx/proxy response buffering
    return resp


class JournalPreviewView(APIView):
    """
    POST /api/accounting/journal/preview/
         Body: { date_to: 'YYYY-MM-DD', date_from?: 'YYYY-MM-DD' }
    GET  /api/accounting/journal/preview/

    POST builds a staging batch — every journal entry a sweep to ``date_to``
    (from ``date_from``, if given; otherwise from the oldest unposted document)
    would post, materialised into StagedJournalEntry/StagedJournalLine rows and
    nothing else. No ledger row is written, no document's posting_status moves,
    no JournalDayLog is touched. Streams the same per-day SSE events as the run
    endpoint so the existing grid animation works unchanged, except a day
    settles as 'staged' rather than 'posted'.

    Starting a preview supersedes any open draft: accounting is one set of
    books, so there is deliberately never more than one.

    GET returns the current open draft (204 when there is none) so the page can
    drop a returning user straight back into review.
    """

    def get(self, request):
        journal_preview.purge_expired_drafts()
        draft = journal_preview.open_draft()
        if draft is None:
            return Response(status=status.HTTP_204_NO_CONTENT)
        payload = JournalStagingBatchSerializer(draft).data
        payload['days'] = _draft_day_counts(draft)
        return Response(payload)

    def post(self, request):
        date_to, err = _parse_date_to(request)
        if err:
            return err
        date_from, err = _parse_date_from(request, date_to)
        if err:
            return err
        actor = _actor(request)

        def frames():
            try:
                for evt in journal_preview.build_preview(actor, date_to, date_from):
                    yield _sse(evt)
            except Exception as exc:  # noqa: BLE001 — cannot become a 500 mid-stream
                yield _sse({'type': 'error', 'date': None,
                            'message': str(exc) or exc.__class__.__name__})

        return _stream(frames())


#: An entry an operator would want to look at twice. Kept in one place so the
#: per-day `flagged` count and the ?only_warnings= list filter can never drift
#: apart — a day badged "3 perlu diperiksa" must return exactly those 3.
NEEDS_REVIEW = Q(is_balanced=False) | Q(has_estimate=True) | ~Q(warnings=[])


def _draft_day_counts(draft):
    """Per-day rollup driving the review UI's day rail.

    `flagged` is what lets the rail show where the problems are without the
    client fetching every page: the operator can jump straight to the three
    days that need attention instead of walking all forty-five.
    """
    rows = (
        draft.entries
        .values('date')
        .annotate(entries=Count('id'),
                  flagged=Count('id', filter=NEEDS_REVIEW),
                  debit=Sum('total_debit'), credit=Sum('total_credit'))
        .order_by('date')
    )
    return [{
        'date': r['date'].isoformat(),
        'entries': r['entries'],
        'flagged': r['flagged'],
        'total_debit': str(r['debit'] or 0),
        'total_credit': str(r['credit'] or 0),
    } for r in rows]


class JournalPreviewEntriesView(APIView):
    """
    GET /api/accounting/journal/preview/entries/

    Paginated staged entries of the open draft. A sweep of a busy clinic is
    realistically tens of thousands of lines, so the review UI pages over them
    server-side and never loads the batch whole.

    Query params: date, date_from, date_to, source_type, account, q,
                  only_warnings, page, page_size
    """

    def get(self, request):
        draft = journal_preview.open_draft()
        if draft is None:
            return Response({'error': 'Tidak ada draf jurnal yang terbuka.'},
                            status=status.HTTP_404_NOT_FOUND)

        qs = draft.entries.all()

        date        = request.query_params.get('date', '').strip()
        date_from   = request.query_params.get('date_from', '').strip()
        date_to     = request.query_params.get('date_to', '').strip()
        source_type = request.query_params.get('source_type', '').strip()
        account     = request.query_params.get('account', '').strip()
        q           = request.query_params.get('q', '').strip()
        only_warn   = request.query_params.get('only_warnings', '').strip().lower() in ('1', 'true', 'yes')

        if date:
            qs = qs.filter(date=date)
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if source_type:
            qs = qs.filter(source_type=source_type)
        if account:
            qs = qs.filter(lines__account_id=account).distinct()
        if q:
            qs = qs.filter(Q(memo__icontains=q) | Q(source_label__icontains=q))
        if only_warn:
            # An entry that does not balance, one whose account does not exist
            # yet, or one carrying a FIFO estimate that may move at commit.
            qs = qs.filter(NEEDS_REVIEW).distinct()

        try:
            page      = max(1, int(request.query_params.get('page', 1)))
            page_size = min(500, max(1, int(request.query_params.get('page_size', 100))))
        except ValueError:
            page, page_size = 1, 100

        total = qs.count()
        offset = (page - 1) * page_size
        entries = qs[offset:offset + page_size]

        return Response({
            'count': total,
            'page': page,
            'page_size': page_size,
            'results': StagedJournalEntryListSerializer(entries, many=True).data,
        })


class JournalPreviewEntryDetailView(APIView):
    """GET /api/accounting/journal/preview/entries/<pk>/ — one staged entry and
    its lines, for the entry detail page in 'draft' mode."""

    def get(self, request, pk):
        entry = (
            StagedJournalEntry.objects
            .filter(pk=pk)
            .select_related('batch')
            .prefetch_related('lines__account')
            .first()
        )
        if entry is None:
            return Response({'error': 'Entri tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)
        payload = StagedJournalEntryDetailSerializer(entry).data
        payload['batch'] = JournalStagingBatchSerializer(entry.batch).data
        return Response(payload)


class JournalPreviewCommitView(APIView):
    """
    POST /api/accounting/journal/preview/commit/
    Body: { staging_batch_id }  (optional — defaults to the open draft)

    Writes the reviewed draft to the ledger, streaming per-day SSE events.

    Emits an extra terminal event the run endpoint never had:

        event: stale  {stale: [{staged_entry_id, source_label, reason}], message}

    …when a source document changed after the operator reviewed it. Nothing is
    posted in that case; the draft stays open and the UI asks for a re-preview.

    There is deliberately NO @transaction.atomic here. The commit owns a
    per-day transaction boundary — a failure on day k leaves days 1…k-1 posted,
    which is what makes a long sweep resumable. An outer atomic block would
    silently restore all-or-nothing behaviour.
    """

    def post(self, request):
        batch_id = request.data.get('staging_batch_id')
        if batch_id:
            batch = JournalStagingBatch.objects.filter(pk=batch_id).first()
        else:
            batch = journal_preview.open_draft()

        if batch is None:
            return Response({'error': 'Draf jurnal tidak ditemukan. Jalankan pratinjau lebih dulu.'},
                            status=status.HTTP_404_NOT_FOUND)
        if batch.status != 'draft':
            return Response({'error': f'Draf ini berstatus "{batch.status}" dan tidak bisa dicatat.'},
                            status=status.HTTP_409_CONFLICT)
        if batch.expires_at <= timezone.now():
            return Response({'error': 'Draf sudah kedaluwarsa. Jalankan pratinjau ulang.'},
                            status=status.HTTP_409_CONFLICT)

        actor = _actor(request)

        def frames():
            try:
                for evt in journal_preview.commit_preview(actor, batch, _build_journal_summary):
                    yield _sse(evt)
            except Exception as exc:  # noqa: BLE001
                yield _sse({'type': 'error', 'date': None,
                            'message': str(exc) or exc.__class__.__name__,
                            'days_committed': 0})

        return _stream(frames())


class JournalPreviewDiscardView(APIView):
    """POST /api/accounting/journal/preview/discard/ — drop the open draft."""

    def post(self, request):
        draft = journal_preview.open_draft()
        if draft is None:
            return Response(status=status.HTTP_204_NO_CONTENT)
        draft.status = 'discarded'
        draft.save(update_fields=['status'])
        return Response({'discarded': draft.id})


# ── Journal Entries (Phase 4 — headers, detail, corrections) ──────────────────

class JournalEntryListView(APIView):
    """
    GET /api/accounting/journal/entries/

    Posted journal entries at header level — the list behind the entry detail
    pages. ``JournalHistoryView`` still serves the flat per-line view; this is
    the document-level companion to it.

    Query params: date_from, date_to, source_type, account, batch, q,
                  has_correction, page, page_size
    """

    def get(self, request):
        qs = (
            JournalEntry.objects
            .select_related('invoice', 'purchase_invoice', 'expense', 'transfer')
            .annotate(line_count=Count('lines', distinct=True),
                      reversal_count=Count('reversed_by', distinct=True))
        )
        qs = filter_by_branch(qs, request)

        date_from   = request.query_params.get('date_from', '').strip()
        date_to     = request.query_params.get('date_to', '').strip()
        source_type = request.query_params.get('source_type', '').strip()
        account     = request.query_params.get('account', '').strip()
        batch       = request.query_params.get('batch', '').strip()
        q           = request.query_params.get('q', '').strip()
        has_corr    = request.query_params.get('has_correction', '').strip().lower() in ('1', 'true', 'yes')

        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if source_type:
            qs = qs.filter(source_type=source_type)
        if account:
            qs = qs.filter(lines__account_id=account).distinct()
        if batch:
            qs = qs.filter(batch_id=batch)
        if q:
            qs = qs.filter(Q(memo__icontains=q) | Q(entry_number__icontains=q))
        if has_corr:
            qs = qs.filter(reversal_count__gt=0)

        try:
            page      = max(1, int(request.query_params.get('page', 1)))
            page_size = min(500, max(1, int(request.query_params.get('page_size', 100))))
        except ValueError:
            page, page_size = 1, 100

        total = qs.count()
        offset = (page - 1) * page_size
        return Response({
            'count': total,
            'page': page,
            'page_size': page_size,
            'results': JournalEntryListSerializer(qs[offset:offset + page_size], many=True).data,
        })


def _entry_or_none(pk):
    return (
        JournalEntry.objects
        .filter(pk=pk)
        .select_related('invoice', 'purchase_invoice', 'expense', 'transfer',
                        'created_by', 'reverses', 'corrects')
        .prefetch_related('lines__account', 'reversed_by', 'corrections')
        .first()
    )


class JournalEntryDetailView(APIView):
    """GET /api/accounting/journal/entries/<pk>/ — header, every line with its
    account, the source document reference, and the correction chain."""

    def get(self, request, pk):
        entry = _entry_or_none(pk)
        if entry is None:
            return Response({'error': 'Jurnal tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(JournalEntryDetailSerializer(entry).data)


class JournalEntryCorrectionDraftView(APIView):
    """
    GET /api/accounting/journal/entries/<pk>/correction-draft/

    Everything the correction editor needs to open prefilled: the original
    entry, the reversal that will be generated for it, and an editable copy of
    its lines as the starting point for the replacement.
    """

    def get(self, request, pk):
        entry = _entry_or_none(pk)
        if entry is None:
            return Response({'error': 'Jurnal tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)

        blocked = _correction_block_reason(entry)
        lines = [{
            'account_id': l.account_id,
            'account_number': l.account.account_number,
            'account_name': l.account.name,
            'entry_type': l.entry_type,
            'amount': str(l.amount),
            'description': l.description,
        } for l in entry.lines.select_related('account').all()]

        return Response({
            'original': JournalEntryDetailSerializer(entry).data,
            'correction_date': timezone.now().date().isoformat(),
            'blocked_reason': blocked,
            'lines': lines,
        })


def _correction_block_reason(entry):
    """Why this entry cannot be corrected, or None when it can."""
    if entry.source_type == 'reversal':
        return 'Entri pembalik tidak dapat dikoreksi. Koreksi jurnal aslinya.'
    if entry.reversed_by.exists():
        existing = entry.reversed_by.first()
        return f'Jurnal ini sudah dibalik oleh {existing.entry_number}.'
    return None


class JournalEntryCorrectView(APIView):
    """
    POST /api/accounting/journal/entries/<pk>/correct/
    Body: { memo, reason, lines: [{account, entry_type, amount, description}] }

    Posts a correction as two new entries, both dated **today**:

      1. a full reversal of the original — auto-generated, every line flipped,
         ``source_type='reversal'``, ``reverses=<original>``;
      2. the operator's replacement — ``source_type='correction'``,
         ``corrects=<original>``.

    The original day's books are never rewritten, which is the whole point: a
    period that has been reported on stays as it was reported.

    Deliberately does NOT touch JournalDayLog. Reversal and correction rows are
    already-posted by construction (same convention as the void/edit memo
    path); marking today is_posted=True would tell the financial-reports guard
    that today had been swept when it has not.

    The source document is left alone. If the *document* is wrong the operator
    should edit it — corrections are for entries whose accounting was wrong.
    """

    def post(self, request, pk):
        entry = _entry_or_none(pk)
        if entry is None:
            return Response({'error': 'Jurnal tidak ditemukan.'}, status=status.HTTP_404_NOT_FOUND)

        blocked = _correction_block_reason(entry)
        if blocked:
            return Response({'error': blocked}, status=status.HTTP_409_CONFLICT)

        memo = (request.data.get('memo') or '').strip()
        if not memo:
            return Response({'error': 'Keterangan koreksi wajib diisi.'},
                            status=status.HTTP_400_BAD_REQUEST)
        reason = (request.data.get('reason') or '').strip()

        legset, errors = _parse_correction_lines(request.data.get('lines') or [])
        if errors:
            return Response({'error': 'Baris jurnal tidak valid.', 'lines': errors},
                            status=status.HTTP_400_BAD_REQUEST)

        today = timezone.now().date()
        actor = _actor(request)

        with transaction.atomic():
            # Re-read under lock so two managers cannot both reverse the same
            # entry and double the reversal.
            locked = JournalEntry.objects.select_for_update().get(pk=entry.pk)
            if locked.reversed_by.exists():
                return Response({'error': 'Jurnal ini baru saja dibalik oleh proses lain.'},
                                status=status.HTTP_409_CONFLICT)

            numbers = reserve_entry_numbers(today.year, 2)

            reversal = write_legs(
                reverse_legset(legset_from_entry(locked)),
                date=today,
                source_type='reversal',
                document=_entry_document(locked),
                actor=actor,
                memo=f'Pembalikan {locked.entry_number}'[:255],
                entry_number=numbers[0],
                reverses=locked,
            )

            correction = write_legs(
                legset,
                date=today,
                source_type='correction',
                document=_entry_document(locked),
                actor=actor,
                memo=memo,
                entry_number=numbers[1],
                corrects=locked,
            )

            AuditLog.objects.create(
                performed_by=actor,
                action='CORRECT',
                entity_type='JournalEntry',
                entity_id=str(locked.pk),
                description=(
                    f'Koreksi {locked.entry_number}: pembalikan {reversal.entry_number}, '
                    f'koreksi {correction.entry_number}. Alasan: {reason or "—"}'
                ),
            )

        return Response({
            'original': JournalEntryRefSerializer(locked).data,
            'reversal': JournalEntryRefSerializer(reversal).data,
            'correction': JournalEntryDetailSerializer(_entry_or_none(correction.pk)).data,
        }, status=status.HTTP_201_CREATED)


def _entry_document(entry):
    """The source document instance an entry hangs off, or None.

    A correction inherits the original's document link so the invoice/expense
    detail page shows the whole story — original, reversal and correction —
    rather than the original alone.
    """
    return entry.invoice or entry.purchase_invoice or entry.expense or entry.transfer


def _parse_correction_lines(raw):
    """Validate the operator's lines and turn them into a LegSet.

    Returns ``(legset, errors)``; ``errors`` maps line index → message and is
    empty when the entry is postable. Rejects anything the ledger should never
    see: fewer than two lines, non-positive amounts, head accounts (which are
    grouping rows, not postable), and — the important one — an entry whose
    debits and credits do not match.
    """
    errors = {}
    legset = LegSet()

    if len(raw) < 2:
        return legset, {'_': 'Jurnal koreksi butuh minimal dua baris.'}

    account_ids = []
    for i, row in enumerate(raw):
        try:
            account_ids.append(int(row.get('account')))
        except (TypeError, ValueError):
            errors[i] = 'Akun wajib dipilih.'
            account_ids.append(None)

    accounts = {a.pk: a for a in ChartOfAccounts.objects.filter(
        pk__in=[a for a in account_ids if a is not None])}

    for i, row in enumerate(raw):
        if i in errors:
            continue
        account = accounts.get(account_ids[i])
        if account is None:
            errors[i] = 'Akun tidak ditemukan.'
            continue
        if account.is_head:
            errors[i] = 'Akun induk tidak bisa dijurnal. Pilih sub-akun.'
            continue

        entry_type = (row.get('entry_type') or '').strip().lower()
        if entry_type not in ('debit', 'credit'):
            errors[i] = 'Tipe harus debit atau kredit.'
            continue

        amount = _safe_decimal(row.get('amount', 0))
        if amount <= 0:
            errors[i] = 'Jumlah harus lebih dari nol.'
            continue

        description = (row.get('description') or '').strip()
        if not description:
            errors[i] = 'Keterangan baris wajib diisi.'
            continue

        legset.add(account, entry_type, amount, description)

    if errors:
        return legset, errors

    if not legset.is_balanced:
        return legset, {'_': (
            f'Jurnal tidak seimbang: debit {legset.total_debit} '
            f'vs kredit {legset.total_credit}.'
        )}

    return legset, {}


# ── Journal History ────────────────────────────────────────────────────────────

class JournalHistoryView(APIView):
    """
    GET /api/accounting/journal/
    Query params:
      date_from       YYYY-MM-DD
      date_to         YYYY-MM-DD
      account         account id
      account_number  account_number (int)
      account_type    asset|liability|equity|revenue|cogs|expense|other_income|other_expense
      source_type     invoice|purchase|transfer|adjustment|stock|opname|manual
      entry_type      debit|credit
      q               search description
      page            page number (default 1)
      page_size       items per page (default 100, max 500)
    """

    def get(self, request):
        qs = filter_by_branch(
            LedgerEntry.objects.select_related('account', 'invoice', 'purchase_invoice', 'transfer'),
            request,
        )

        date_from    = request.query_params.get('date_from', '').strip()
        date_to      = request.query_params.get('date_to', '').strip()
        acct_id      = request.query_params.get('account', '').strip()
        acct_number  = request.query_params.get('account_number', '').strip()
        acct_type    = request.query_params.get('account_type', '').strip()
        source_type  = request.query_params.get('source_type', '').strip()
        entry_type   = request.query_params.get('entry_type', '').strip().lower()
        q            = request.query_params.get('q', '').strip()

        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if acct_id:
            qs = qs.filter(account_id=acct_id)
        if acct_number:
            qs = qs.filter(account__account_number=acct_number)
        if acct_type:
            qs = qs.filter(account__account_type=acct_type)
        if source_type:
            qs = qs.filter(source_type=source_type)
        if entry_type in ('debit', 'credit'):
            qs = qs.filter(entry_type=entry_type)
        if q:
            qs = qs.filter(description__icontains=q)

        totals = qs.aggregate(
            total_debit=Sum('amount', filter=Q(entry_type='debit')),
            total_credit=Sum('amount', filter=Q(entry_type='credit')),
        )

        try:
            page      = max(1, int(request.query_params.get('page', 1)))
            page_size = min(500, max(1, int(request.query_params.get('page_size', 100))))
        except ValueError:
            page, page_size = 1, 100

        total_count = qs.count()
        offset = (page - 1) * page_size
        entries = qs[offset:offset + page_size]

        return Response({
            'count': total_count,
            'page': page,
            'page_size': page_size,
            'total_debit': str(totals['total_debit'] or 0),
            'total_credit': str(totals['total_credit'] or 0),
            'results': LedgerEntrySerializer(entries, many=True).data,
        })


# ── Accounting Landing ─────────────────────────────────────────────────────────

class AccountingDashboardView(APIView):
    """
    GET /api/accounting/dashboard/
    Returns unpaid/partial purchase invoices and a summary.
    """

    def get(self, request):
        today = timezone.now().date()

        unpaid_qs = PurchaseInvoice.objects.filter(
            status__in=['unpaid', 'partial']
        ).select_related('supplier', 'payment_account', 'payment_method', 'payment_method__linked_account').order_by('due_date', '-purchase_date')

        overdue_count = unpaid_qs.filter(due_date__lt=today).count()
        total_unpaid  = unpaid_qs.aggregate(
            s=Sum('total_amount') - Sum('amount_paid')
        )['s'] or Decimal('0')

        return Response({
            'unpaid_invoices': PurchaseInvoiceListSerializer(unpaid_qs, many=True).data,
            'overdue_count':   overdue_count,
            'total_unpaid':    str(total_unpaid),
        })


# ── Payment Plan (Rencana Pembayaran) ──────────────────────────────────────────

INDO_MONTHS = [
    'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
    'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember',
]


def _payment_plan_queryset(date_from, date_to, include_no_due_date=False):
    """Outstanding (unpaid/partial) purchase invoices whose due_date (jatuh tempo)
    falls in the inclusive range, ordered by due date (nulls last).

    When include_no_due_date is True, invoices without a due date are always
    included regardless of the range so that nothing outstanding is missed.
    """
    date_filter = Q(due_date__gte=date_from, due_date__lte=date_to)
    if include_no_due_date:
        date_filter |= Q(due_date__isnull=True)
    return (
        PurchaseInvoice.objects
        .select_related('supplier')
        .filter(status__in=['unpaid', 'partial'])
        .filter(date_filter)
        .order_by(F('due_date').asc(nulls_last=True), 'purchase_date', 'id')
    )


def _parse_plan_include_no_due_date(request):
    """Whether outstanding invoices without a due date should be included."""
    return request.query_params.get('include_no_due_date', '').strip().lower() in ('1', 'true', 'yes')


def _parse_plan_range(request):
    """Returns (date_from, date_to, error_response)."""
    import datetime
    date_from = request.query_params.get('date_from', '').strip()
    date_to   = request.query_params.get('date_to', '').strip()
    if not date_from or not date_to:
        return None, None, Response(
            {'error': 'date_from dan date_to wajib diisi (YYYY-MM-DD).'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        d_from = datetime.date.fromisoformat(date_from)
        d_to   = datetime.date.fromisoformat(date_to)
    except ValueError:
        return None, None, Response(
            {'error': 'Format tanggal tidak valid. Gunakan YYYY-MM-DD.'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    if d_from > d_to:
        return None, None, Response(
            {'error': 'date_from tidak boleh setelah date_to.'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    return d_from, d_to, None


def _default_plan_title(d_from, d_to):
    """e.g. 'Rencana Pembayaran ( 1 - 15 Juni ) 2026'."""
    if d_from.month == d_to.month and d_from.year == d_to.year:
        period = f'( {d_from.day} - {d_to.day} {INDO_MONTHS[d_to.month - 1]} ) {d_to.year}'
    elif d_from.year == d_to.year:
        period = (f'( {d_from.day} {INDO_MONTHS[d_from.month - 1]} - '
                  f'{d_to.day} {INDO_MONTHS[d_to.month - 1]} ) {d_to.year}')
    else:
        period = (f'( {d_from.day} {INDO_MONTHS[d_from.month - 1]} {d_from.year} - '
                  f'{d_to.day} {INDO_MONTHS[d_to.month - 1]} {d_to.year} )')
    return f'Rencana Pembayaran {period}'


class PaymentPlanPreviewView(APIView):
    """
    GET /api/accounting/payment-plan/preview/?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD
                                             &include_no_due_date=true
    Lists outstanding purchase invoices (unpaid/partial) whose due date falls in
    the range for review before generating the Excel file.
    """

    def get(self, request):
        d_from, d_to, err = _parse_plan_range(request)
        if err:
            return err
        include_no_dd = _parse_plan_include_no_due_date(request)

        rows = []
        total = Decimal('0')
        for inv in _payment_plan_queryset(d_from, d_to, include_no_dd):
            balance = inv.total_amount - inv.amount_paid
            total += balance
            rows.append({
                'id':                 inv.id,
                'internal_id':        inv.internal_id,
                'external_invoice_no': inv.external_invoice_no,
                'supplier_name':      inv.supplier.name,
                'balance_due':        str(balance),
                'due_date':           inv.due_date.isoformat() if inv.due_date else None,
                'purchase_date':      inv.purchase_date.isoformat(),
                'status':             inv.status,
            })

        return Response({
            'date_from':     d_from.isoformat(),
            'date_to':       d_to.isoformat(),
            'default_title': _default_plan_title(d_from, d_to),
            'count':         len(rows),
            'total':         str(total),
            'results':       rows,
        })


class PaymentPlanExportView(APIView):
    """
    GET /api/accounting/payment-plan/export/?date_from=&date_to=&title=
                                            &include_no_due_date=true
    Streams an .xlsx file styled to match the clinic's 'Rencana Pembayaran' template.
    Filters outstanding invoices by due date (jatuh tempo).
    """

    def get(self, request):
        from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side

        d_from, d_to, err = _parse_plan_range(request)
        if err:
            return err
        include_no_dd = _parse_plan_include_no_due_date(request)

        title = (request.query_params.get('title', '') or '').strip()
        if not title:
            title = _default_plan_title(d_from, d_to)

        # Period label reused for the credit-card block title.
        if title.lower().startswith('rencana pembayaran'):
            period_label = title[len('rencana pembayaran'):].strip()
        else:
            period_label = _default_plan_title(d_from, d_to)[len('Rencana Pembayaran'):].strip()

        # ── Style primitives (mirrors the reference workbook) ──────────────────
        ACCT_FMT  = '_-* #,##0_-;\\-* #,##0_-;_-* "-"_-;_-@_-'
        DATE_FMT  = '[$-409]d\\-mmm;@'
        thin      = Side(style='thin')
        all_thin  = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill(
            patternType='solid',
            fgColor=Color(theme=8, tint=0.7999816888943144),
        )
        f_normal = Font(name='Calibri', size=11)
        f_bold   = Font(name='Calibri', size=11, bold=True)
        f_title  = Font(name='Calibri', size=14, bold=True)
        center   = Alignment(horizontal='center', vertical='center')
        center_v = Alignment(vertical='center')
        left_v   = Alignment(horizontal='left', vertical='center')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Rencana Pembayaran'

        # Column widths copied from the reference file
        widths = {'A': 33.28515625, 'B': 31.42578125, 'C': 13.28515625,
                  'D': 10.42578125, 'E': 14.5703125, 'F': 18.28515625, 'G': 39.28515625}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        # ── Title (merged A1:G1) ───────────────────────────────────────────────
        ws.merge_cells('A1:G1')
        t = ws['A1']
        t.value = title
        t.font = f_title
        t.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 18.75

        # ── Header row (row 3) ─────────────────────────────────────────────────
        headers = ['No. Invoice', 'Supplier', 'Tagihan', 'DD', 'Tgl Byr', 'No. Rek', 'a/n']
        for idx, label in enumerate(headers):
            c = ws.cell(row=3, column=idx + 1, value=label)
            c.font = f_bold
            c.alignment = center
            c.border = all_thin
            c.fill = header_fill

        # ── Data rows ──────────────────────────────────────────────────────────
        row = 4
        first_data_row = row
        for inv in _payment_plan_queryset(d_from, d_to, include_no_dd):
            balance = inv.total_amount - inv.amount_paid

            a = ws.cell(row=row, column=1, value=inv.external_invoice_no or '')
            a.font = f_normal; a.alignment = left_v; a.border = all_thin

            b = ws.cell(row=row, column=2, value=inv.supplier.name)
            b.font = f_normal; b.alignment = center_v; b.border = all_thin

            c = ws.cell(row=row, column=3, value=float(balance))
            c.font = f_normal; c.alignment = center_v
            c.number_format = ACCT_FMT; c.border = all_thin

            d = ws.cell(row=row, column=4)
            if inv.due_date:
                d.value = inv.due_date
                d.number_format = DATE_FMT
            d.font = f_normal; d.alignment = center; d.border = all_thin

            # E (Tgl Byr), F (No. Rek), G (a/n) — left blank for manual entry
            e = ws.cell(row=row, column=5)
            e.font = f_normal; e.alignment = center; e.border = all_thin
            fcell = ws.cell(row=row, column=6)
            fcell.font = f_normal; fcell.alignment = center; fcell.border = all_thin
            g = ws.cell(row=row, column=7)
            g.font = f_normal; g.alignment = left_v; g.border = all_thin

            row += 1

        last_data_row = row - 1
        if last_data_row < first_data_row:
            last_data_row = first_data_row  # keep SUM range valid even when empty

        # ── Total row (one blank row below the table) ──────────────────────────
        total_row = row + 1
        total_border = Border(top=thin, bottom=Side(style='double'))
        tb = ws.cell(row=total_row, column=2, value='Total')
        tb.font = f_bold; tb.border = total_border
        tc = ws.cell(row=total_row, column=3, value=f'=SUM(C{first_data_row}:C{last_data_row})')
        tc.font = f_bold; tc.number_format = ACCT_FMT; tc.border = total_border

        # ── Static credit-card block ───────────────────────────────────────────
        cc_title_row = total_row + 4
        ws.merge_cells(start_row=cc_title_row, start_column=2, end_row=cc_title_row, end_column=4)
        cc = ws.cell(row=cc_title_row, column=2, value=f'Pembayaran Kartu kredit {period_label}')
        cc.font = f_bold; cc.alignment = Alignment(horizontal='center')
        for col in (2, 3, 4):
            ws.cell(row=cc_title_row, column=col).border = all_thin

        # Spacer row (bordered, empty)
        spacer_row = cc_title_row + 1
        for col in (2, 3, 4):
            ws.cell(row=spacer_row, column=col).border = all_thin

        # Column headers
        hdr_row = spacer_row + 1
        for col, label in ((2, 'BANK'), (3, 'TGL BAYAR'), (4, 'TGL LUNAS')):
            hc = ws.cell(row=hdr_row, column=col, value=label)
            hc.font = f_normal; hc.border = all_thin

        # Static bank rows (TGL LUNAS left blank for manual entry)
        bank_rows = [('CITIBANK', '10-15'), ('MANDIRI', '17-20'),
                     ('BCA', '15-17'), ('BNI', '7-9')]
        r = hdr_row + 1
        for bank, tgl in bank_rows:
            ws.cell(row=r, column=2, value=bank).font = f_normal
            ws.cell(row=r, column=3, value=tgl).font = f_normal
            for col in (2, 3, 4):
                ws.cell(row=r, column=col).border = all_thin
            r += 1

        # ── Page setup (landscape A4, matching the reference) ──────────────────
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 9
        ws.page_margins.left = 0.59
        ws.page_margins.right = 0
        ws.page_margins.top = 0
        ws.page_margins.bottom = 0

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_name = title.replace('/', '-').replace('\\', '-')
        return HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{safe_name}.xlsx"'},
        )


# ── Daily Sales ────────────────────────────────────────────────────────────────

class DailySalesView(APIView):
    """
    GET /api/accounting/daily-sales/?date=YYYY-MM-DD
    Returns total sales and breakdown per cash (payment) account for the given date.
    Defaults to today in Jakarta time (WIB, UTC+7).
    """

    def get(self, request):
        import datetime
        from zoneinfo import ZoneInfo

        WIB = ZoneInfo('Asia/Jakarta')
        date_str = request.query_params.get('date', '').strip()
        if date_str:
            try:
                target_date = datetime.date.fromisoformat(date_str)
            except ValueError:
                return Response(
                    {'error': 'Invalid date. Use YYYY-MM-DD.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            target_date = timezone.now().astimezone(WIB).date()

        # Filter invoices for the target date (database stores UTC; compare naive date in local time)
        day_start = datetime.datetime(target_date.year, target_date.month, target_date.day,
                                      tzinfo=WIB)
        day_end   = day_start + datetime.timedelta(days=1)

        # Voided invoices are out: this figure is counted against the physical
        # drawer, and a voided sale put nothing in it.
        invoices = filter_by_branch(
            Invoice.objects
            .filter(datetime__gte=day_start, datetime__lt=day_end, is_voided=False)
            .select_related('payment_method', 'payment_method__linked_account'),
            request,
        )

        grand_total = Decimal('0')
        total_count = 0
        for inv in invoices:
            grand_total += inv.grand_total
            total_count += 1

        # Split payments are broken out per method rather than heaped onto the
        # first one — this report is what the drawer gets counted against.
        breakdown = [
            {
                'account_id':     r['payment_method_id'],
                'account_number': r['account_number'],
                'account_name':   r['method'],
                'total':          str(r['total']),
                'invoice_count':  r['invoice_count'],
            }
            for r in payment_method_breakdown(invoices)
        ]

        return Response({
            'date':          str(target_date),
            'total':         str(grand_total),
            'invoice_count': total_count,
            'by_account':    breakdown,
        })
