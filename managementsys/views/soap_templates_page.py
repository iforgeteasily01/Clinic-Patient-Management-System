import io

import openpyxl
from openpyxl.styles import Font, PatternFill
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView

from ..api.serializers import SoapTemplateSerializer
from ..models import AppUser, AuditLog, SoapTemplate


def _actor(request):
    return request.user if isinstance(request.user, AppUser) else None


class SoapTemplateListCreateView(APIView):
    def get(self, request):
        field = request.GET.get('field')
        qs = SoapTemplate.objects.all()
        if field:
            qs = qs.filter(field=field)
        return Response(SoapTemplateSerializer(qs, many=True).data)

    def post(self, request):
        serializer = SoapTemplateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='CREATE',
            entity_type='SoapTemplate',
            entity_id=str(instance.id),
            description=f'SOAP template created: [{instance.field}] {instance.title}',
        )
        return Response(SoapTemplateSerializer(instance).data, status=status.HTTP_201_CREATED)


class SoapTemplateDetailView(APIView):
    def _get(self, pk):
        try:
            return SoapTemplate.objects.get(pk=pk)
        except SoapTemplate.DoesNotExist:
            return None

    def put(self, request, pk):
        instance = self._get(pk)
        if instance is None:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = SoapTemplateSerializer(instance, data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='SoapTemplate',
            entity_id=str(pk),
            description=f'SOAP template updated: [{instance.field}] {instance.title}',
        )
        return Response(serializer.data)

    def delete(self, request, pk):
        instance = self._get(pk)
        if instance is None:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='DELETE',
            entity_type='SoapTemplate',
            entity_id=str(pk),
            description=f'SOAP template deleted: [{instance.field}] {instance.title}',
        )
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


VALID_FIELDS = {'subjective', 'objective', 'assessment', 'plan'}

_HEADER_FILL = PatternFill('solid', fgColor='4F81BD')
_HEADER_FONT = Font(bold=True, color='FFFFFF')


class SoapTemplateExportView(APIView):
    """GET — download all SOAP templates as an .xlsx file."""

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'SOAP Templates'

        headers = ['field', 'title', 'body', 'sort_order']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL

        for tmpl in SoapTemplate.objects.all():
            ws.append([tmpl.field, tmpl.title, tmpl.body, tmpl.sort_order])

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from django.http import HttpResponse
        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="soap_templates.xlsx"'
        return response


class SoapTemplateTemplateDownloadView(APIView):
    """GET — download a blank sample .xlsx showing the expected import format."""

    def get(self, request):
        from django.http import HttpResponse

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'SOAP Templates'

        headers = ['field', 'title', 'body', 'sort_order']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL

        samples = [
            ('subjective', 'Chief complaint', 'Patient reports skin concern with…', 1),
            ('objective',  'Skin assessment', 'Skin appears well-hydrated with…', 1),
            ('assessment', 'Acne vulgaris',   'Consistent with mild acne vulgaris.', 1),
            ('plan',       'Follow-up',       'Return in 4 weeks for review.', 1),
        ]
        for row in samples:
            ws.append(list(row))

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="soap_templates_sample.xlsx"'
        return response


class SoapTemplateImportView(APIView):
    """
    POST multipart/form-data with file=<xlsx>
    Columns (row 1 = header, skipped):
      field | title | body | sort_order
    field must be one of: subjective, objective, assessment, plan
    sort_order defaults to 0 if blank.
    Existing templates matched by (field + title) are updated; new ones created.
    Returns { created, updated, errors: [{row, message}] }
    """

    def post(self, request):
        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), read_only=True, data_only=True)
        except Exception:
            return Response({'error': 'Could not parse file. Upload a valid .xlsx file.'}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'error': 'File is empty.'}, status=status.HTTP_400_BAD_REQUEST)

        created = updated = 0
        errors = []

        for i, row in enumerate(rows[1:], start=2):
            padded = list(row) + [None] * 4
            field      = str(padded[0]).strip().lower() if padded[0] is not None else ''
            title      = str(padded[1]).strip()         if padded[1] is not None else ''
            body       = str(padded[2]).strip()         if padded[2] is not None else ''
            order_raw  = padded[3]

            if not field:
                errors.append({'row': i, 'message': 'field is required.'})
                continue
            if field not in VALID_FIELDS:
                errors.append({'row': i, 'message': f'Invalid field "{field}". Must be one of: {", ".join(sorted(VALID_FIELDS))}.'})
                continue
            if not title:
                errors.append({'row': i, 'message': 'title is required.'})
                continue
            if not body:
                errors.append({'row': i, 'message': 'body is required.'})
                continue

            try:
                sort_order = int(order_raw) if order_raw not in (None, '') else 0
            except (ValueError, TypeError):
                sort_order = 0

            obj, was_created = SoapTemplate.objects.update_or_create(
                field=field,
                title=title,
                defaults={'body': body, 'sort_order': sort_order},
            )
            AuditLog.objects.create(
                performed_by=_actor(request),
                action='CREATE' if was_created else 'UPDATE',
                entity_type='SoapTemplate',
                entity_id=str(obj.id),
                description=f'SOAP template {"imported" if was_created else "updated via import"}: [{obj.field}] {obj.title}',
            )
            if was_created:
                created += 1
            else:
                updated += 1

        return Response({'created': created, 'updated': updated, 'errors': errors})
