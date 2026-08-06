import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import generics, status
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView
from ..auth_backend import IsAppAuthenticated
from .financial_reports_utils import ledger_rows_with_balance, unposted_dates_in_range

_HEADER_FONT = Font(bold=True, color='FFFFFF')
_HEADER_FILL = PatternFill('solid', fgColor='0284C7')
from ..api.serializers import (
    AppUserAdminSerializer,
    BeauticianAdminStatusSerializer,
    BeauticiansSerializer,
    ChartOfAccountsSerializer,
    ColorPaletteSerializer,
    DoctorsSerializer,
    LedgerEntrySerializer,
    PatientSerializer,
    PaymentMethodSerializer,
    SiteConfigSerializer,
    TreatmentCategorySerializer,
    TreatmentMaterialSerializer,
    TreatmentPackageSerializer,
    TreatmentSerializer,
)
from ..models import (
    AppUser, AuditLog, Beauticians, ChartOfAccounts, ColorPalette, Doctors, InventoryItem,
    LedgerEntry, Patient, PaymentMethod, SiteConfig, Treatment, TreatmentCategory, TreatmentMaterial, TreatmentPackage,
)


def _actor(request):
    return request.user if isinstance(request.user, AppUser) else None


class DoctorListCreateAdminView(generics.ListCreateAPIView):
    queryset = Doctors.objects.all().order_by('doctor_name')
    serializer_class = DoctorsSerializer

    def perform_create(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='CREATE',
            entity_type='Doctor',
            entity_id=str(instance.id),
            description=f'Doctor added: {instance.doctor_name}',
        )


class DoctorDetailAdminView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Doctors.objects.all()
    serializer_class = DoctorsSerializer

    def perform_update(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='UPDATE',
            entity_type='Doctor',
            entity_id=str(instance.id),
            description=f'Doctor updated: {instance.doctor_name}',
        )

    def perform_destroy(self, instance):
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='DELETE',
            entity_type='Doctor',
            entity_id=str(instance.id),
            description=f'Doctor deleted: {instance.doctor_name}',
        )
        instance.delete()


class PatientListCreateAdminView(generics.ListCreateAPIView):
    serializer_class = PatientSerializer

    def get_queryset(self):
        search = self.request.GET.get('search', '').strip()
        qs = Patient.objects.all().order_by('name')
        if search:
            qs = qs.filter(
                Q(name__icontains=search) | Q(patient_no__icontains=search)
            )
        return qs[:50]

    def perform_create(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='CREATE',
            entity_type='Patient',
            entity_id=instance.patient_no,
            description=f'Patient added via admin: {instance.name} ({instance.patient_no})',
        )


class PatientDetailAdminView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Patient.objects.all()
    serializer_class = PatientSerializer
    lookup_field = 'patient_no'
    lookup_url_kwarg = 'patient_no'

    def perform_update(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='UPDATE',
            entity_type='Patient',
            entity_id=instance.patient_no,
            description=f'Patient updated: {instance.name} ({instance.patient_no})',
        )

    def perform_destroy(self, instance):
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='DELETE',
            entity_type='Patient',
            entity_id=instance.patient_no,
            description=f'Patient deleted: {instance.name} ({instance.patient_no})',
        )
        instance.delete()


class TreatmentListCreateAdminView(generics.ListCreateAPIView):
    queryset = Treatment.objects.all().order_by('category', 'name')
    serializer_class = TreatmentSerializer

    def perform_create(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='CREATE',
            entity_type='Treatment',
            entity_id=str(instance.id),
            description=f'Treatment added: {instance.name} ({instance.code})',
        )


class TreatmentDetailAdminView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Treatment.objects.all()
    serializer_class = TreatmentSerializer

    def perform_update(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='UPDATE',
            entity_type='Treatment',
            entity_id=str(instance.id),
            description=f'Treatment updated: {instance.name} ({instance.code})',
        )

    def perform_destroy(self, instance):
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='DELETE',
            entity_type='Treatment',
            entity_id=str(instance.id),
            description=f'Treatment deleted: {instance.name} ({instance.code})',
        )
        instance.delete()


class TreatmentMaterialListCreateView(APIView):
    """GET/POST treatment materials for a given treatment (admin only)."""

    def get(self, request, pk):
        try:
            treatment = Treatment.objects.get(pk=pk)
        except Treatment.DoesNotExist:
            return Response({'error': 'Treatment not found.'}, status=status.HTTP_404_NOT_FOUND)
        materials = TreatmentMaterial.objects.filter(treatment=treatment).select_related('item')
        return Response(TreatmentMaterialSerializer(materials, many=True).data)

    def post(self, request, pk):
        try:
            treatment = Treatment.objects.get(pk=pk)
        except Treatment.DoesNotExist:
            return Response({'error': 'Treatment not found.'}, status=status.HTTP_404_NOT_FOUND)
        data = {**request.data, 'treatment': treatment.id}
        serializer = TreatmentMaterialSerializer(data=data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        item_id = serializer.validated_data['item'].id
        if InventoryItem.objects.filter(pk=item_id, is_service=True).exists():
            return Response(
                {'error': 'Cannot use a service item as a treatment material.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='CREATE',
            entity_type='TreatmentMaterial',
            entity_id=str(instance.id),
            description=f'Material added to {treatment.name}: {instance.quantity_small} {instance.item.unit_small} of {instance.item.name}',
        )
        return Response(TreatmentMaterialSerializer(instance).data, status=status.HTTP_201_CREATED)


class TreatmentMaterialDetailView(APIView):
    """PUT/DELETE a single treatment material (admin only)."""

    def _get(self, pk, mid):
        try:
            return TreatmentMaterial.objects.select_related('item').get(pk=mid, treatment_id=pk)
        except TreatmentMaterial.DoesNotExist:
            return None

    def put(self, request, pk, mid):
        instance = self._get(pk, mid)
        if instance is None:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        data = {**request.data, 'treatment': pk, 'item': instance.item_id}
        serializer = TreatmentMaterialSerializer(instance, data=data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        updated = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='TreatmentMaterial',
            entity_id=str(updated.id),
            description=f'Material updated: {updated.quantity_small} {updated.item.unit_small} of {updated.item.name}',
        )
        return Response(TreatmentMaterialSerializer(updated).data)

    def delete(self, request, pk, mid):
        instance = self._get(pk, mid)
        if instance is None:
            return Response({'error': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        AuditLog.objects.create(
            performed_by=_actor(request),
            action='DELETE',
            entity_type='TreatmentMaterial',
            entity_id=str(instance.id),
            description=f'Material removed: {instance.quantity_small} {instance.item.unit_small} of {instance.item.name}',
        )
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class BeauticianListCreateAdminView(generics.ListCreateAPIView):
    queryset = Beauticians.objects.all().order_by('beautician_name')

    def get_serializer_class(self):
        # Use rich status serializer for reads; plain one for writes
        if self.request.method == 'GET':
            return BeauticianAdminStatusSerializer
        return BeauticiansSerializer

    def perform_create(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='CREATE',
            entity_type='Beautician',
            entity_id=str(instance.id),
            description=f'Beautician added: {instance.beautician_name}',
        )


class BeauticianDetailAdminView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Beauticians.objects.all()
    serializer_class = BeauticiansSerializer

    def perform_update(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='UPDATE',
            entity_type='Beautician',
            entity_id=str(instance.id),
            description=f'Beautician updated: {instance.beautician_name}',
        )

    def perform_destroy(self, instance):
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='DELETE',
            entity_type='Beautician',
            entity_id=str(instance.id),
            description=f'Beautician deleted: {instance.beautician_name}',
        )
        instance.delete()


class BeauticianReleaseView(APIView):
    """
    POST /api/admin/beauticians/<pk>/release/
    Force-releases a beautician who is stuck (available=False but patient is no longer
    in active treatment). Sets available=True and logs the action.
    """
    def post(self, request, pk):
        try:
            beautician = Beauticians.objects.get(id=pk)
        except Beauticians.DoesNotExist:
            return Response({'error': 'Beautician not found.'}, status=status.HTTP_404_NOT_FOUND)

        if beautician.available:
            return Response(
                BeauticianAdminStatusSerializer(beautician).data,
                status=status.HTTP_200_OK,
            )

        beautician.available = True
        beautician.save(update_fields=['available'])

        AuditLog.objects.create(
            performed_by=_actor(request),
            action='UPDATE',
            entity_type='Beautician',
            entity_id=str(beautician.id),
            description=f'Beautician {beautician.beautician_name} manually released by admin',
        )

        return Response(BeauticianAdminStatusSerializer(beautician).data, status=status.HTTP_200_OK)


class AppUserListCreateAdminView(generics.ListCreateAPIView):
    queryset = AppUser.objects.all().order_by('display_name')
    serializer_class = AppUserAdminSerializer

    def perform_create(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='CREATE',
            entity_type='AppUser',
            entity_id=str(instance.id),
            description=f'User created: {instance.display_name} ({instance.role})',
        )


class AppUserDetailAdminView(generics.RetrieveUpdateDestroyAPIView):
    queryset = AppUser.objects.all()
    serializer_class = AppUserAdminSerializer

    def perform_update(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='UPDATE',
            entity_type='AppUser',
            entity_id=str(instance.id),
            description=f'User updated: {instance.display_name} ({instance.role})',
        )

    def perform_destroy(self, instance):
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='DELETE',
            entity_type='AppUser',
            entity_id=str(instance.id),
            description=f'User deleted: {instance.display_name}',
        )
        instance.delete()


# ── Chart of Accounts ──────────────────────────────────────────────────────

class ChartOfAccountsListCreateView(generics.ListCreateAPIView):
    serializer_class = ChartOfAccountsSerializer

    def get_serializer_context(self):
        context = super().get_serializer_context()
        # Treatment / item counts per category, so the COA page can show how
        # much each per-category account actually covers without an N+1.
        counts = {}
        for cat_id, is_service, total in (
            InventoryItem.objects
            .filter(item_category__isnull=False)
            .values_list('item_category_id', 'is_service')
            .annotate(total=Count('id'))
        ):
            treatments, items = counts.get(cat_id, (0, 0))
            if is_service:
                treatments = total
            else:
                items = total
            counts[cat_id] = (treatments, items)
        context['category_counts'] = counts
        return context

    def get_queryset(self):
        qs = (
            ChartOfAccounts.objects
            .select_related(
                'parent',
                'treatment_category',
                'supplier_ap',
            )
            .order_by('account_number')
        )
        range_param = self.request.query_params.get('range', '').strip()
        if range_param == 'cash':
            # Return sub-accounts of the system cash head (account 1100000).
            # This is the authoritative filter for POS payment method accounts.
            qs = qs.filter(parent__account_number=1100000, is_head=False)
        elif range_param == 'inventory':
            qs = qs.filter(account_number__gte=1200000, account_number__lte=1299999)
        return qs

    def perform_create(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='CREATE',
            entity_type='ChartOfAccounts',
            entity_id=str(instance.id),
            description=f'Account created: {instance.account_number} – {instance.name}',
        )


class ChartOfAccountsDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = ChartOfAccounts.objects.all()
    serializer_class = ChartOfAccountsSerializer

    def perform_update(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='UPDATE',
            entity_type='ChartOfAccounts',
            entity_id=str(instance.id),
            description=f'Account updated: {instance.account_number} – {instance.name}',
        )

    def perform_destroy(self, instance):
        if instance.is_system:
            raise ValidationError('System accounts cannot be deleted.')
        if instance.is_head:
            if instance.sub_accounts.exists():
                raise ValidationError('Head accounts with sub-accounts cannot be deleted. Remove all sub-accounts first.')
            raise ValidationError('Head accounts cannot be deleted.')
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='DELETE',
            entity_type='ChartOfAccounts',
            entity_id=str(instance.id),
            description=f'Account deleted: {instance.account_number} – {instance.name}',
        )
        instance.delete()


class AccountLedgerView(APIView):
    """GET /api/admin/accounts/<pk>/ledger/

    Query params (all optional):
      date_from   YYYY-MM-DD
      date_to     YYYY-MM-DD
      entry_type  'debit' | 'credit'

    Entries come back ASCENDING (oldest first) — a running-balance ledger has
    to read top-down — each carrying `balance`, the account's balance after
    that entry. `opening_balance` is the natural-signed balance before
    `date_from` (0 when no `date_from`); `closing_balance` is the balance after
    every entry in the window, regardless of the `entry_type` display filter.
    """

    def get(self, request, pk):
        try:
            account = ChartOfAccounts.objects.get(pk=pk)
        except ChartOfAccounts.DoesNotExist:
            return Response({'error': 'Account not found.'}, status=status.HTTP_404_NOT_FOUND)

        date_from = request.query_params.get('date_from', '').strip()
        date_to   = request.query_params.get('date_to', '').strip()
        etype     = request.query_params.get('entry_type', '').strip().lower()

        rows, opening, closing, total_debit, total_credit = ledger_rows_with_balance(
            account, date_from=date_from, date_to=date_to, entry_type=etype,
        )

        # LedgerEntrySerializer is shared with other endpoints, so `balance` is
        # not a serializer field — zip it onto the serialized dicts instead.
        # `rows` and `.data` are the same list in the same order.
        entries = LedgerEntrySerializer(rows, many=True).data
        for row, entry in zip(rows, entries):
            entry['balance'] = str(row.running_balance)

        return Response({
            # Full account serializer, not a hand-rolled dict: the ledger page
            # is where an account's linkage (the treatment category or vendor
            # it serves) is shown now that the COA list dropped that column.
            'account': ChartOfAccountsSerializer(account).data,
            'entries': entries,
            'opening_balance': str(opening),
            'closing_balance': str(closing),
            'total_debit': str(total_debit),
            'total_credit': str(total_credit),
        })


# ── Ledger print (PDF) ───────────────────────────────────────────────────────

# A full-history print of a busy cash account materialises every row twice —
# once as data, once as a platypus flowable. Cap the window rather than let a
# stray click OOM the server.
MAX_LEDGER_PRINT_DAYS = 366

_BOOL_TRUE = {'1', 'true', 'yes', 'on', 't', 'y'}
_BOOL_FALSE = {'0', 'false', 'no', 'off', 'f', 'n'}

_MAX_TITLE_LEN = 120
_MAX_SUBTITLE_LEN = 300


def _bad(message):
    return Response({'error': message}, status=status.HTTP_400_BAD_REQUEST)


def _param_date(raw, name):
    """-> (date|None, error|None). Blank is None, garbage is an error — never a
    silent fallback to 'today' or 'all time'."""
    raw = (raw or '').strip()
    if not raw:
        return None, None
    try:
        return datetime.strptime(raw, '%Y-%m-%d').date(), None
    except ValueError:
        return None, f"Format {name} tidak valid: '{raw}'. Gunakan YYYY-MM-DD."


def _param_enum(raw, allowed, name):
    """-> (value, error|None). An unrecognised value is a 400, so a typo in the
    query string can never quietly print a different report than the one asked
    for."""
    value = (raw or '').strip().lower()
    if value not in allowed:
        shown = ', '.join(repr(a) for a in allowed)
        return None, f"Nilai {name} tidak dikenal: '{raw}'. Pilihan: {shown}."
    return value, None


def _param_bool(raw, name, default=True):
    if raw is None or not str(raw).strip():
        return default, None
    value = str(raw).strip().lower()
    if value in _BOOL_TRUE:
        return True, None
    if value in _BOOL_FALSE:
        return False, None
    return None, f"Nilai {name} tidak dikenal: '{raw}'. Gunakan true atau false."


class AccountLedgerPrintView(APIView):
    """GET /api/admin/accounts/<pk>/ledger/print/ → application/pdf

    The printable twin of ``AccountLedgerView``. Both read the same rows from
    ``ledger_rows_with_balance()``, so the PDF and the screen can never
    disagree about a balance.

    Query params — see docs/DESIGN_expense_redesign_and_coa_print.md §4.2:
      date_from, date_to   YYYY-MM-DD, both required
      entry_type           '' | debit | credit          (default '')
      title                str                          (default 'Buku Besar')
      subtitle             str                          (default '')
      group_by             none | day | month           (default none)
      page_break           none | group                 (default none)
      show_opening / show_running / show_subtotals  bool (default true)
      orientation          portrait | landscape         (default portrait)

    Served ``inline`` so the browser opens it in a tab the user can Ctrl+P,
    rather than dropping a file in Downloads.
    """

    def get(self, request, pk):
        try:
            account = ChartOfAccounts.objects.get(pk=pk)
        except ChartOfAccounts.DoesNotExist:
            return Response({'error': 'Account not found.'}, status=status.HTTP_404_NOT_FOUND)

        q = request.query_params

        date_from, err = _param_date(q.get('date_from'), 'date_from')
        if err:
            return _bad(err)
        date_to, err = _param_date(q.get('date_to'), 'date_to')
        if err:
            return _bad(err)
        if not date_from or not date_to:
            return _bad('date_from dan date_to wajib diisi untuk mencetak buku besar.')
        if date_from > date_to:
            return _bad('date_from tidak boleh melewati date_to.')

        span_days = (date_to - date_from).days + 1
        if span_days > MAX_LEDGER_PRINT_DAYS:
            return _bad(
                f'Rentang {span_days} hari terlalu panjang untuk dicetak '
                f'(maksimal {MAX_LEDGER_PRINT_DAYS} hari). Persempit rentang tanggalnya.'
            )

        entry_type, err = _param_enum(q.get('entry_type'), ('', 'debit', 'credit'), 'entry_type')
        if err:
            return _bad(err)
        group_by, err = _param_enum(q.get('group_by') or 'none', ('none', 'day', 'month'), 'group_by')
        if err:
            return _bad(err)
        page_break, err = _param_enum(q.get('page_break') or 'none', ('none', 'group'), 'page_break')
        if err:
            return _bad(err)
        orientation, err = _param_enum(
            q.get('orientation') or 'portrait', ('portrait', 'landscape'), 'orientation',
        )
        if err:
            return _bad(err)

        if page_break == 'group' and group_by == 'none':
            return _bad(
                'page_break=group memerlukan group_by=day atau group_by=month — '
                'tanpa pengelompokan tidak ada kelompok untuk dipisah per halaman.'
            )

        show_opening, err = _param_bool(q.get('show_opening'), 'show_opening')
        if err:
            return _bad(err)
        show_running, err = _param_bool(q.get('show_running'), 'show_running')
        if err:
            return _bad(err)
        show_subtotals, err = _param_bool(q.get('show_subtotals'), 'show_subtotals')
        if err:
            return _bad(err)

        rows, opening, closing, total_debit, total_credit = ledger_rows_with_balance(
            account, date_from=date_from, date_to=date_to, entry_type=entry_type,
        )

        # An accountant printing an incomplete ledger with no warning is the
        # worst failure mode this endpoint has.
        unposted_count = len(unposted_dates_in_range(date_from, date_to))

        actor = _actor(request)
        title = (q.get('title') or 'Buku Besar').strip()[:_MAX_TITLE_LEN] or 'Buku Besar'
        subtitle = (q.get('subtitle') or '').strip()[:_MAX_SUBTITLE_LEN]

        # Imported here, not at module scope: reportlab is a heavy third-party
        # dependency and a missing/broken install should take down this one
        # endpoint, not every admin route that shares this module.
        from ..services.ledger_pdf import build_ledger_pdf

        pdf = build_ledger_pdf(
            account, rows, opening, closing,
            {'debit': total_debit, 'credit': total_credit},
            {
                'date_from': date_from,
                'date_to': date_to,
                'entry_type': entry_type,
                'title': title,
                'subtitle': subtitle,
                'group_by': group_by,
                'page_break': page_break,
                'show_opening': show_opening,
                'show_running': show_running,
                'show_subtotals': show_subtotals,
                'orientation': orientation,
                'printed_by': getattr(actor, 'display_name', '') or '',
                'printed_at': timezone.localtime(),
                'unposted_count': unposted_count,
            },
        )

        filename = f'buku-besar-{account.account_number}-{date_from}-{date_to}.pdf'
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        response['Content-Length'] = str(len(pdf))
        return response


# ── Payment Methods ──────────────────────────────────────────────────────────

class PaymentMethodListCreateView(generics.ListCreateAPIView):
    serializer_class = PaymentMethodSerializer

    def get_queryset(self):
        qs = PaymentMethod.objects.select_related('linked_account').order_by('sort_order', 'name')
        if self.request.query_params.get('active_only', '').strip().lower() in ('1', 'true', 'yes'):
            qs = qs.filter(is_active=True)
        return qs

    def perform_create(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='CREATE',
            entity_type='PaymentMethod',
            entity_id=str(instance.id),
            description=f'Payment method created: {instance.name}',
        )


class PaymentMethodDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = PaymentMethod.objects.select_related('linked_account').all()
    serializer_class = PaymentMethodSerializer

    def perform_update(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='UPDATE',
            entity_type='PaymentMethod',
            entity_id=str(instance.id),
            description=f'Payment method updated: {instance.name}',
        )

    def perform_destroy(self, instance):
        if instance.is_system:
            raise ValidationError('System payment methods cannot be deleted.')
        if instance.invoices.exists() or instance.purchase_invoices.exists():
            raise ValidationError('Payment methods already used on invoices cannot be deleted. Deactivate it instead.')
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='DELETE',
            entity_type='PaymentMethod',
            entity_id=str(instance.id),
            description=f'Payment method deleted: {instance.name}',
        )
        instance.delete()


# ── Treatment Categories ────────────────────────────────────────────────────

class TreatmentCategoryListCreateView(generics.ListCreateAPIView):
    queryset = TreatmentCategory.objects.all()
    serializer_class = TreatmentCategorySerializer

    def perform_create(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='CREATE',
            entity_type='TreatmentCategory',
            entity_id=str(instance.id),
            description=f'Treatment category added: {instance.name}',
        )


class TreatmentCategoryDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = TreatmentCategory.objects.all()
    serializer_class = TreatmentCategorySerializer

    def perform_update(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='UPDATE',
            entity_type='TreatmentCategory',
            entity_id=str(instance.id),
            description=f'Treatment category updated: {instance.name}',
        )

    def perform_destroy(self, instance):
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='DELETE',
            entity_type='TreatmentCategory',
            entity_id=str(instance.id),
            description=f'Treatment category deleted: {instance.name}',
        )
        instance.delete()


class TreatmentCategoryAuditAccountsView(APIView):
    """GET — return every category with per-account presence flags."""

    def get(self, request):
        cats = TreatmentCategory.objects.select_related('revenue_account').order_by('name')
        results = []
        for cat in cats:
            results.append({
                'id': cat.id,
                'name': cat.name,
                'revenue_account': str(cat.revenue_account) if cat.revenue_account_id else None,
                'needs_provisioning': not cat.revenue_account_id,
            })
        any_missing = any(r['needs_provisioning'] for r in results)
        return Response({'categories': results, 'any_missing': any_missing})


class TreatmentCategoryProvisionAccountsView(APIView):
    """POST — create missing COA accounts for all (or specified) categories.

    Body (optional): { "category_ids": [1, 2, 3] }
    Omit body / send empty list to provision ALL categories that need it.
    """

    def post(self, request):
        category_ids = request.data.get('category_ids') or []
        qs = TreatmentCategory.objects.select_related('revenue_account')
        if category_ids:
            qs = qs.filter(pk__in=category_ids)

        provisioned = []
        for cat in qs:
            needs = not cat.revenue_account_id
            if needs:
                created = cat.ensure_accounts()
                if created:
                    provisioned.append({'id': cat.id, 'name': cat.name, 'created': created})
                    AuditLog.objects.create(
                        performed_by=_actor(request),
                        action='CREATE',
                        entity_type='TreatmentCategory',
                        entity_id=str(cat.id),
                        description=f'Auto-provisioned COA accounts for category: {cat.name} ({", ".join(created)})',
                    )

        return Response({'provisioned': provisioned, 'count': len(provisioned)})


# ── Excel template + import ───────────────────────────────────────────────────

class TreatmentTemplateView(APIView):
    """GET — download a blank .xlsx template for bulk treatment import."""

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Treatments'

        headers = ['code', 'name', 'category', 'price', 'active']
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = _HEADER_FONT
            cell.fill      = _HEADER_FILL
            cell.alignment = Alignment(horizontal='center')

        notes = [
            'Unique treatment code (required)',
            'Treatment name (required)',
            'Category name (optional)',
            'Price in IDR (required, e.g. 150000)',
            'yes / no (optional, default yes)',
        ]
        ws.append(notes)
        note_font = Font(italic=True, color='595959')
        for cell in ws[2]:
            cell.font      = note_font
            cell.alignment = Alignment(wrap_text=True)

        examples = [
            ('FC-001', 'Facial Basic',    'Facial',    150000, 'yes'),
            ('FC-002', 'Facial Premium',  'Facial',    250000, 'yes'),
            ('SK-001', 'Skin Brightening', 'Skincare', 320000, 'yes'),
        ]
        for row in examples:
            ws.append(list(row))

        col_widths = [16, 30, 20, 16, 12]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        ws.row_dimensions[2].height = 36

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="treatments_template.xlsx"'
        return response


class TreatmentImportView(APIView):
    """
    POST multipart/form-data with file=<xlsx>
    Columns (row 1 = header, skipped):
      code | name | category | price | active
    active: yes/true/1 → True; no/false/0 → False; blank → True

    ?preview=true  — parse only, no DB writes.
      Returns { rows: [{row, code, name, category, price, active, action, error}] }
      action: 'create' | 'update' | 'error'

    Normal POST — commits all valid rows.
      Returns { created, updated, errors: [{row, message}] }
    """

    @staticmethod
    def _parse_rows(ws_rows):
        """Parse worksheet rows (skipping header) into a list of dicts."""
        result = []
        for i, row in enumerate(ws_rows[1:], start=2):
            padded = list(row) + [None] * 5
            code       = str(padded[0]).strip() if padded[0] is not None else ''
            name       = str(padded[1]).strip() if padded[1] is not None else ''
            category   = str(padded[2]).strip() if padded[2] is not None else ''
            price_raw  = padded[3]
            active_raw = padded[4]

            error = None
            price = None

            if not code:
                error = 'Code is required.'
            elif not name:
                error = 'Name is required.'
            else:
                try:
                    price = float(price_raw) if price_raw not in (None, '') else None
                    if price is None:
                        raise ValueError
                except (ValueError, TypeError):
                    error = f'Invalid price "{price_raw}".'

            if active_raw is None or str(active_raw).strip() == '':
                active = True
            else:
                active = str(active_raw).strip().lower() in ('yes', 'true', '1')

            result.append({
                'row': i,
                'code': code,
                'name': name,
                'category': category,
                'price': price,
                'active': active,
                'error': error,
            })
        return result

    def post(self, request):
        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), read_only=True, data_only=True)
        except Exception:
            return Response({'error': 'Could not parse file. Upload a valid .xlsx file.'}, status=status.HTTP_400_BAD_REQUEST)

        ws = wb.active
        ws_rows = list(ws.iter_rows(values_only=True))
        if not ws_rows:
            return Response({'error': 'File is empty.'}, status=status.HTTP_400_BAD_REQUEST)

        parsed = self._parse_rows(ws_rows)

        # ── Preview mode ────────────────────────────────────────────────────
        if request.query_params.get('preview') == 'true':
            existing_codes = set(
                Treatment.objects.filter(
                    code__in=[r['code'] for r in parsed if not r['error']]
                ).values_list('code', flat=True)
            )
            rows = []
            for r in parsed:
                if r['error']:
                    action = 'error'
                elif r['code'] in existing_codes:
                    action = 'update'
                else:
                    action = 'create'
                rows.append({
                    'row':      r['row'],
                    'code':     r['code'],
                    'name':     r['name'],
                    'category': r['category'],
                    'price':    r['price'],
                    'active':   r['active'],
                    'action':   action,
                    'error':    r['error'],
                })
            return Response({'rows': rows}, status=status.HTTP_200_OK)

        # ── Commit mode ──────────────────────────────────────────────────────
        created = updated = 0
        errors = []

        for r in parsed:
            if r['error']:
                errors.append({'row': r['row'], 'message': r['error']})
                continue
            try:
                with transaction.atomic():
                    obj, was_created = Treatment.objects.update_or_create(
                        code=r['code'],
                        defaults={
                            'name':     r['name'],
                            'category': r['category'],
                            'price':    r['price'],
                            'active':   r['active'],
                        },
                    )
                    AuditLog.objects.create(
                        performed_by=_actor(request),
                        action='CREATE' if was_created else 'UPDATE',
                        entity_type='Treatment',
                        entity_id=str(obj.id),
                        description=f'Treatment {"imported" if was_created else "updated via import"}: {obj.name} ({obj.code})',
                    )
                if was_created:
                    created += 1
                else:
                    updated += 1
            except Exception as exc:
                errors.append({'row': r['row'], 'message': str(exc)})

        return Response({'created': created, 'updated': updated, 'errors': errors}, status=status.HTTP_200_OK)


# ── Treatment Packages ─────────────────────────────────────────────────────

class TreatmentPackageListCreateAdminView(generics.ListCreateAPIView):
    queryset = TreatmentPackage.objects.prefetch_related('items__treatment').order_by('name')
    serializer_class = TreatmentPackageSerializer

    def perform_create(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='CREATE',
            entity_type='TreatmentPackage',
            entity_id=str(instance.id),
            description=f'Package added: {instance.name} ({instance.code})',
        )


class TreatmentPackageDetailAdminView(generics.RetrieveUpdateDestroyAPIView):
    queryset = TreatmentPackage.objects.prefetch_related('items__treatment')
    serializer_class = TreatmentPackageSerializer

    def perform_update(self, serializer):
        instance = serializer.save()
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='UPDATE',
            entity_type='TreatmentPackage',
            entity_id=str(instance.id),
            description=f'Package updated: {instance.name} ({instance.code})',
        )

    def perform_destroy(self, instance):
        AuditLog.objects.create(
            performed_by=_actor(self.request),
            action='DELETE',
            entity_type='TreatmentPackage',
            entity_id=str(instance.id),
            description=f'Package deleted: {instance.name} ({instance.code})',
        )
        instance.delete()



class TreatmentPackageSyncView(generics.ListAPIView):
    """Read-only active package list — accessible to all authenticated users for POS sync."""
    queryset = TreatmentPackage.objects.filter(active=True).prefetch_related('items__treatment').order_by('name')
    serializer_class = TreatmentPackageSerializer


class SiteConfigView(APIView):
    """GET/PUT the singleton clinic receipt configuration."""

    def get(self, request):
        return Response(SiteConfigSerializer(SiteConfig.get_solo()).data)

    def put(self, request):
        obj = SiteConfig.get_solo()
        serializer = SiteConfigSerializer(obj, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ColorPaletteListCreateView(APIView):
    permission_classes = [IsAppAuthenticated]

    def get(self, request):
        palettes = ColorPalette.objects.all()
        return Response(ColorPaletteSerializer(palettes, many=True).data)

    def post(self, request):
        if request.user.role not in ('superuser', 'manager'):
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        serializer = ColorPaletteSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ColorPaletteDetailView(APIView):
    permission_classes = [IsAppAuthenticated]

    def _get(self, pk):
        try:
            return ColorPalette.objects.get(pk=pk)
        except ColorPalette.DoesNotExist:
            return None

    def patch(self, request, pk):
        if request.user.role not in ('superuser', 'manager'):
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        obj = self._get(pk)
        if obj is None:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ColorPaletteSerializer(obj, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        if request.user.role not in ('superuser', 'manager'):
            return Response({'error': 'Forbidden'}, status=status.HTTP_403_FORBIDDEN)
        obj = self._get(pk)
        if obj is None:
            return Response({'error': 'Not found'}, status=status.HTTP_404_NOT_FOUND)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
