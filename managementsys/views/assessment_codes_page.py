import io

import openpyxl
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView

from ..models import AssessmentCode
from ..api.serializers import AssessmentCodeSerializer

_HEADER_FONT = Font(bold=True, color='FFFFFF')
_HEADER_FILL = PatternFill('solid', fgColor='4F81BD')


class AssessmentCodeListCreateView(generics.ListCreateAPIView):
    serializer_class = AssessmentCodeSerializer

    def get_queryset(self):
        qs = AssessmentCode.objects.all()
        search = self.request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(Q(code__icontains=search) | Q(description__icontains=search))
        return qs


class AssessmentCodeDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = AssessmentCode.objects.all()
    serializer_class = AssessmentCodeSerializer


class AssessmentCodeTemplateDownloadView(APIView):
    """GET — download a blank sample .xlsx with ICD-10 import format and example rows."""

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'ICD-10 Codes'

        headers = ['Code', 'Description', 'Category', 'Active']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL

        samples = [
            ('L70.0', 'Acne vulgaris', 'Common', 'TRUE'),
            ('L81.1', 'Chloasma', 'Common', 'TRUE'),
            ('L57.0', 'Actinic keratosis', 'Uncommon', 'TRUE'),
            ('L90.5', 'Scar conditions and fibrosis of skin', 'Uncommon', 'FALSE'),
        ]
        for row in samples:
            ws.append(list(row))

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 10

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="icd10_import_template.xlsx"'
        return response


class AssessmentCodeImportPreviewView(APIView):
    """Parse uploaded .xlsx and return rows for user review — no DB writes.

    Expected columns (row 1 = header, skipped):
      A: Code        — ICD-10 code, e.g. L70.0
      B: Description — human-readable label
      C: Category    — 1 / "Common" or 2 / "Uncommon" (default: Common)
      D: Active      — TRUE/1/Yes or FALSE/0/No (default: True)
    """

    def post(self, request):
        uploaded = request.FILES.get('file')
        if not uploaded:
            return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(uploaded.read()), read_only=True, data_only=True)
        except Exception:
            return Response(
                {'error': 'Could not parse file. Upload a valid .xlsx file.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb.active
        rows = []
        errors = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None for cell in row):
                continue

            code_val = str(row[0]).strip().upper() if row[0] is not None else ''
            desc_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''

            raw_cat = row[2] if len(row) > 2 else None
            if raw_cat is None or str(raw_cat).strip() == '':
                category = 1
            elif str(raw_cat).strip().lower() in ('1', 'common'):
                category = 1
            elif str(raw_cat).strip().lower() in ('2', 'uncommon'):
                category = 2
            else:
                errors.append(f'Row {i}: invalid category "{raw_cat}" — use 1/Common or 2/Uncommon.')
                continue

            raw_active = row[3] if len(row) > 3 else None
            if raw_active is None or str(raw_active).strip() == '':
                active = True
            else:
                active = str(raw_active).strip().lower() not in ('false', '0', 'no')

            if not code_val:
                errors.append(f'Row {i}: Code is empty.')
                continue
            if not desc_val:
                errors.append(f'Row {i}: Description is empty.')
                continue

            rows.append({'code': code_val, 'description': desc_val, 'category': category, 'active': active})

        return Response({'rows': rows, 'errors': errors})


class AssessmentCodeImportConfirmView(APIView):
    """Bulk create / update assessment codes from previewed rows."""

    def post(self, request):
        rows = request.data.get('rows', [])
        if not rows:
            return Response({'error': 'No rows to import.'}, status=status.HTTP_400_BAD_REQUEST)

        to_upsert = []
        skipped = 0
        for row in rows:
            code = str(row.get('code', '')).strip().upper()
            desc = str(row.get('description', '')).strip()
            try:
                category = int(row.get('category', 1))
            except (TypeError, ValueError):
                category = 1
            active = bool(row.get('active', True))
            if not code or not desc:
                skipped += 1
                continue
            to_upsert.append(AssessmentCode(code=code, description=desc, category=category, active=active))

        if not to_upsert:
            return Response({'created': 0, 'updated': 0, 'skipped': skipped})

        existing_codes = set(
            AssessmentCode.objects.filter(code__in=[obj.code for obj in to_upsert])
            .values_list('code', flat=True)
        )
        new_objs = [obj for obj in to_upsert if obj.code not in existing_codes]
        update_objs = [obj for obj in to_upsert if obj.code in existing_codes]

        with transaction.atomic():
            if new_objs:
                AssessmentCode.objects.bulk_create(new_objs, ignore_conflicts=True)
            if update_objs:
                AssessmentCode.objects.bulk_update(
                    update_objs, ['description', 'category', 'active'], batch_size=500
                )

        return Response({'created': len(new_objs), 'updated': len(update_objs), 'skipped': skipped})
