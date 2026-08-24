"""Parse a bank statement export into signed statement lines.

Indonesian banks export wildly different spreadsheets, so this deliberately does
**not** try to be clever about any one of them. It looks for a header row by
name, accepts the handful of column spellings that actually turn up, and reports
anything it cannot read as a per-row problem the operator sees before importing.

Two-phase, like every other import in this codebase: ``parse`` returns rows for
review and writes nothing; the view writes only what the operator confirmed.

Sign convention
---------------
Everything comes out as one signed number — **positive money in, negative money
out** — because that is what makes the matching arithmetic in
``bank_reconciliation`` plain subtraction. Three input shapes are supported and
all collapse to it:

  * separate debit/credit (or *masuk*/*keluar*) columns
  * a single signed amount column
  * an amount column plus a direction column (DB/CR, D/K)

The ambiguous case — a single *unsigned* amount column with no direction — is
rejected rather than guessed. Guessing here silently reverses half a statement.
"""
import csv
import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

# Header spellings seen in the wild, lowercased and stripped of punctuation.
DATE_HEADERS = {
    'tanggal', 'tgl', 'date', 'transaction date', 'tanggal transaksi',
    'posting date', 'tgl transaksi',
}
DESCRIPTION_HEADERS = {
    'keterangan', 'description', 'uraian', 'deskripsi', 'remark', 'remarks',
    'transaction description', 'narasi', 'berita',
}
REFERENCE_HEADERS = {
    'referensi', 'reference', 'ref', 'no referensi', 'no ref', 'trace no',
    'nomor referensi', 'kode transaksi',
}
DEBIT_HEADERS = {'debit', 'debet', 'keluar', 'pengeluaran', 'withdrawal', 'dr', 'kredit keluar'}
CREDIT_HEADERS = {'kredit', 'credit', 'masuk', 'pemasukan', 'deposit', 'cr', 'setoran'}
AMOUNT_HEADERS = {'jumlah', 'amount', 'nominal', 'nilai', 'mutasi'}
DIRECTION_HEADERS = {'dk', 'd/k', 'db/cr', 'tipe', 'type', 'jenis', 'arah'}

DEBIT_MARKERS = {'d', 'db', 'dr', 'debit', 'debet', 'k', 'keluar', 'out'}
CREDIT_MARKERS = {'c', 'cr', 'k', 'kredit', 'credit', 'masuk', 'in'}

# 'k' means kredit in some exports and keluar in others, so it is never decisive.
AMBIGUOUS_MARKERS = {'k'}

DATE_FORMATS = [
    '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%d-%m-%y',
    '%m/%d/%Y', '%d %b %Y', '%d %B %Y', '%Y/%m/%d',
]


class StatementParseError(Exception):
    """The file as a whole could not be read. Per-row problems are not errors."""


def _norm(value):
    return ' '.join(str(value or '').strip().lower().replace('.', '').split())


def _parse_date(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(value):
    """A number out of whatever the bank wrote, or None.

    Handles both Indonesian (1.234.567,89) and English (1,234,567.89) grouping
    by deciding from which separator appears last, and strips currency prefixes
    and parenthesised negatives.
    """
    if value in (None, ''):
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return None

    negative = text.startswith('(') and text.endswith(')')
    if negative:
        text = text[1:-1]

    text = (text.replace('Rp', '').replace('rp', '').replace('IDR', '')
                .replace('idr', '').replace(' ', '').replace(' ', ''))
    if not text:
        return None

    last_dot = text.rfind('.')
    last_comma = text.rfind(',')
    if last_comma > last_dot:
        # Indonesian: dots group, comma is the decimal point.
        text = text.replace('.', '').replace(',', '.')
    else:
        text = text.replace(',', '')

    try:
        amount = Decimal(text)
    except InvalidOperation:
        return None
    return -amount if negative else amount


def _rows_from_file(upload):
    """Every cell row from an .xlsx or .csv upload, as lists."""
    name = (getattr(upload, 'name', '') or '').lower()

    if name.endswith(('.xlsx', '.xlsm', '.xls')):
        try:
            import openpyxl
        except ImportError as exc:  # pragma: no cover - openpyxl is a hard dep
            raise StatementParseError('Pembaca Excel tidak tersedia di server.') from exc
        try:
            wb = openpyxl.load_workbook(upload, data_only=True, read_only=True)
        except Exception as exc:
            raise StatementParseError(f'File Excel tidak dapat dibaca: {exc}') from exc
        ws = wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]

    raw = upload.read()
    if isinstance(raw, bytes):
        for encoding in ('utf-8-sig', 'utf-8', 'cp1252', 'latin-1'):
            try:
                raw = raw.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise StatementParseError('Encoding file tidak dikenali.')

    sample = raw[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(io.StringIO(raw), dialect)]


def _find_header(rows):
    """(index, {role: column}) for the first row that looks like a header.

    Statements routinely carry several lines of account-holder preamble above
    the table, so the header is searched for rather than assumed to be row 1.
    """
    for index, row in enumerate(rows[:30]):
        mapping = {}
        for col, cell in enumerate(row):
            key = _norm(cell)
            if not key:
                continue
            if key in DATE_HEADERS and 'date' not in mapping:
                mapping['date'] = col
            elif key in DESCRIPTION_HEADERS and 'description' not in mapping:
                mapping['description'] = col
            elif key in REFERENCE_HEADERS and 'reference' not in mapping:
                mapping['reference'] = col
            elif key in DEBIT_HEADERS and 'debit' not in mapping:
                mapping['debit'] = col
            elif key in CREDIT_HEADERS and 'credit' not in mapping:
                mapping['credit'] = col
            elif key in AMOUNT_HEADERS and 'amount' not in mapping:
                mapping['amount'] = col
            elif key in DIRECTION_HEADERS and 'direction' not in mapping:
                mapping['direction'] = col

        has_money = ('debit' in mapping or 'credit' in mapping or 'amount' in mapping)
        if 'date' in mapping and has_money:
            return index, mapping

    raise StatementParseError(
        'Kolom tanggal dan jumlah tidak ditemukan. Pastikan file memiliki baris '
        'judul kolom seperti "Tanggal" dan "Debit"/"Kredit" atau "Jumlah".'
    )


def _signed(row, mapping):
    """(amount, problem) for one row, already signed. Amount is None on failure."""
    def cell(role):
        col = mapping.get(role)
        return row[col] if col is not None and col < len(row) else None

    debit = _parse_amount(cell('debit'))
    credit = _parse_amount(cell('credit'))

    if mapping.get('debit') is not None or mapping.get('credit') is not None:
        if debit and credit:
            return None, 'Baris mengisi kolom debit dan kredit sekaligus.'
        if credit:
            return abs(credit), None
        if debit:
            return -abs(debit), None
        return None, 'Tidak ada nilai debit maupun kredit.'

    amount = _parse_amount(cell('amount'))
    if amount is None:
        return None, 'Jumlah tidak dapat dibaca.'

    direction = _norm(cell('direction'))
    if direction:
        # 'k' is kredit in some exports and keluar in others, so it never
        # decides on its own -- see AMBIGUOUS_MARKERS.
        if direction in AMBIGUOUS_MARKERS:
            return None, ('Arah transaksi "K" ambigu (kredit atau keluar). '
                          'Gunakan kolom Debit/Kredit terpisah.')
        if direction in CREDIT_MARKERS:
            return abs(amount), None
        if direction in DEBIT_MARKERS:
            return -abs(amount), None
        return None, f'Arah transaksi "{direction}" tidak dikenali.'

    if amount == 0:
        return None, 'Jumlah nol.'
    if amount > 0:
        # An unsigned amount column with no direction cannot be read. Guessing
        # silently reverses half a statement -- see the module docstring.
        return None, ('Kolom jumlah tidak bertanda dan tidak ada kolom arah. '
                      'Gunakan kolom Debit/Kredit terpisah atau jumlah bertanda.')
    return amount, None


def parse(upload, *, date_from=None, date_to=None):
    """Read a statement file into review rows. Writes nothing.

    Returns ``{'rows': [...], 'summary': {...}}``. Each row carries ``ok``, and
    when False, ``problem`` — a sentence explaining what could not be read.
    Out-of-period rows are marked ``out_of_period`` rather than dropped, so an
    operator who picked the wrong file finds out from the preview instead of
    from an empty import.
    """
    rows = _rows_from_file(upload)
    if not rows:
        raise StatementParseError('File kosong.')

    header_index, mapping = _find_header(rows)

    def cell(row, role):
        col = mapping.get(role)
        if col is None or col >= len(row):
            return ''
        return str(row[col] or '').strip()

    out = []
    for offset, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        if not any(str(c or '').strip() for c in row):
            continue

        parsed_date = _parse_date(row[mapping['date']] if mapping['date'] < len(row) else None)
        amount, problem = _signed(row, mapping)

        if parsed_date is None:
            # A statement's totals row has no date and no readable amount; it is
            # noise, not a problem worth showing.
            if amount is None:
                continue
            problem = problem or 'Tanggal tidak dapat dibaca.'

        out_of_period = bool(
            parsed_date and (
                (date_from and parsed_date < date_from) or
                (date_to and parsed_date > date_to)
            )
        )

        out.append({
            'row_number':    offset,
            'date':          parsed_date.isoformat() if parsed_date else None,
            'description':   cell(row, 'description')[:255],
            'reference':     cell(row, 'reference')[:100],
            'amount':        str(amount) if amount is not None else None,
            'ok':            bool(parsed_date and amount is not None and not out_of_period),
            'out_of_period': out_of_period,
            'problem':       (problem if problem else
                              ('Di luar periode rekening koran.' if out_of_period else '')),
        })

    usable = [r for r in out if r['ok']]
    return {
        'rows': out,
        'summary': {
            'total_rows':    len(out),
            'usable_rows':   len(usable),
            'skipped_rows':  len(out) - len(usable),
            'total_amount':  str(sum((Decimal(r['amount']) for r in usable), Decimal('0'))),
            'columns_found': sorted(mapping.keys()),
        },
    }
