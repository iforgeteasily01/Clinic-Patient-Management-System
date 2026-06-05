"""
SO template builder — uses so_template_sample.xlsx as the source of truth
for valid production item codes (local dev DB has no production data).
"""
import openpyxl
from difflib import SequenceMatcher

# ── Load production codes from template sample ─────────────────────────────
wb_sample = openpyxl.load_workbook('../so_template_sample.xlsx')
ws_sample = wb_sample.active
prod_codes = [
    row[0] for row in ws_sample.iter_rows(min_row=2, max_col=1, values_only=True)
    if row[0]
]
print(f"Production codes available: {len(prod_codes)}")
prod_codes_upper = {c.upper(): c for c in prod_codes}

def prod_lookup(query):
    q = query.strip()
    # exact
    hit = prod_codes_upper.get(q.upper())
    if hit:
        return hit, 1.0, 'EXACT'
    # fuzzy
    best_score, best_code = 0, None
    for code in prod_codes:
        score = SequenceMatcher(None, q.lower(), code.lower()).ratio()
        if score > best_score:
            best_score, best_code = score, code
    if best_score >= 0.70:
        return best_code, best_score, f'FUZZY {best_score:.0%}'
    return None, best_score, f'NOT FOUND ({best_score:.0%})'

# ── Manual overrides (paper name -> production code) ───────────────────────
# Verified against iPos DB; fuzzy misses corrected manually
OVERRIDES = {
    # Cosmels
    'Cosmels HyGlow':               'COSMELS HYGLOW FACIAL WASH',
    'Cosmels Acne Care':            'COSMELS ACNE CARE FACIAL WASH',
    'Cosmels Serum Glow':           'Cosmels Glow Serum',
    'Cosmels Brigtening Cream':     'Cosmels Brightening Cream',
    'Cosmels Acne Dry lotion':      'Cosmels Acne Dry Lotion',
    # Dermaverse
    'Dermaverse Serum AHA BHA':     'Dermaverse AHA BHA Booster Serum',
    'Dermaverse Serum Light 2.0':   'Serum lightening 2.0',
    'Dermaverse Serum Light 3.0':   'DERMAVERSE SERUM LIGHTENING 3.0',
    'Dermaverse Hyaluromide Complex': 'Dermaverse Serum Hyaluromide Complex',
    'Toner Dermaverse Light':       'Dermaverse Toner Normal',
    'Toner Dermaverse Oily':        'Dermaverse Toner Oily',
    # Derma XP
    'Derma XP Ad Pro Face & Body Wash':  'Derma Xp Fw Barrir Skin',
    'Derma XP Ad Pro Face & Body Moist': 'Derma XP AD Pro Face&Body Moisturizer',
    'Derma XP Ad Pro Soothing Gel':      'Derma XP AD Pro Soothing Gel',
    'Acne Clear UV Sunscreen':           'Acne Clear UV',
    # Lenfi
    'Lenfi Reti+ Pure':             'LENFI Reti Pure+',
    'Lenfi Acne Calming FW':        'Lenfi Acne Facial Wash',
    'Lenfi Gentle Peeling Solution':'Lenfi Glow Peeling',   # iPos confirmed
    'Lenfi Skin Rescue':            'Lenfi Colla-Tela',
    'Lenfi Compact Ivory':          'Lenfi Flawless Compact IVORY',
    'Lenfi Compact Beige':          'Lenfi Flawless Compact BEIGE',
    'Lenfi AHA Exfoliating Lotion': 'Lenfi AHA Exfoliating',
    # Ufines
    'Ufines Hair Tonic':            'Ufines Dandruff HairTonic',
    'Ufines Neck & Bright':         'Ufines Neck Firm',
    'Ufines Acne Drying Lot':       'UFINES ACNE DRY LOTION',
    'Ufines SPF50PA+':              'UFINES Sunscreen SPF50PA+++',
    # Sunscreen
    'Kitoderm Sunscreen Pink':      'Kitoderm Suncreen Light Pink',
    'Sunscreen Mist':               'Sunscreen Spray Mist with Microbiome Care',
    'UV Lot 50':                    'UVL 50',
    # Glow Brite
    'Glow Brite Toner':             'GLOWBRITE TONER',
    'Glow Brite Facial Wash':       'Glowbrite Facial Wash',
    # Other corrections
    'Cosmels Acne Sun Friendly':    'COSMELS Acne Sunfriendly',
    'Acne Care Body Spray Plus':    'Acne Care Body Plus',
    'Acne Care Foundation':         'Acne Care Foundation XP',
    'Eminox CL':                    'EminoxCL',
    'Eminox Plus':                  'EMINOX+',
    'Elo 2':                        'ELO2',
    'Ecto-Derm':                    'Ecto Derm',
    'RCTI 5':                       'RCTI5',
    'Sooting Toner':                'Soothing Toner',
    'FWS':                          'FWNS',             # iPos confirmed alias
    'Cleansing Milk':               'Cleansing',
    'Retinol Night Cream Serum':    'Retinol Night Cream B',
    'Lumi White':                   'Lumi Whitening Cream',
    'Ketomed Shampo':               'Ketomed',
    'Dermaten Lotion':              'DERMATEN LOTION 250ML',
    'Sodermix':                     'Sodermix Cream',
    'Translution Body Lotion':      'Trans Hand & Body Lotion',
    'XP Hair Grow Serum':           'Serum Hair Grow',
    'XP Hair Grow Shampo':          'Shampo Hair Grow',
    'Dermacept Sunscreen Acne':     'Dermacept Acne Sunscreen 47PA++',
    'Kito Infusion Mist':           'Kitoderm pro+ infusion mist',
    'Kito Milky Toner':             'Kitoderm pro+ milky toner',
    'Glam White':                   'Glam White Day Cream',
    'Zenobia Lip Serum':            'Zenobia Magic Lip Serum',
    'Diprosta Cream':               'Diprosta Cream 5g',
    'Eye Serum Skin Bright':        'Skin Bright Eye Serum',
    'Genoint':                      'Genoint Eye Cream',
    # iPos-verified — correct code different from auto-guess
    'Serum Anti Aging 1.0':         'Serum Anti Aging',   # iPos: Serum Anti Aging
    'HP 4':                         'HP 4',               # template has HP 4 (not HP 4 besar)
    'BP':                           'SCRUB BP',           # SCRUB BP is in template; BP likely refers to this
    'NF 8':                         'NF 8',               # confirmed in template
    'SB':                           'SB',                 # confirmed in template (Suu Balm)
    # Oral meds — paper uses full names, template uses shorter codes
    'Lansoprazole 30mg':            'Lansoprazole',
    'Cefixime 200mg':               'Cefixime Trihydrate',
    'Cetirizine 10mg':              'Cetirizine 10mg',
    'Clindamycin 300mg':            'Clindamycin 300mg',
    'Lincomycin 500mg':             'Linco',
    'Loratadine 10mg':              'Loratadine',
    'Metcor 16mg':                  'Metcor 16',
    'MP 8mg':                       'MP8',
    'MP 16mg':                      'MP16',
    'Kalnex 500mg':                 'Kalnex',
    'Orlibes':                      'ORLIBES KAP',
    # Partial codes from paper → template code
    'NC 1 (dikulkas)':              'NC 1',
    'NC 2 (dikulkas)':              'NC 2',
    'NCS 1 (dikulkas)':             'NCS 1',
    'ASMM (dikulkas)':              'ASMM',
    'NC 3 (dikulkas)':              'NC 3',    # NC 3 NOT in template — check below
    'AMG 1':                        'AMG1',
    'AMG 3':                        'AMG3',
    'AMG 5':                        'AMG5',
    'Pharmalene B':                 'PHARMALEN B',
    'R 1,5':                        'R1.5',
    'Betathion':                    "Betathione12'S",
    'Fyber Green':                  'GREEN V-FYBER',
    'Meditamin':                    'Meditamin per Tab',
    'Caps 22/3':                    'Caps 2 2/3',
    # R series — R0.5 and R1 are NOT in the template (only R1.5 and TONER R2 are)
    'R0,5 R':   '__MISSING__',  # NOT in production template
    'R1':       '__MISSING__',  # NOT in production template
    'R 1,5':    'R1.5',
    'R 2':      'TONER R2',
    # SPF — fix wrong fuzzy match
    'SPF 50N Tube': 'SPF Tube 50N',   # template has SPF Tube 50N, not SPF 50N Pot
    'Physical Tinted UV': 'SPF Physical T',  # template has SPF Physical T
    # ANC 4, NCT 2, Serum A2, Serum W2 — NOT in production template (diff products)
    'ANC 4':    '__MISSING__',
    'NCT 2':    '__MISSING__',
    'Serum A2': '__MISSING__',
    'Serum W2': '__MISSING__',
}

def resolve(old_name):
    # Check override first
    candidate = OVERRIDES.get(old_name)
    if candidate is not None:
        if candidate == '__MISSING__':
            return old_name, False, 'CONFIRMED-MISSING'
        # verify candidate is in production
        if candidate.upper() in prod_codes_upper:
            actual = prod_codes_upper[candidate.upper()]
            return actual, True, 'OVERRIDE'
        else:
            return candidate, False, 'OVERRIDE-MISSING'
    # Direct lookup
    code, score, method = prod_lookup(old_name)
    if code:
        return code, True, method
    return old_name, False, method

# ── Load old SO data ────────────────────────────────────────────────────────
wb_old = openpyxl.load_workbook(
    '../@ STOCK OBAT GUDANG 1 DAN OBAT MINUM (Autosaved).xlsx', data_only=True
)

def extract(ws, start_row):
    out = []
    for row in ws.iter_rows(min_row=start_row, max_col=1, values_only=True):
        v = row[0]
        if v and isinstance(v, str) and v.strip():
            out.append(v.strip())
    return out

g1 = [x for x in extract(wb_old['GUDANG 1'], 6) if not x.startswith('Note')]

section_hdrs = {'Obat Jarang Terjual', 'Obat Display', 'Obat Minum'}
om_jarang, om_display, om_minum = [], [], []
cur_sec = None
for name in extract(wb_old['OBAT MINUM & OBT JARANG'], 7):
    if name in section_hdrs:
        cur_sec = name; continue
    if cur_sec == 'Obat Jarang Terjual': om_jarang.append(name)
    elif cur_sec == 'Obat Display':      om_display.append(name)
    elif cur_sec == 'Obat Minum':        om_minum.append(name)

groups = [
    ('GUDANG 1',           g1),
    ('OBAT JARANG TERJUAL', om_jarang),
    ('OBAT DISPLAY',       om_display),
    ('OBAT MINUM',         om_minum),
]

all_results = {}
not_in_prod = []

print("=" * 80)
print("MAPPING REPORT  (source of truth: so_template_sample.xlsx)")
print("=" * 80)

for gname, glist in groups:
    all_results[gname] = []
    print(f"\n--- {gname} ({len(glist)} items) ---")
    for old_name in glist:
        code, in_prod, method = resolve(old_name)
        all_results[gname].append({'old': old_name, 'code': code, 'in_prod': in_prod, 'method': method})
        flag = 'OK' if in_prod else 'MISSING'
        print(f"  [{flag:7s}] {old_name!r:42s} => {code!r:35s}  ({method})")
        if not in_prod:
            not_in_prod.append((gname, old_name, code))

total = sum(len(v) for v in all_results.values())
found = sum(1 for v in all_results.values() for e in v if e['in_prod'])
print(f"\n\nSUMMARY: {found}/{total} mapped to production codes")
if not_in_prod:
    print(f"\nNot in production ({len(not_in_prod)}) — needs investigation:")
    for gname, old, code in not_in_prod:
        print(f"  [{gname}] {old!r}")

# ── Build upload Excel ──────────────────────────────────────────────────────
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'SO Template'

section_names = list(all_results.keys())
ws_out.append(section_names)

from openpyxl.styles import PatternFill
red_fill  = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
warn_fill = PatternFill(start_color='FFF3CC', end_color='FFF3CC', fill_type='solid')

max_len = max(len(v) for v in all_results.values())
for i in range(max_len):
    row = []
    for gname in section_names:
        entries = all_results[gname]
        if i < len(entries):
            e = entries[i]
            row.append(e['code'] if e['in_prod'] else f"[MISSING: {e['old']}]")
        else:
            row.append(None)
    ws_out.append(row)

for row in ws_out.iter_rows(min_row=2):
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            if cell.value.startswith('[MISSING:'):
                cell.fill = red_fill
            elif 'FUZZY' in (all_results.get(
                section_names[cell.column - 1], [{}] * (cell.row - 1)
            )[cell.row - 2] if cell.row - 2 < len(all_results.get(section_names[cell.column - 1], [])) else {}).get('method', ''):
                cell.fill = warn_fill

out_path = '../so_template_final.xlsx'
wb_out.save(out_path)
print(f"\nSaved: {out_path}")
print("Red = not in production | Yellow = fuzzy match (verify)")
